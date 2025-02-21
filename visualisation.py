# -*- coding: utf-8 -*-

"""
This file contains functions related to visualisation of the model results.

Functions:
    plot_train_losses
    plot_metrics
    plot_decoder_predictions
    one_sample_permutation_test
    compute_permutation
    significant_voxels
    two_samples_permutation_test
    diff_means
    print_dict_tree
"""

from imports import np, torch, F, plt, sns, explained_variance_score, stats, Pool, perm_test, nib, ssim
from dataset import normalize
import pandas as pd
from openpyxl import load_workbook
import matplotlib.colors as mcolors
import nibabel as nib
from nilearn import datasets, plotting, image
from scipy.ndimage import rotate

### TRAINING VISUALISATION ###


def plot_train_losses(history, start_epoch):
    """
    Plot the training losses and other metrics from the training history.

    Parameters:
        history (dict): Contains the training history data including loss and other metrics.
        start_epoch (int): The starting epoch number for plotting.

    Output:
        Displays plots for each metric in the history.
    """
    x = np.arange(len(history['total_loss'])) + start_epoch 
    
    plt.plot(x, history['total_loss'])
    plt.title(history['metrics_names'][-1])
    plt.xlabel('epoch')
    plt.ylabel('value')
    plt.show()

    for i in range(len(history['metrics_names']) - 1):
        plt.plot(x, history['other_metrics'][:, i])
        plt.title(history['metrics_names'][i])
        plt.xlabel('epoch')
        plt.ylabel('value')
        plt.show()


### TESTING VISUALISATION ###

def z_score(data):
    return (data - np.mean(data, axis=0)) / np.std(data, axis=0)


def plot_metrics(label, prediction, movie, plot_TR = False, performance_dict = None, display_plots=True):
    """
    Plot various statistical metrics and distributions for the prediction quality of a model.

    Parameters:
        label (np.array): Ground truth values, shape (n_samples, n_features).
        prediction (np.array): Model predictions, shape (n_samples, n_features).
        movie (str): Name of the dataset or context for the plot titles.
        plot_TR (bool, optional): If true, also plots time-resolved (TR) metrics. Default is False.

    Output:
        Visualizations of correlation, explained variance, and other statistical measures between label and prediction.
    """

    if movie == 'all':
        all_labels, all_predictions = label.copy(), prediction.copy()
        label, prediction = [], []
        for key in all_labels.keys():
            all_predictions[key] = z_score(all_predictions[key])
            all_labels[key] = z_score(all_labels[key])
            prediction.append(all_predictions[key])
            label.append(all_labels[key])
        prediction = np.concatenate(prediction, axis=0)
        label = np.concatenate(label, axis=0)
    else: 
        label = z_score(label)
        prediction = z_score(prediction)
    
    TR, voxels = prediction.shape

    correlations_voxel = np.array([
        stats.pearsonr(label[:, i], prediction[:, i])[0] 
        for i in range(voxels) 
        if np.std(label[:, i]) != 0 and np.std(prediction[:, i]) != 0
    ])
    mean_correlation_voxel = np.mean(correlations_voxel)
    median_correlation_voxel = np.median(correlations_voxel)
    std_correlation_voxel = np.std(correlations_voxel)

    best_voxel = np.argmax(correlations_voxel)
    worst_voxel = np.argmin(correlations_voxel)

    mse_voxel = np.array([
        F.mse_loss(torch.from_numpy(label[:, i]).unsqueeze(0), torch.from_numpy(prediction[:, i]).unsqueeze(0), reduction='mean').item()
        for i in range(voxels)
        if np.std(label[:, i]) != 0  
    ])
    mean_mse_voxel = np.mean(mse_voxel)
    median_mse_voxel = np.median(mse_voxel)
    std_mse_voxel = np.std(mse_voxel)

    if display_plots:
        print(f'### {movie}: ({TR}, {voxels}) ###')
    
    if movie == 'all':
        all_correlation = {'all': correlations_voxel}
        all_mse = {'all': mse_voxel}
        for key in all_labels.keys():
            voxels = all_labels[key].shape[1]
            corr_movie = np.array([
                stats.pearsonr(all_labels[key][:, i], all_predictions[key][:, i])[0] 
                for i in range(voxels) 
                if np.std(all_labels[key][:, i]) != 0 and np.std(all_predictions[key][:, i]) != 0
            ])
            mse_movie = np.array([
                F.mse_loss(torch.from_numpy(all_labels[key][:, i]).unsqueeze(0), torch.from_numpy(all_predictions[key][:, i]).unsqueeze(0), reduction='mean').item()
                for i in range(voxels)
                if np.std(all_labels[key][:, i]) != 0  
            ])
            all_correlation[key] = corr_movie
            all_mse[key] = mse_movie
        performance_dict['mean_corr_E'], performance_dict['median_corr_E'] = plot_scatter_metrics(all_correlation, 'Correlations', display_plots)
        performance_dict['mean_mse_E'], performance_dict['median_mse_E'] = plot_scatter_metrics(all_mse, 'MSE', display_plots)

    if display_plots:
        plt.figure(figsize=(17, 3))
        plt.subplot(121)
        plt.hist(correlations_voxel, bins=30, alpha=0.5, color='green', density=True, label='Histogram')
        sns.kdeplot(correlations_voxel, fill=True, color='g', label='KDE')
        plt.axvline(x=median_correlation_voxel, color='r', linestyle=':', linewidth=1, label=f'Median: {median_correlation_voxel:.3f}')
        plt.axvline(x=mean_correlation_voxel, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_correlation_voxel:.3f}')
        plt.axvspan(mean_correlation_voxel - std_correlation_voxel, mean_correlation_voxel + std_correlation_voxel, color='red', alpha=0.05, label=f'std: {std_correlation_voxel:.3f}')
        plt.xlabel('Correlation Coefficient')
        plt.ylabel('Density')
        plt.title(f'Distribution of Correlations per Voxel (N = {voxels})')
        plt.legend() 
        plt.subplot(122)
        plt.hist(mse_voxel, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(mse_voxel, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_voxel, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_voxel:.3f}')
        plt.axvline(x=mean_mse_voxel, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_voxel:.3f}')
        plt.axvspan(mean_mse_voxel - std_mse_voxel, mean_mse_voxel + std_mse_voxel, color='red', alpha=0.05, label=f'std: {std_mse_voxel:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE per Voxel (N = {voxels})')
        plt.legend() 
        plt.show()
    
        plt.figure(figsize=(22, 7))
        plt.subplot(2, 1, 1)
        plt.title(f'Best Voxel (Corr({best_voxel}) = {np.max(correlations_voxel):.3f})')
        plt.plot(label[:,best_voxel], label = 'ground truth')
        plt.plot(prediction[:,best_voxel], label = 'prediction')
        plt.legend()
        plt.subplot(2, 1, 2)
        plt.title(f'Worst Voxel (Corr({worst_voxel}) = {np.min(correlations_voxel):.3f})')
        plt.plot(label[:,worst_voxel], label = 'ground truth')
        plt.plot(prediction[:,worst_voxel], label = 'prediction')
        plt.legend()
        plt.show()

    if plot_TR and display_plots:
    
        correlations_TR = np.array([
            stats.pearsonr(label[i, :], prediction[i, :])[0] 
            for i in range(TR) 
            if np.std(label[i, :]) != 0 and np.std(prediction[i, :]) != 0
        ])
        mean_correlation_TR = np.mean(correlations_TR)
        median_correlation_TR = np.median(correlations_TR)
        std_correlation_TR = np.std(correlations_TR)
    
        best_TR = np.argmax(correlations_TR)
        worst_TR = np.argmin(correlations_TR)
    
        mse_TR = np.array([
            F.mse_loss(torch.from_numpy(label[i, :]).unsqueeze(0), torch.from_numpy(prediction[i, :]).unsqueeze(0), reduction='mean').item()
            for i in range(TR)
            if np.std(label[i, :]) != 0  
        ])
        mean_mse_TR = np.mean(mse_TR)
        median_mse_TR = np.median(mse_TR)
        std_mse_TR = np.std(mse_TR)
    
        plt.figure(figsize=(17, 3))
        plt.subplot(121)
        plt.hist(correlations_TR, bins=30, alpha=0.5, color='green', density=True, label='Histogram')
        sns.kdeplot(correlations_TR, fill=True, color='g', label='KDE')
        plt.axvline(x=median_correlation_TR, color='r', linestyle=':', linewidth=1, label=f'Median: {median_correlation_TR:.3f}')
        plt.axvline(x=mean_correlation_TR, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_correlation_TR:.3f}')
        plt.axvspan(mean_correlation_TR - std_correlation_TR, mean_correlation_TR + std_correlation_TR, color='red', alpha=0.05, label=f'std: {std_correlation_TR:.3f}')
        plt.xlabel('Correlation Coefficient')
        plt.ylabel('Density')
        plt.title(f'Distribution of Correlations per TR (N = {TR})')
        plt.legend() 
        plt.subplot(122)
        plt.hist(mse_TR, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(mse_TR, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_TR, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_TR:.3f}')
        plt.axvline(x=mean_mse_TR, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_TR:.3f}')
        plt.axvspan(mean_mse_TR - std_mse_TR, mean_mse_TR + std_mse_TR, color='red', alpha=0.05, label=f'std: {std_mse_TR:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE per TR (N = {TR})')
        plt.legend() 
        plt.show()
    
        plt.figure(figsize=(22, 7))
        plt.subplot(2, 1, 1)
        plt.title(f'Best TR (Corr({best_TR}) = {np.max(correlations_TR):.3f})')
        plt.plot(label[best_TR, :], label = 'ground truth')
        plt.plot(prediction[best_TR, :], label = 'prediction')
        plt.legend()
        plt.subplot(2, 1, 2)
        plt.title(f'Worst TR (Corr({worst_TR}) = {np.min(correlations_TR):.3f})')
        plt.plot(label[worst_TR, :], label = 'ground truth')
        plt.plot(prediction[worst_TR, :], label = 'prediction')
        plt.legend()
        plt.show()

    if performance_dict:
        return performance_dict

def plot_scatter_metrics(metrics, name, display_plots=True):
    colors = {'Correlations': 'green',
              'MSE': 'red',
              'SSIM': 'blue'}
    
    # Convert dictionary to DataFrame
    df = pd.DataFrame([
        {'Movie': movie, 'Value': value}
        for movie, values in metrics.items()
        for value in values
    ])

    # Retrieve statistics for the 'all' movie
    first_values = np.array(metrics['all'])
    median_value = np.median(first_values)
    mean_value = np.mean(first_values)
    std_dev_value = np.std(first_values)

    if display_plots:
        movie_abreviations = {'Sintel': '1-St',
                          'Payload': '2-Pld',
                          'TearsOfSteel': '3-TOS',
                          'Superhero': '4-Sh',
                          'BigBuckBunny': '5-BBB',
                          'FirstBite': '6-FB',
                          'BetweenViewings': '7-BW',
                          'AfterTheRain': '8-ATR',
                          'TheSecretNumber': '9-TSN',
                          'Chatter': '10-Cht',
                          'Spaceman': '11-Spm',
                          'LessonLearned': '12-LL',
                          'YouAgain': '*13-YA',
                          'ToClaireFromSonny': '14-TCFS'}
        for key, value in movie_abreviations.items():
            df['Movie'] = df['Movie'].replace(key, value)
    
        
        categories = df['Movie'].unique()
        
        fig, ax = plt.subplots(figsize=(15, 9))

        # Create violin plot for the categories
        sns.violinplot(x='Movie', y='Value', data=df,
                       ax=ax, density_norm='width', fill=True,
                       linecolor='k', linewidth=2,
                       color='k', alpha=0.07,
                       inner_kws=dict(box_width=8, whis_width=2, color="0.7"))

        jitter, size = 0.08, 1
        if first_values.shape[0] > 3000:
            jitter, size = 0.04, 0.5
        # Plot scatter points for each category
        for idx, category in enumerate(categories):
            y = df[df['Movie'] == category]['Value']
            x = np.random.normal(idx, jitter, size=len(y))  # Add jitter
    
            # Assign color based on the category
            if category == 'all':
                color = colors.get(name, '0.5')  # Use the color corresponding to 'name'
            else:
                color = 'k'  # Default grey color for other categories
            
            ax.scatter(x, y, alpha=0.5, s=size, color=color)

        # Statistical lines for the 'all' movie
        ax.axhline(mean_value, color='r', linestyle='--', linewidth=3, label=f'Mean: {mean_value:.3f}')
        ax.axhline(median_value, color='g', linestyle=':', linewidth=3,label=f'Median: {median_value:.3f}')
        ax.axhspan(mean_value - std_dev_value, mean_value + std_dev_value, color='red', alpha=0.05, label=f'Std: {std_dev_value:.3f}')
        
        # Set x-ticks and labels
        ax.set_xticks(np.arange(len(categories)))
        ax.set_xticklabels(categories, rotation=30, fontsize=25, ha='center')
        
        # Customize plot
        ax.set_title(f'{name} Distribution by Movie', fontsize=20, y=1.1)
        ax.set_ylabel(name, fontsize=33)
        ax.set_xlabel('Movie', fontsize=33)
        ax.tick_params(axis='y', labelsize=25, rotation=60)
        ax.legend(fontsize=29, loc='upper center', bbox_to_anchor=(0.5, 1.12),
              fancybox=True, shadow=True, ncol=3)
        ax.grid(color='grey', linestyle=':', linewidth=0.8)
        plt.tight_layout()
        plt.show()
        
    return np.round(mean_value, 3), np.round(median_value, 3)

# def plot_scatter_metrics(metrics, name, display_plots=True):
#     colors = {'Correlations': 'green',
#               'MSE': 'red',
#               'SSIM': 'blue'}
    
#     # Convert dictionary to DataFrame
#     df = pd.DataFrame([
#         {'Movie': movie, 'Value': value}
#         for movie, values in metrics.items()
#         for value in values
#     ])

#     # Retrieve statistics for the 'all' movie
#     first_values = np.array(metrics['all'])
#     median_value = np.median(first_values)
#     mean_value = np.mean(first_values)
#     std_dev_value = np.std(first_values)

#     #np.save(f'encoder_15364_2_{name}', first_values)

#     if display_plots:
#         movie_abreviations = {'Sintel': '1-St',
#                           'Payload': '2-Pld',
#                           'TearsOfSteel': '3-TOS',
#                           'Superhero': '4-Sh',
#                           'BigBuckBunny': '5-BBB',
#                           'FirstBite': '6-FB',
#                           'BetweenViewings': '7-BW',
#                           'AfterTheRain': '8-ATR',
#                           'TheSecretNumber': '9-TSN',
#                           'Chatter': '10-Cht',
#                           'Spaceman': '11-Spm',
#                           'LessonLearned': '12-LL',
#                           'YouAgain': '*13-YA',
#                           'ToClaireFromSonny': '14-TCFS'}
#         for key, value in movie_abreviations.items():
#             df['Movie'] = df['Movie'].replace(key, value)
    
        
#         categories = df['Movie'].unique()
#         positions = np.arange(len(categories))
        
#         # Prepare data for boxplot
#         box_data = [df[df['Movie'] == category]['Value'] for category in categories]
        
#         fig, ax = plt.subplots(figsize=(15, 10))
        
#         # Plot the boxplot shifted slightly to the left
#         box_positions = positions - 0.1
#         bp = ax.boxplot(box_data, positions=box_positions, widths=0.1, patch_artist=True, showfliers=False)
        
#         # Customize boxplot appearance
#         for box in bp['boxes']:
#             box.set(facecolor='none', linewidth=1)
#         for whisker in bp['whiskers']:
#             whisker.set(color='black', linewidth=1)
#         for cap in bp['caps']:
#             cap.set(color='black', linewidth=1)
#         for median in bp['medians']:
#             median.set(color='black', linewidth=1)
        
#         # Plot the scatter points shifted slightly to the right
#         strip_positions = positions + 0.2
#         for idx, category in enumerate(categories):
#             y = df[df['Movie'] == category]['Value']
#             x = np.random.normal(strip_positions[idx], 0.08, size=len(y))  # Add jitter
    
#             # Assign color based on the category
#             if category == 'all':
#                 color = colors.get(name, '0.5')  # Use the color corresponding to 'name'
#             else:
#                 color = 'k'  # Default grey color for other categories
            
#             ax.scatter(x, y, alpha=0.5, s=1, color=color)
        
#         # Set x-ticks and labels
#         ax.set_xticks(positions)
#         ax.set_xticklabels(categories, rotation=30, fontsize=25, ha='center')
        
#         # Statistical lines for the 'all' movie
#         ax.axhline(mean_value, color='r', linestyle='--', linewidth=3, label=f'Mean: {mean_value:.3f}')
#         ax.axhline(median_value, color='g', linestyle=':', linewidth=3,label=f'Median: {median_value:.3f}')
#         ax.axhspan(mean_value - std_dev_value, mean_value + std_dev_value, color='red', alpha=0.08, label=f'Std: {std_dev_value:.3f}')
        
#         # Customize plot
#         ax.set_title(f'{name} Distribution by Movie', fontsize=20, y=1.1)
#         ax.set_ylabel(name, fontsize=24)
#         ax.set_xlabel('Movie', fontsize=24)
#         ax.tick_params(axis='y', labelsize=25, rotation=60)
#         ax.legend(fontsize=26, loc='upper center', bbox_to_anchor=(0.5, 1.09),
#               fancybox=True, shadow=True, ncol=3)
#         ax.grid(color='grey', linestyle=':', linewidth=0.8)
#         plt.tight_layout()
#         plt.show()

#     return np.round(mean_value, 3), np.round(median_value, 3)

def plot_decoder_predictions(predictions, videos, performance_dict = None, display_plots=True):
    """
    Display comparison plots between original videos and their corresponding predictions.

    Parameters:
        predictions (dict): A dictionary containing predicted frames, indexed by video names.
        videos (dict): A dictionary containing original video frames, indexed by video names.

    This function iterates through each video key, retrieves frames from both the original and predicted data,
    and plots them side by side for visual comparison. It supports different intervals for videos with a large
    number of frames, adjusting the subplot layout accordingly.
    """

    total_ssim, total_mse = [], []
    all_ssim = {'all': None}
    all_mse = {'all': None}
    for key in videos.keys():
        prediction = predictions[key]
        video = videos[key][..., 15]
        N = video.shape[0]

        ssim_frames = np.array([
            ssim(torch.from_numpy(video[i]).unsqueeze(0), torch.from_numpy(prediction[i]).unsqueeze(0), data_range=1, size_average=True).item()
            for i in range(N) 
        ])
        mean_ssim_frames = np.mean(ssim_frames)
        median_ssim_frames = np.median(ssim_frames)
        std_ssim_frames = np.std(ssim_frames)
        all_ssim[key] = ssim_frames

        mse_frames = np.array([
            F.mse_loss(normalize(torch.from_numpy(video[i])), normalize(torch.from_numpy(prediction[i]))).numpy()
            for i in range(N) 
        ])
        mean_mse_frames = np.mean(mse_frames)
        median_mse_frames = np.median(mse_frames)
        std_mse_frames = np.std(mse_frames)
        all_mse[key] = mse_frames

        total_ssim = np.concatenate((total_ssim, ssim_frames))
        total_mse = np.concatenate((total_mse, mse_frames))

        if display_plots:
            print(f'### {key} ({N}) ###')
            plt.figure(figsize=(17, 3))
            plt.subplot(121)
            plt.hist(ssim_frames, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
            sns.kdeplot(ssim_frames, fill=True, color='b', label='KDE')
            plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
            plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
            plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
            plt.xlabel('SSIM')
            plt.ylabel('Density')
            plt.title(f'Distribution of Structural Similarity per Frame (N = {N})')
            plt.legend() 
            plt.subplot(122)
            plt.hist(mse_frames, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
            sns.kdeplot(mse_frames, fill=True, color='r', label='KDE')
            plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
            plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
            plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
            plt.xlabel('MSE')
            plt.ylabel('Density')
            plt.title(f'Distribution of MSE per Frame (N = {N})')
            plt.legend() 
            plt.show()
            
            if prediction.shape[0] > 400:
                indices = np.arange(0, prediction.shape[0], 15)
            else:
                indices = np.arange(0, prediction.shape[0], 3)
            print(f'Display {len(indices)} frames:', indices)
            col = int(indices.shape[0] / 3) + 1
            k = 0
    
            if len(indices) < 20:
                plt.figure(figsize=(8, 10))
            elif len(indices) >= 20 and len(indices) < 40:
                plt.figure(figsize=(10, 15))
            else:
                plt.figure(figsize=(15, 20))
                
            for i in indices:
                k += 1
                plt.subplot(col,6,k)
                plt.imshow(np.transpose(normalize(video[i]), (1, 2, 0)))
                plt.axis('off')
                k += 1
                plt.subplot(col,6,k)
                plt.imshow(np.transpose(normalize(prediction[i]), (1, 2, 0)))
                plt.axis('off')
            plt.show()
            del prediction, video, indices

    mean_ssim_frames = np.mean(total_ssim)
    median_ssim_frames = np.median(total_ssim)
    std_ssim_frames = np.std(total_ssim)

    mean_mse_frames = np.mean(total_mse)
    median_mse_frames = np.median(total_mse)
    std_mse_frames = np.std(total_mse)

    all_ssim['all'] = total_ssim
    all_mse['all'] = total_mse

    if display_plots:
        print(f'### all ({total_ssim.shape[0]}) ###')
        plt.figure(figsize=(17, 3))
        plt.subplot(121)
        plt.hist(total_ssim, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_ssim, fill=True, color='b', label='KDE')
        plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
        plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
        plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
        plt.xlabel('SSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Structural Similarity per Frame (N = {total_ssim.shape[0]})')
        plt.legend() 
        plt.subplot(122)
        plt.hist(total_mse, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(total_mse, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
        plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
        plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE per Frame (N = {total_mse.shape[0]})')
        plt.legend() 
        plt.show()

    if performance_dict:
        performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_ssim, 'SSIM', display_plots)
        performance_dict['mean_mse_D'], performance_dict['median_mse_D'] = plot_scatter_metrics(all_mse, 'MSE', display_plots)
        return performance_dict

def plot_saliency_distribution(saliency_values):

    print('### Saliency Mapping to Anatomical Regions ###')
    percentile = 20
    threshold = np.percentile(saliency_values, q=100-percentile)
    print(f'Top {percentile}% saliency values are above {np.round(threshold,3)}')

    # Define the dictionary with indexes as keys and region names as values
    brain_regions = {
        4112: 'ParaHippocampal_R',
        5001: 'Calcarine_L',
        5002: 'Calcarine_R',
        5011: 'Cuneus_L',
        5012: 'Cuneus_R',
        5021: 'Lingual_L',
        5022: 'Lingual_R',
        5101: 'Occipital_Sup_L',
        5102: 'Occipital_Sup_R',
        5201: 'Occipital_Mid_L',
        5202: 'Occipital_Mid_R',
        5301: 'Occipital_Inf_L',
        5302: 'Occipital_Inf_R',
        5401: 'Fusiform_L',
        5402: 'Fusiform_R',
        6101: 'Parietal_Sup_L',
        6102: 'Parietal_Sup_R',
        6201: 'Parietal_Inf_L',
        6222: 'Angular_R',
        6301: 'Precuneus_L',
        6302: 'Precuneus_R',
        8201: 'Temporal_Mid_L',
        8202: 'Temporal_Mid_R',
        8301: 'Temporal_Inf_L',
        8302: 'Temporal_Inf_R',
        9001: 'Cerebelum_Crus1_L',
        9002: 'Cerebelum_Crus1_R',
        9031: 'Cerebelum_4_5_L',
        9032: 'Cerebelum_4_5_R',
        9041: 'Cerebelum_6_L',
        9042: 'Cerebelum_6_R',
        9130: 'Vermis_6'
    }
    
    main_rois = {'Calcarine': 1,
            'Lingual': 2,
            'Occipital_Sup': 3,
            'Occipital_Mid': 4,
            'Occipital_Inf': 5,
            'Fusiform': 6,
            'Temporal_Mid': 7,
            'Temporal_Inf': 8,
            'Other': 9}

    colors = [
        "black",             # Black
        "mediumseagreen",
        "hotpink",
        "lightskyblue",
        "#0072B2",           # Medium blue
        "navy",
        "darkgreen",
        "red",
        "coral",
        "dimgray"
    ]


    brain_path = '/media/miplab-nas-shadow/Data2/Movies_Emo/Preprocessed_data/sub-S01/ses-1/pp_sub-S01_ses-1_YouAgain.feat/filtered_func_data_res_MNI.nii'
    brain_img = nib.load(brain_path)
    template_img = datasets.load_mni152_template()
    template_img = image.resample_to_img(source_img=template_img,
                                        target_img=brain_img,
                                        interpolation='nearest')
    template_data = np.asanyarray(template_img.dataobj)
    
    atlas_path = './aal_SPM12/aal/atlas/AAL.nii'
    atlas_img = nib.load(atlas_path)
    atlas_data = np.asanyarray(atlas_img.dataobj)
    
    mask3d_path_s = f'mask_schaefer1000_{saliency_values.shape[0]}.npy'
    mask3d_path_a = f'mask_schaefer1000_4609.npy'
    #mask3d_path_a = f'mask_schaefer1000_15364.npy'
    mask_3d_s = np.load(mask3d_path_s, mmap_mode = 'r')
    mask_3d_a = np.load(mask3d_path_a, mmap_mode = 'r')
    masked_anatomical = atlas_data * mask_3d_a
    mask_2d = mask_3d_s.reshape(-1,) 
    indices = np.where(mask_2d == 1)[0]

    # Step 1: Create a zero-filled array with the same shape as the original 3D data
    reconstructed_3d = np.zeros(mask_3d_s.shape, dtype=np.float32)
    # Step 2: Flatten the zero-filled array (this step is optional and for clarity)
    reconstructed_flat = reconstructed_3d.flatten()
    # Step 3: Use the stored indices to place the masked_fMRI values into the flattened array
    reconstructed_flat[indices] = saliency_values
    # Step 4: Reshape back to the original 3D shape
    masked_saliency = reconstructed_flat.reshape(mask_3d_s.shape)

    #np.save('saliency_map', masked_saliency)

    # Step 1: Create a zero-filled array with the same shape as the original 3D data
    reconstructed_3d_n = np.zeros(mask_3d_s.shape, dtype=np.float32)
    # Step 2: Flatten the zero-filled array (this step is optional and for clarity)
    reconstructed_flat_n = reconstructed_3d_n.flatten()
    # Step 3: Use the stored indices to place the masked_fMRI values into the flattened array
    saliency_values_n = normalize(saliency_values)
    reconstructed_flat_n[indices] = saliency_values_n
    # Step 4: Reshape back to the original 3D shape
    masked_saliency_n = reconstructed_flat_n.reshape(mask_3d_s.shape)

    labels, counts = np.unique(masked_anatomical, return_counts=True)

    # Create a dictionary to hold the voxel counts for each category
    voxel_counts = {key: 0 for key in main_rois}
    total_voxels = 0

    best_saliency_counts = {key: 0 for key in main_rois}
    total_best_saliency = 0
    
    # Create a dictionary to hold saliency values for each category
    saliency_dict = {key: [] for key in main_rois}

    # Assign counts to categories and collect saliency values
    for label, count in zip(labels, counts):
        if label == 0:
            continue  # Skip background or unlabeled
        roi_name = brain_regions.get(label, 'Other')

        # Determine the category for each roi_name
        matched = False
        for roi in main_rois.keys():
            if roi in roi_name:  # Check for partial matches
                # Get indices where mask == label
                indices = np.where(masked_anatomical == label)
                # Get saliency values at those indices
                saliency_vals = masked_saliency[indices]
                # Append to saliency_dict[roi]
                saliency_dict[roi].extend(saliency_vals.tolist())
                
                voxel_counts[roi] += count
                matched = True
                masked_anatomical[masked_anatomical == label] = main_rois[roi]

                best_saliency_counts[roi] += (saliency_vals > threshold).sum()
                break

        if not matched:
            indices = np.where(masked_anatomical == label)
            saliency_vals = masked_saliency[indices]
            saliency_dict['Other'].extend(saliency_vals.tolist())

            voxel_counts['Other'] += count
            masked_anatomical[masked_anatomical == label] = main_rois['Other']
            best_saliency_counts['Other'] += (saliency_vals > threshold).sum()

        total_voxels += count
        #print(f"({int(label)}) | {roi_name}: {count}")

    #print(f"Total: {total_voxels} voxels")

    # Convert counts to percentages
    roi_percentages = {k: (v / total_voxels * 100) for k, v in voxel_counts.items()}
    best_saliency_percentages = {k: (v / (saliency_values.shape[0] * percentile) * 10000) for k, v in best_saliency_counts.items()}

    print('Region' + ' '*10, '| Number of voxels    | Number of best saliencies')
    for k in voxel_counts.keys():
        print(k, ' '*(15-len(k)), f'| {voxel_counts[k]}', ' '*(18-len(str(voxel_counts[k]))), f'| {best_saliency_counts[k]}')

    print()
    
    print('Region' + ' '*10, '| Total % of voxels   | % of best saliencies')
    for k in voxel_counts.keys():
        print(k, ' '*(15-len(k)), f'| {np.round(roi_percentages[k],1)}%', ' '*(17-len(str(np.round(roi_percentages[k],1)))), f'| {np.round(best_saliency_percentages[k],1)}%')

    df = pd.DataFrame([
        {'Region': region, 'Value': value}
        for region, values in saliency_dict.items()
        for value in values
    ])
    
    categories = df['Region'].unique()
    positions = np.arange(len(categories))
    
    # Prepare data for boxplot
    box_data = [df[df['Region'] == category]['Value'] for category in categories]
    
    fig, ax = plt.subplots(figsize=(10, 5))
    
    # Plot the boxplot shifted slightly to the left
    box_positions = positions - 0.1
    bp = ax.boxplot(box_data, positions=box_positions, widths=0.1, patch_artist=True, showfliers=False)
    
    # Customize boxplot appearance
    for box in bp['boxes']:
        box.set(facecolor='none', linewidth=1)
    for whisker in bp['whiskers']:
        whisker.set(color='black', linewidth=1)
    for cap in bp['caps']:
        cap.set(color='black', linewidth=1)
    for median in bp['medians']:
        median.set(color='black', linewidth=1)
    
    # Plot the scatter points shifted slightly to the right
    strip_positions = positions + 0.1
    for idx, category in enumerate(categories):
        y = df[df['Region'] == category]['Value']
        x = np.random.normal(strip_positions[idx], 0.05, size=len(y))  # Add jitter
        color = colors[idx+1]  # Default grey color for other categories
        
        ax.scatter(x, y, alpha=1, s=1, color=color)

    ax.axhline(threshold, color='r', linestyle='--', linewidth=3, label=f'Top {percentile}% above {np.round(threshold,3)}')

    x_labels = [f'{category}\n({roi_percentages[category]:.1f}%)' for category in categories]
    # Set x-ticks and labels
    ax.set_xticks(positions)
    ax.set_xticklabels(x_labels, rotation=40, fontsize=13)
    # Customize plot
    ax.set_title(f'Saliency Distribution by Region', fontsize=15)
    ax.set_ylabel('Absolute Saliency', fontsize=15)
    ax.tick_params(axis='y', labelsize=15)
    ax.grid(color='grey', linestyle=':', linewidth=0.4)
    plt.tight_layout()
    plt.legend()
    plt.show()

    n_labels = len(main_rois.keys())
    # Create a custom colormap with a color for each ROI plus one for 'Other'
    custom_cmap = mcolors.ListedColormap(colors)
    custom_cmap.set_under(color='black', alpha=0)
    
    # Custom normalization function
    class CustomNormalize(mcolors.Normalize):
        def __init__(self, vmin=None, vmax=None, midpoint=0.4, clip=False):
            self.midpoint = midpoint
            super().__init__(vmin, vmax, clip)
        
        def __call__(self, value, clip=None):
            x, y = [self.vmin, self.midpoint, self.vmax], [0, 0.5, 1]
            return np.ma.masked_array(np.interp(value, x, y))

    n_colors = 256
    color_array = plt.cm.hot(np.linspace(0, 1, n_colors))  # Reds colormap
    color_array[0] = (0, 0, 0, 0)  # Set 0 to transparent black

    colors = [(0, 0, 0, 0), (1, 1, 0, 1)]  # Transparent black, Yellow
    custom_cmap_s = mcolors.ListedColormap(colors)

    # Example: Creating a norm to map values (0 maps to transparent, >0 to yellow)
    bounds = [0, 0.1, 1]  # Define boundaries for 0 and non-zero
    norm = mcolors.BoundaryNorm(bounds, custom_cmap_s.N)

    masked_saliency[masked_saliency < threshold] = 0

    color_bar = True
    for i in range(0, 90, 10):

        plt.figure(figsize=(15, 6))
        
        # Sagittal slices
        ax1 = plt.subplot(1, 3, 1)
        ax1.axis('off')
        plt.imshow(rotate(template_data[i, :, :], 90, reshape=True), cmap='gray')
        img1a = plt.imshow(rotate(masked_anatomical[i, :, :], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[i, :, :], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.title(f'Sagittal slice at x={i}')
        
        # Coronal slices
        ax2 = plt.subplot(1, 3, 2)
        ax2.axis('off')
        plt.imshow(rotate(template_data[:, i, :], 90, reshape=True), cmap='gray')
        img2a = plt.imshow(rotate(masked_anatomical[:, i, :], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[:, i, :], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.title(f'Coronal slice at y={i}')
        
        # Axial slices
        ax3 = plt.subplot(1, 3, 3)
        ax3.axis('off')
        plt.imshow(rotate(template_data[:, :, i], 90, reshape=True), cmap='gray')
        img3a = plt.imshow(rotate(masked_anatomical[:, :, i], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[:, :, i], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.title(f'Axial slice at z={i}')
        
        plt.suptitle('Visual Mask', fontsize=16)
        plt.show()

    
    best_slices = [27, 31, 33]
    for i in best_slices:        # for i in range(90):

        plt.figure(figsize=(6, 6))
        plt.axis('off')
        plt.imshow(rotate(template_data[i, :, :], 90, reshape=True), cmap='gray')
        img1a = plt.imshow(rotate(masked_anatomical[i, :, :], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[i, :, :], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.show()
        print(f'Sagittal slice at x={i}')

        plt.figure(figsize=(6, 6))
        plt.axis('off')
        plt.imshow(rotate(template_data[:, i, :], 90, reshape=True), cmap='gray')
        img2a = plt.imshow(rotate(masked_anatomical[:, i, :], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[:, i, :], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.show()
        print(f'Coronal slice at y={i}')

        plt.figure(figsize=(6, 6))
        plt.axis('off')
        plt.imshow(rotate(template_data[:, :, i], 90, reshape=True), cmap='gray')
        img3a = plt.imshow(rotate(masked_anatomical[:, :, i], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[:, :, i], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.show()
        print(f'Axial slice at z={i}')

# def plot_saliency_distribution_old(saliency_values):

#     print('### Saliency Mapping to Anatomical Regions ###')

#     # Define the dictionary with indexes as keys and region names as values
#     brain_regions = {
#         4112: 'ParaHippocampal_R',
#         5001: 'Calcarine_L',
#         5002: 'Calcarine_R',
#         5011: 'Cuneus_L',
#         5012: 'Cuneus_R',
#         5021: 'Lingual_L',
#         5022: 'Lingual_R',
#         5101: 'Occipital_Sup_L',
#         5102: 'Occipital_Sup_R',
#         5201: 'Occipital_Mid_L',
#         5202: 'Occipital_Mid_R',
#         5301: 'Occipital_Inf_L',
#         5302: 'Occipital_Inf_R',
#         5401: 'Fusiform_L',
#         5402: 'Fusiform_R',
#         6101: 'Parietal_Sup_L',
#         6102: 'Parietal_Sup_R',
#         6201: 'Parietal_Inf_L',
#         6222: 'Angular_R',
#         6301: 'Precuneus_L',
#         6302: 'Precuneus_R',
#         8201: 'Temporal_Mid_L',
#         8202: 'Temporal_Mid_R',
#         8301: 'Temporal_Inf_L',
#         8302: 'Temporal_Inf_R',
#         9001: 'Cerebelum_Crus1_L',
#         9002: 'Cerebelum_Crus1_R',
#         9031: 'Cerebelum_4_5_L',
#         9032: 'Cerebelum_4_5_R',
#         9041: 'Cerebelum_6_L',
#         9042: 'Cerebelum_6_R',
#         9130: 'Vermis_6'
#     }
    
#     main_rois = {'Calcarine': 1,
#             'Lingual': 2,
#             'Occipital_Sup': 3,
#             'Occipital_Mid': 4,
#             'Occipital_Inf': 5,
#             'Fusiform': 6,
#             'Temporal_Mid': 7,
#             'Temporal_Inf': 8,
#             'Other': 9}

#     colors = [
#         "black",
#         "#ff7f0e",  # Safety orange.
#         "salmon",
#         "#1f77b4",  # Muted blue.
#         "#17becf",  # Blue-teal.
#         "#9467bd",  # Muted purple.
#         "#d62728",  # Brick red.
#         "#2ca02c",  # Cooked asparagus green.
#         "#bcbd22",  # Curry yellow-green.
#         "#8c564b",  # Chestnut brown.
#     ]

#     brain_path = '/media/miplab-nas-shadow/Data2/Movies_Emo/Preprocessed_data/sub-S01/ses-1/pp_sub-S01_ses-1_YouAgain.feat/filtered_func_data_res_MNI.nii'
#     brain_img = nib.load(brain_path)
#     template_img = datasets.load_mni152_template()
#     template_img = image.resample_to_img(source_img=template_img,
#                                         target_img=brain_img,
#                                         interpolation='nearest')
#     template_data = np.asanyarray(template_img.dataobj)
    
#     atlas_path = './aal_SPM12/aal/atlas/AAL.nii'
#     atlas_img = nib.load(atlas_path)
#     atlas_data = np.asanyarray(atlas_img.dataobj)
    
#     mask3d_path = f'mask_schaefer1000_{saliency_values.shape[0]}.npy'
#     mask_3d = np.load(mask3d_path, mmap_mode = 'r')
#     masked_anatomical = atlas_data * mask_3d
#     mask_2d = mask_3d.reshape(-1,) 
#     indices = np.where(mask_2d == 1)[0]

#     # Step 1: Create a zero-filled array with the same shape as the original 3D data
#     reconstructed_3d = np.zeros(mask_3d.shape, dtype=np.float32)
#     # Step 2: Flatten the zero-filled array (this step is optional and for clarity)
#     reconstructed_flat = reconstructed_3d.flatten()
#     # Step 3: Use the stored indices to place the masked_fMRI values into the flattened array
#     reconstructed_flat[indices] = saliency_values
#     # Step 4: Reshape back to the original 3D shape
#     masked_saliency = reconstructed_flat.reshape(mask_3d.shape)

#     #np.save('saliency_map', masked_saliency)

#     # Step 1: Create a zero-filled array with the same shape as the original 3D data
#     reconstructed_3d_n = np.zeros(mask_3d.shape, dtype=np.float32)
#     # Step 2: Flatten the zero-filled array (this step is optional and for clarity)
#     reconstructed_flat_n = reconstructed_3d_n.flatten()
#     # Step 3: Use the stored indices to place the masked_fMRI values into the flattened array
#     reconstructed_flat_n[indices] = normalize(saliency_values)
#     # Step 4: Reshape back to the original 3D shape
#     masked_saliency_n = reconstructed_flat_n.reshape(mask_3d.shape)

#     labels, counts = np.unique(masked_anatomical, return_counts=True)

#     # Create a dictionary to hold the voxel counts for each category
#     voxel_counts = {key: 0 for key in main_rois}
#     total_voxels = 0

#     # Create a dictionary to hold saliency values for each category
#     saliency_dict = {key: [] for key in main_rois}

#     # Assign counts to categories and collect saliency values
#     for label, count in zip(labels, counts):
#         if label == 0:
#             continue  # Skip background or unlabeled
#         roi_name = brain_regions.get(label, 'Other')

#         # Determine the category for each roi_name
#         matched = False
#         for roi in main_rois.keys():
#             if roi in roi_name:  # Check for partial matches
#                 # Get indices where mask == label
#                 indices = np.where(masked_anatomical == label)
#                 # Get saliency values at those indices
#                 saliency_vals = masked_saliency[indices]
#                 # Append to saliency_dict[roi]
#                 saliency_dict[roi].extend(saliency_vals.tolist())
                
#                 voxel_counts[roi] += count
#                 matched = True
#                 masked_anatomical[masked_anatomical == label] = main_rois[roi]
#                 break

#         if not matched:
#             indices = np.where(masked_anatomical == label)
#             saliency_vals = masked_saliency[indices]
#             saliency_dict['Other'].extend(saliency_vals.tolist())

#             voxel_counts['Other'] += count
#             masked_anatomical[masked_anatomical == label] = main_rois['Other']

#         total_voxels += count
#         print(f"({int(label)}) | {roi_name}: {count}")

#     print(f"Total: {total_voxels} voxels")

#     # Convert counts to percentages
#     roi_percentages = {k: (v / total_voxels * 100) for k, v in voxel_counts.items()}

#     df = pd.DataFrame([
#         {'Region': region, 'Value': value}
#         for region, values in saliency_dict.items()
#         for value in values
#     ])
    
#     categories = df['Region'].unique()
#     positions = np.arange(len(categories))
    
#     # Prepare data for boxplot
#     box_data = [df[df['Region'] == category]['Value'] for category in categories]
    
#     fig, ax = plt.subplots(figsize=(10, 5))
    
#     # Plot the boxplot shifted slightly to the left
#     box_positions = positions - 0.1
#     bp = ax.boxplot(box_data, positions=box_positions, widths=0.1, patch_artist=True, showfliers=False)
    
#     # Customize boxplot appearance
#     for box in bp['boxes']:
#         box.set(facecolor='none', linewidth=1)
#     for whisker in bp['whiskers']:
#         whisker.set(color='black', linewidth=1)
#     for cap in bp['caps']:
#         cap.set(color='black', linewidth=1)
#     for median in bp['medians']:
#         median.set(color='black', linewidth=1)
    
#     # Plot the scatter points shifted slightly to the right
#     strip_positions = positions + 0.1
#     for idx, category in enumerate(categories):
#         y = df[df['Region'] == category]['Value']
#         x = np.random.normal(strip_positions[idx], 0.05, size=len(y))  # Add jitter
#         color = colors[idx+1]  # Default grey color for other categories
        
#         ax.scatter(x, y, alpha=1, s=1, color=color)

#     x_labels = [f'{category}\n({roi_percentages[category]:.1f}%)' for category in categories]
#     # Set x-ticks and labels
#     ax.set_xticks(positions)
#     ax.set_xticklabels(x_labels, rotation=40, fontsize=13)
#     # Customize plot
#     ax.set_title(f'Saliency Distribution by Region', fontsize=15)
#     ax.set_ylabel('Absolute Saliency', fontsize=15)
#     ax.tick_params(axis='y', labelsize=15)
#     ax.grid(color='grey', linestyle=':', linewidth=0.4)
#     plt.tight_layout()
#     plt.show()

#     n_labels = len(main_rois.keys())
#     # Create a custom colormap with a color for each ROI plus one for 'Other'
#     custom_cmap = mcolors.ListedColormap(colors)
#     custom_cmap.set_under(color='black', alpha=0)
    
#     # Custom normalization function
#     class CustomNormalize(mcolors.Normalize):
#         def __init__(self, vmin=None, vmax=None, midpoint=0.4, clip=False):
#             self.midpoint = midpoint
#             super().__init__(vmin, vmax, clip)
        
#         def __call__(self, value, clip=None):
#             x, y = [self.vmin, self.midpoint, self.vmax], [0, 0.5, 1]
#             return np.ma.masked_array(np.interp(value, x, y))
    
#     # Define a custom colormap that emphasizes higher values (above 0.4)
#     n_colors = 256
#     color_array = plt.cm.Reds(np.linspace(0, 1, n_colors))  # Reds colormap
#     color_array[0] = (0, 0, 0, 0)  # Set 0 to transparent black
#     custom_cmap_s = mcolors.ListedColormap(color_array)

#     # Use custom normalization for better contrast
#     norm = CustomNormalize(vmin=0, vmax=1, midpoint=0.4)
    
#     color_bar = True
#     for i in range(10, 90, 10):

#         plt.figure(figsize=(15, 5))
        
#         # Sagittal slices
#         ax1 = plt.subplot(1, 3, 1)
#         ax1.axis('off')
#         plt.imshow(rotate(template_data[i, :, :], 90, reshape=True), cmap='gray')
#         img1 = plt.imshow(rotate(masked_anatomical[i, :, :], 90, reshape=True), cmap=custom_cmap, alpha=0.8, vmin=0.1, vmax=n_labels)
#         plt.title(f'Sagittal slice at x={i}')
        
#         # Coronal slices
#         ax2 = plt.subplot(1, 3, 2)
#         ax2.axis('off')
#         plt.imshow(rotate(template_data[:, i, :], 90, reshape=True), cmap='gray')
#         img2 = plt.imshow(rotate(masked_anatomical[:, i, :], 90, reshape=True), cmap=custom_cmap, alpha=0.8, vmin=0.1, vmax=n_labels)
#         plt.title(f'Coronal slice at y={i}')
        
#         # Axial slices
#         ax3 = plt.subplot(1, 3, 3)
#         ax3.axis('off')
#         plt.imshow(rotate(template_data[:, :, i], 90, reshape=True), cmap='gray')
#         img3 = plt.imshow(rotate(masked_anatomical[:, :, i], 90, reshape=True), cmap=custom_cmap, alpha=0.8, vmin=0.1, vmax=n_labels)
#         plt.title(f'Axial slice at z={i}')
        
#         plt.suptitle('Visual Mask', fontsize=16)
#         plt.show()

#         plt.figure(figsize=(15, 5))
        
#         # Sagittal slices
#         ax1 = plt.subplot(1, 3, 1)
#         ax1.axis('off')
#         plt.imshow(rotate(template_data[i, :, :], 90, reshape=True), cmap='gray')
#         img1 = plt.imshow(rotate(masked_saliency_n[i, :, :], 90, reshape=True), cmap=custom_cmap_s, norm=norm)
#         plt.title(f'Sagittal slice at x={i}')
        
#         # Coronal slices
#         ax2 = plt.subplot(1, 3, 2)
#         ax2.axis('off')
#         plt.imshow(rotate(template_data[:, i, :], 90, reshape=True), cmap='gray')
#         img2 = plt.imshow(rotate(masked_saliency_n[:, i, :], 90, reshape=True), cmap=custom_cmap_s, norm=norm)
#         plt.title(f'Coronal slice at y={i}')
        
#         # Axial slices
#         ax3 = plt.subplot(1, 3, 3)
#         ax3.axis('off')
#         plt.imshow(rotate(template_data[:, :, i], 90, reshape=True), cmap='gray')
#         img3 = plt.imshow(rotate(masked_saliency_n[:, :, i], 90, reshape=True), cmap=custom_cmap_s, norm=norm)
#         plt.title(f'Axial slice at z={i}')

#         if color_bar:
#             cbar = plt.colorbar(img3, ax=[ax1, ax2, ax3], orientation='vertical', fraction=0.046, pad=0.04)
#             cbar.set_label('Normalized Saliency')
#             color_bar = False

#         plt.suptitle('Saliency Distribution', fontsize=16)
#         plt.show()

### STATISTICAL TESTS ###

# Define global variables to hold data, assuming they fit comfortably in memory
# and can be accessed by subprocesses in multiprocessing environment.

def one_sample_permutation_test(lab, pred, alpha = 0.05, num_permutations = 999):
    """
    Conducts a one-sample permutation test to evaluate the significance of observed correlations against a null distribution.

    Parameters:
        lab (numpy array): The ground truth labels, expected to be of shape (TR, voxels).
        pred (numpy array): The model predictions, expected to be of the same shape as 'lab'.
        alpha (float): Significance level for determining statistical significance.
        num_permutations (int): Number of permutations to generate the null distribution.

    Returns:
        numpy array: P-values for each voxel, providing insight into the significance of the observed correlations.
    
    This function computes the correlations between predicted and actual data for each voxel and compares these
    correlations against a null distribution generated by randomly shuffling the labels.
    """
    global tot_lab, tot_pred, valid_indices
    tot_lab, tot_pred = lab, pred

    TR, voxels = tot_pred.shape
    
    # Calculate valid indices only once
    valid_indices = [i for i in range(voxels) if np.std(tot_lab[:, i]) != 0 and np.std(tot_pred[:, i]) != 0]
    
    # Compute observed correlations using valid indices
    correlations_voxel = np.array([
        stats.pearsonr(tot_lab[:, i], tot_pred[:, i])[0] 
        for i in valid_indices
    ])
    mean_correlation_voxel = np.mean(correlations_voxel)
    median_correlation_voxel = np.median(correlations_voxel)
    std_correlation_voxel = np.std(correlations_voxel)

    # Using multiprocessing to handle the computationally intensive task
    with Pool() as pool:
        # Create unique seeds based on the process index, could be random or sequential
        seeds = np.random.randint(0, 1000000, size=num_permutations)
        correlations_voxel_perm = np.asarray(pool.map(compute_permutation, seeds)).T
    mean_correlations_perm = np.concatenate(correlations_voxel_perm[:, :10], axis=0) #for visualisation of the null distribution

    plt.figure(figsize=(8, 4))
    plt.hist(correlations_voxel, bins=30, alpha=0.5, color='green', density=True, label='Prediction Distribution')
    sns.kdeplot(correlations_voxel, fill=True, color='g')
    plt.axvline(x=median_correlation_voxel, color='r', linestyle=':', linewidth=1, label=f'Median: {median_correlation_voxel:.3f}')
    plt.axvline(x=mean_correlation_voxel, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_correlation_voxel:.3f}')
    plt.axvspan(mean_correlation_voxel - std_correlation_voxel, mean_correlation_voxel + std_correlation_voxel, color='red', alpha=0.05, label=f'std: {std_correlation_voxel:.3f}')
    
    # Plotting the histogram for mean_correlations_perm
    plt.hist(mean_correlations_perm, bins=100, alpha=0.5, color='grey', density=True, label=f'Null Distribution\n(Permutation Test)')
    sns.kdeplot(mean_correlations_perm, fill=True, color='grey')

    plt.xlabel('Correlation Coefficient')
    plt.ylabel('Density')
    plt.title(f'Distribution of Correlations per Voxel (N = {voxels})')
    plt.legend()
    plt.show()
    
    # Calculate the p-values for each observed correlation
    p_values = np.sum(correlations_voxel[:, np.newaxis] < correlations_voxel_perm, axis=1)/num_permutations
    # Assuming you have an array of p-values called `p_values` and other variables `num_permutations`, `voxels`
    plt.figure(figsize=(8, 2))
    # Add a very small value to p_values to avoid log(0) issue, assuming no p-value is exactly 0
    adjusted_p_values = np.maximum(p_values, 1e-6)
    plt.hist(adjusted_p_values, bins=1000, alpha=0.2, color='b', cumulative=True, density=False)
    plt.axvline(alpha, color='k', linestyle=':', linewidth=0.6)
    plt.text(alpha, 170, f' alpha = {alpha}', color='k', ha='left')
    plt.axhline(y=np.sum(p_values < alpha), color='k', linestyle=':', linewidth=0.6)
    plt.text(alpha, 170+np.sum(p_values < alpha), f' {100*np.sum(p_values < alpha)/voxels:.1f}%', color='r', ha='left')
    plt.plot(alpha, np.sum(p_values < alpha), 'x', markersize=6, markeredgewidth=2, color='red')
    plt.axvline(alpha/voxels, color='k', linestyle=':', linewidth=0.6)
    plt.text(alpha/voxels, 170, f' alpha = {alpha}/{voxels}', color='k', ha='left')
    plt.axhline(y=np.sum(p_values < alpha/voxels), color='k', linestyle=':', linewidth=0.6)
    plt.text(alpha/voxels, 170+np.sum(p_values < alpha/voxels), f' {100*np.sum(p_values < alpha/voxels)/voxels:.1f}%', color='r', ha='left')
    plt.plot(alpha/voxels, np.sum(p_values < alpha/voxels), 'x', markersize=6, markeredgewidth=2, color='red')
    plt.xscale('log')
    plt.xlabel('p-value (log scale)')
    plt.ylabel('Cumulative count')
    plt.xlim([0.000001, 1.001])
    plt.title(f'Cumulative Distribution of p-values (Permutation Test with {num_permutations} permutations)')
    plt.show()
    return p_values

# Define the compute_permutation function at the top level
def compute_permutation(seed):
    """
    Compute a permutation of Pearson correlation coefficients between shuffled labels and actual predictions.

    Parameters:
        seed (int): A seed number used to ensure reproducibility of the random shuffling process.

    Returns:
        numpy array: An array of Pearson correlation coefficients calculated between shuffled labels
                     and actual predictions for each valid index.

    This function shuffles the labels array using a given seed to ensure the randomness is reproducible, 
    then calculates the Pearson correlation coefficient between these shuffled labels and the original predictions.
    It operates on the global variables `tot_lab`, `tot_pred`, and `valid_indices`, which must be defined outside
    this function. These variables represent the total labels, total predictions, and indices of valid data points, 
    respectively.
    """
    np.random.seed(seed)  # Set a unique seed for each process
    shuffled_labels = np.copy(tot_lab)
    np.random.shuffle(shuffled_labels)  # Shuffle labels
    perm = np.array([
        stats.pearsonr(shuffled_labels[:, i], tot_pred[:, i])[0] 
        for i in valid_indices
    ])
    return perm

def significant_voxels(p_values, alpha = 0.05, voxels = 4330):
    """
    Identify and visualize (spacially) significant voxels from fMRI data based on provided p-values.

    Parameters:
        p_values (numpy array): Array of p-values for each voxel.
        alpha (float): Significance level for the testing, default is 0.05.
        voxels (int): Total number of voxels considered in the analysis, default is 4330.
    """
    # you can choose another fMRI data as template
    brain_path = '/media/miplab-nas2/Data2/Movies_Emo/Preprocessed_data/sub-S01/ses-1/pp_sub-S01_ses-1_YouAgain.feat/filtered_func_data_res_MNI.nii'
    brain_img = nib.load(brain_path)
    brain_data = np.asanyarray(brain_img.dataobj)[..., 50] #only look at 1 TR

    # path to the mask
    mask_path = '/home/chchan/Michael-Nas2/ml-students2023/resources/vis_mask.nii'
    img_mask = nib.load(mask_path)
    mask_2d = np.asanyarray(img_mask.dataobj).reshape(-1,)
    indices = np.where(mask_2d == 1)[0] # 4330 relevant voxels
    
    #zero-filled array with the same shape as the original 3D data, and flatten
    significant_voxels = np.zeros(brain_data.shape, dtype=np.float32).flatten()
    #set to 1 the voxels that have a p-value < alpha/voxels
    significant_voxels[indices] = np.where(p_values > alpha/voxels, 0, 1)
    #reshape back to the original 3D shape
    significant_voxels = significant_voxels.reshape(brain_data.shape)

    #plot some brain slices
    for i in range(0, 90, 5):
        plt.subplot(1,3,1)
        plt.imshow(np.maximum(significant_voxels[i,:,:],normalize(brain_data[i,:,:])), cmap='gray')
        plt.subplot(1,3,2)
        plt.imshow(np.maximum(significant_voxels[:,i,:],normalize(brain_data[:,i,:])), cmap='gray')
        plt.subplot(1,3,3)
        plt.imshow(np.maximum(significant_voxels[:,:,i],normalize(brain_data[:,:,i])), cmap='gray')
        plt.show()


def two_samples_permutation_test(model1, lab1, pred1, model2, lab2, pred2, num_permutations = 999):
    """
    Conducts a two-sample permutation test to compare two sets of correlations between predicted and actual data.

    Parameters:
        model1 (str): Name or identifier for the first model.
        lab1 (numpy array): Ground truth labels for the first model.
        pred1 (numpy array): Predictions from the first model.
        model2 (str): Name or identifier for the second model.
        lab2 (numpy array): Ground truth labels for the second model.
        pred2 (numpy array): Predictions from the second model.
        num_permutations (int): Number of permutations to perform.

    Returns:
        Displays a visual comparison of correlation distributions and reports the p-value from the permutation test.
    
    This function compares the mean difference in correlations between two different models or conditions using a
    permutation test, providing a statistical comparison of their performance.
    """
    all_encoded, all_labels = [], []
    for key in pred1.keys():
        all_encoded.append(pred1[key])
        all_labels.append(lab1[key])
    tot_pred = np.concatenate(all_encoded, axis=0)
    tot_lab = np.concatenate(all_labels, axis=0)
    voxels = tot_pred.shape[1]
    valid_indices = [i for i in range(voxels) if np.std(tot_lab[:, i]) != 0 and np.std(tot_pred[:, i]) != 0]
    sample1 = np.array([
        stats.pearsonr(tot_lab[:, i], tot_pred[:, i])[0] 
        for i in valid_indices
    ])
    mean1= np.mean(sample1)

    all_encoded, all_labels = [], []
    for key in pred2.keys():
        all_encoded.append(pred2[key])
        all_labels.append(lab2[key])
    tot_pred = np.concatenate(all_encoded, axis=0)
    tot_lab = np.concatenate(all_labels, axis=0)
    voxels = tot_pred.shape[1]
    valid_indices = [i for i in range(voxels) if np.std(tot_lab[:, i]) != 0 and np.std(tot_pred[:, i]) != 0]
    sample2 = np.array([
        stats.pearsonr(tot_lab[:, i], tot_pred[:, i])[0] 
        for i in valid_indices
    ])
    mean2= np.mean(sample2)
    
    test_result = perm_test((sample1, sample2), diff_means, permutation_type='independent', alternative='less', n_resamples=num_permutations)

    plt.figure(figsize=(8, 4))
    plt.hist(sample1, bins=30, alpha=0.5, color='salmon', density=True)
    sns.kdeplot(sample1, fill=True, color='red', label=model1)
    plt.axvline(mean1, color='r', linestyle='--', linewidth=1)
    plt.text(mean1, 8.8, f'Mean: {mean1:.3f} ', color='r', ha='right')
    plt.hist(sample2, bins=30, alpha=0.5, color='lightgreen', density=True)
    sns.kdeplot(sample2, fill=True, color='g', label=model2)
    plt.axvline(mean2, color='g', linestyle='--', linewidth=1)
    plt.text(mean2, 8.8, f' Mean: {mean2:.3f}', color='g', ha='left')
    plt.title(f'Two samples permutation test ({num_permutations} permutations)\nMean difference: {test_result.statistic} (p = {test_result.pvalue:.3e})')
    plt.xlabel('Correlation Coefficient')
    plt.ylabel('Density')
    plt.legend()
    plt.show()

# Define the statistic function
def diff_means(x, y):
    return np.mean(x) - np.mean(y)

from openpyxl import load_workbook

def save_data(excel_file, values_dict, save_model_as, mask_size, num_epochs, lr, encoder_weight):
    values_dict['save_model_as'] = save_model_as
    values_dict['mask_size'] = mask_size
    values_dict['num_epochs'] = num_epochs
    values_dict['lr'] = lr
    values_dict['encoder_weight'] = encoder_weight

    # Load the existing workbook
    workbook = load_workbook(excel_file)

    # Select the active worksheet (or specify the sheet name)
    worksheet = workbook.active

    # Retrieve the headers from the first row of the worksheet
    headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]

    # Ensure that all headers are strings (in case they are None or other types)
    headers = [str(header) if header is not None else '' for header in headers]

    # Check if 'save_model_as' is in headers
    if 'save_model_as' not in headers:
        # Handle the case where 'save_model_as' is not a column in the Excel file
        raise ValueError("'save_model_as' column not found in Excel file.")

    # Find the index of 'save_model_as' column (zero-based index)
    save_model_as_index = headers.index('save_model_as')

    # Collect all existing values in 'save_model_as' column
    existing_save_model_as_values = set()
    for row in worksheet.iter_rows(min_row=2, min_col=save_model_as_index+1, max_col=save_model_as_index+1):
        cell_value = row[0].value
        existing_save_model_as_values.add(cell_value)

    # Check if the 'save_model_as' value already exists
    if save_model_as in existing_save_model_as_values:
        print(f"{save_model_as} already exists in {excel_file}. Skipping save.")
        return  # Do not proceed with appending

    # Prepare the new row data, ordering the values according to the headers
    new_row = [values_dict.get(header, '') for header in headers]

    # Append the new row to the worksheet
    worksheet.append(new_row)

    # Save the workbook to persist the changes
    workbook.save(excel_file)

    print(f"{save_model_as} has been saved to {excel_file}.")
    print(f"num_epochs: {num_epochs}, lr: {lr}, encoder_weight: {encoder_weight}")

def print_dict_tree(d, indent=0):
    """
    Recursively prints the structure of a nested dictionary like a tree.

    Args:
    - d (dict): The nested dictionary to print.
    - indent (int): The indentation level (used for recursion).
    """
    for key, value in d.items():
        if isinstance(value, dict):
            print('  ' * indent + f'{key}')
            print_dict_tree(value, indent + 1)
        elif isinstance(value, np.ndarray):
            print('  ' * indent + f'{key} (shape: {value.shape})')
        else:
            print('  ' * indent + f'{key} (type: {type(value)})')