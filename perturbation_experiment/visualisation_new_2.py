# -*- coding: utf-8 -*-

"""
This file contains functions related to visualisation of the model results.

Functions:
    plot_train_losses
    plot_metrics
    plot_decoder_predictions
    one_sample_permutation_test
    compute_permutation
    significant_voxels
    two_samples_permutation_test
    diff_means
    print_dict_tree
"""

from imports import np, torch, F, plt, sns, explained_variance_score, stats, Pool, perm_test, nib, ssim, TotalVariation
from dataset import normalize
import pandas as pd
from openpyxl import load_workbook
import matplotlib.colors as mcolors
import nibabel as nib
from nilearn import datasets, plotting, image
from scipy.ndimage import rotate

### TRAINING VISUALISATION ###


'''
def plot_train_losses(history, start_epoch):
    """
    Plot the training losses and other metrics from the training history.

    Parameters:
        history (dict): Contains the training history data including loss and other metrics.
        start_epoch (int): The starting epoch number for plotting.

    Output:
        Displays plots for each metric in the history.
    """
    x = np.arange(len(history['total_loss'])) + start_epoch 
    
    plt.plot(x, history['total_loss'])
    plt.title(history['metrics_names'][-1])
    plt.xlabel('epoch')
    plt.ylabel('value')
    plt.show()

    for i in range(len(history['metrics_names']) - 1):
        plt.plot(x, history['other_metrics'][:, i])
        plt.title(history['metrics_names'][i])
        plt.xlabel('epoch')
        plt.ylabel('value')
        plt.show()
'''


def plot_train_losses(history, start_epoch, save_plots=False, save_path=None):
    """
    Plot the training losses and other metrics from the training history.

    Parameters:
        history (dict): Contains the training history data including loss and other metrics.
        start_epoch (int): The starting epoch number for plotting.
        save_plots (bool, optional): Whether to save plots to files. Default is False.
        save_path (str, optional): Base path to save plots. Default is None.

    Output:
        Displays plots for each metric in the history.
    """
    x = np.arange(len(history['total_loss'])) + start_epoch 
    
    plt.figure(figsize=(10, 6))
    plt.plot(x, history['total_loss'])
    plt.title(history['metrics_names'][-1])
    plt.xlabel('epoch')
    plt.ylabel('value')
    
    # Save total loss plot
    if save_plots and save_path:
        # Save with specific filename for total loss
        total_loss_path = save_path.replace('.png', '_total_loss.png')
        plt.savefig(total_loss_path, bbox_inches='tight', dpi=300)
    
    plt.show()

    for i in range(len(history['metrics_names']) - 1):
        plt.figure(figsize=(10, 6))
        plt.plot(x, history['other_metrics'][:, i])
        plt.title(history['metrics_names'][i])
        plt.xlabel('epoch')
        plt.ylabel('value')
        
        # Save individual metric plots
        if save_plots and save_path:
            # Create specific filename for each metric
            metric_path = save_path.replace('.png', f'_{history["metrics_names"][i]}.png')
            plt.savefig(metric_path, bbox_inches='tight', dpi=300)
        
        plt.show()





def plot_train_losses_with_val(history, start_epoch, save_plots=False, save_path=None):
    """
    Plot training and validation losses over epochs.
    
    Arguments:
        history (dict): Dictionary containing 'total_loss', 'validation_loss', and other metrics
        start_epoch (int): Starting epoch number
        save_plots (bool): Whether to save the plot
        save_path (str): Path to save the plot if save_plots is True
    """
    import matplotlib.pyplot as plt
    
    epochs = range(start_epoch, start_epoch + len(history['total_loss']))
    
    plt.figure(figsize=(12, 8))
    
    # Plot total loss (training) and validation loss
    plt.subplot(2, 1, 1)
    plt.plot(epochs, history['total_loss'], 'b-', label='Training Loss', linewidth=2)
    
    # Validation loss is only recorded every 5 epochs
    if len(history['validation_loss']) > 0:
        val_epochs = range(start_epoch + 4, start_epoch + len(history['total_loss']) + 1, 5)  # Every 5th epoch starting from epoch 5
        # Adjust if validation array is shorter
        val_epochs_actual = val_epochs[:len(history['validation_loss'])]
        plt.plot(val_epochs_actual, history['validation_loss'], 'r-', label='Validation Loss', linewidth=2, marker='o')
    
    plt.xlabel('Epoch')
    plt.ylabel('Loss')
    plt.title('Training and Validation Loss')
    plt.legend()
    plt.grid(True, alpha=0.3)
    
    # Plot other metrics if available
    if len(history['other_metrics']) > 0 and len(history['other_metrics'][0]) > 0:
        plt.subplot(2, 1, 2)
        other_metrics = np.array(history['other_metrics'])
        
        if 'metrics_names' in history and len(history['metrics_names']) > 0:
            # Plot each metric separately
            for i, metric_name in enumerate(history['metrics_names'][:len(other_metrics[0])]):
                plt.plot(epochs, other_metrics[:, i], label=f'Training {metric_name}', linewidth=2)
        else:
            # Plot all metrics without specific names
            for i in range(other_metrics.shape[1]):
                plt.plot(epochs, other_metrics[:, i], label=f'Metric {i+1}', linewidth=2)
        
        plt.xlabel('Epoch')
        plt.ylabel('Metric Value')
        plt.title('Training Metrics')
        plt.legend()
        plt.grid(True, alpha=0.3)
    
    plt.tight_layout()
    
    if save_plots and save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"Training plot saved to {save_path}")
    
    plt.show()





### TESTING VISUALISATION ###

def z_score(data):
    return (data - np.mean(data, axis=0)) / np.std(data, axis=0)

def plot_metrics(label, prediction, movie, plot_TR = False, performance_dict = None, display_plots=True, save_plots=False, save_path=None):
#def plot_metrics(label, prediction, movie, plot_TR = False, performance_dict = None, display_plots=True):
    """
    Plot various statistical metrics and distributions for the prediction quality of a model.

    Parameters:
        label (np.array): Ground truth values, shape (n_samples, n_features).
        prediction (np.array): Model predictions, shape (n_samples, n_features).
        movie (str): Name of the dataset or context for the plot titles.
        plot_TR (bool, optional): If true, also plots time-resolved (TR) metrics. Default is False.

    Output:
        Visualizations of correlation, explained variance, and other statistical measures between label and prediction.
    """

    # Create output directory if it doesn't exist
    if save_plots and save_path:
        import os
        os.makedirs(os.path.dirname(save_path) if os.path.dirname(save_path) else 'outputs', exist_ok=True) #what?
    

    if movie == 'all':
        all_labels, all_predictions = label.copy(), prediction.copy()
        label, prediction = [], []
        for key in all_labels.keys():
            all_predictions[key] = z_score(all_predictions[key])
            all_labels[key] = z_score(all_labels[key])
            prediction.append(all_predictions[key])
            label.append(all_labels[key])
        prediction = np.concatenate(prediction, axis=0)
        label = np.concatenate(label, axis=0)
    else: 
        label = z_score(label)
        prediction = z_score(prediction)
    
    TR, voxels = prediction.shape

    correlations_voxel = np.array([
        stats.pearsonr(label[:, i], prediction[:, i])[0] 
        for i in range(voxels) 
        if np.std(label[:, i]) != 0 and np.std(prediction[:, i]) != 0
    ])
    mean_correlation_voxel = np.mean(correlations_voxel)
    median_correlation_voxel = np.median(correlations_voxel)
    std_correlation_voxel = np.std(correlations_voxel)

    best_voxel = np.argmax(correlations_voxel)
    worst_voxel = np.argmin(correlations_voxel)

    mse_voxel = np.array([
        F.mse_loss(torch.from_numpy(label[:, i]).unsqueeze(0), torch.from_numpy(prediction[:, i]).unsqueeze(0), reduction='mean').item()
        for i in range(voxels)
        if np.std(label[:, i]) != 0  
    ])
    mean_mse_voxel = np.mean(mse_voxel)
    median_mse_voxel = np.median(mse_voxel)
    std_mse_voxel = np.std(mse_voxel)

    if display_plots:
        print(f'### {movie}: ({TR}, {voxels}) ###')
    
    if movie == 'all':
        all_correlation = {'all': correlations_voxel}
        all_mse = {'all': mse_voxel}
        for key in all_labels.keys():
            voxels = all_labels[key].shape[1]
            corr_movie = np.array([
                stats.pearsonr(all_labels[key][:, i], all_predictions[key][:, i])[0] 
                for i in range(voxels) 
                if np.std(all_labels[key][:, i]) != 0 and np.std(all_predictions[key][:, i]) != 0
            ])
            mse_movie = np.array([
                F.mse_loss(torch.from_numpy(all_labels[key][:, i]).unsqueeze(0), torch.from_numpy(all_predictions[key][:, i]).unsqueeze(0), reduction='mean').item()
                for i in range(voxels)
                if np.std(all_labels[key][:, i]) != 0  
            ])
            all_correlation[key] = corr_movie
            all_mse[key] = mse_movie
        performance_dict['mean_corr_E'], performance_dict['median_corr_E'] = plot_scatter_metrics(all_correlation, 'Correlations', display_plots)
        performance_dict['mean_mse_E'], performance_dict['median_mse_E'] = plot_scatter_metrics(all_mse, 'MSE', display_plots)

    if display_plots:
        plt.figure(figsize=(17, 3))
        plt.subplot(121)
        plt.hist(correlations_voxel, bins=30, alpha=0.5, color='green', density=True, label='Histogram')
        sns.kdeplot(correlations_voxel, fill=True, color='g', label='KDE')
        plt.axvline(x=median_correlation_voxel, color='r', linestyle=':', linewidth=1, label=f'Median: {median_correlation_voxel:.3f}')
        plt.axvline(x=mean_correlation_voxel, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_correlation_voxel:.3f}')
        plt.axvspan(mean_correlation_voxel - std_correlation_voxel, mean_correlation_voxel + std_correlation_voxel, color='red', alpha=0.05, label=f'std: {std_correlation_voxel:.3f}')
        plt.xlabel('Correlation Coefficient')
        plt.ylabel('Density')
        plt.title(f'Distribution of Correlations per Voxel (N = {voxels})')
        plt.legend() 
        plt.subplot(122)
        plt.hist(mse_voxel, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(mse_voxel, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_voxel, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_voxel:.3f}')
        plt.axvline(x=mean_mse_voxel, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_voxel:.3f}')
        plt.axvspan(mean_mse_voxel - std_mse_voxel, mean_mse_voxel + std_mse_voxel, color='red', alpha=0.05, label=f'std: {std_mse_voxel:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE per Voxel (N = {voxels})')
        plt.legend()
        # Save the metrics plot if requested
        if save_plots and save_path:
            plt.savefig(save_path, bbox_inches='tight', dpi=300)
#        plt.show()
        plt.show()
    
        plt.figure(figsize=(22, 7))
        plt.subplot(2, 1, 1)
        plt.title(f'Best Voxel (Corr({best_voxel}) = {np.max(correlations_voxel):.3f})')
        plt.plot(label[:,best_voxel], label = 'ground truth')
        plt.plot(prediction[:,best_voxel], label = 'prediction')
        plt.legend()
        plt.subplot(2, 1, 2)
        plt.title(f'Worst Voxel (Corr({worst_voxel}) = {np.min(correlations_voxel):.3f})')
        plt.plot(label[:,worst_voxel], label = 'ground truth')
        plt.plot(prediction[:,worst_voxel], label = 'prediction')
        plt.legend()
        # Save the voxel comparison plot if requested
        if save_plots and save_path:
            voxel_path = save_path.replace('.png', '_voxels.png')
            plt.savefig(voxel_path, bbox_inches='tight', dpi=300)
#        plt.show()
        plt.show()

    if plot_TR and display_plots:
    
        correlations_TR = np.array([
            stats.pearsonr(label[i, :], prediction[i, :])[0] 
            for i in range(TR) 
            if np.std(label[i, :]) != 0 and np.std(prediction[i, :]) != 0
        ])
        mean_correlation_TR = np.mean(correlations_TR)
        median_correlation_TR = np.median(correlations_TR)
        std_correlation_TR = np.std(correlations_TR)
    
        best_TR = np.argmax(correlations_TR)
        worst_TR = np.argmin(correlations_TR)
    
        mse_TR = np.array([
            F.mse_loss(torch.from_numpy(label[i, :]).unsqueeze(0), torch.from_numpy(prediction[i, :]).unsqueeze(0), reduction='mean').item()
            for i in range(TR)
            if np.std(label[i, :]) != 0  
        ])
        mean_mse_TR = np.mean(mse_TR)
        median_mse_TR = np.median(mse_TR)
        std_mse_TR = np.std(mse_TR)
    
        #plot MSE and SSIM
        plt.figure(figsize=(17, 3))
        plt.subplot(121)
        plt.hist(correlations_TR, bins=30, alpha=0.5, color='green', density=True, label='Histogram')
        sns.kdeplot(correlations_TR, fill=True, color='g', label='KDE')
        plt.axvline(x=median_correlation_TR, color='r', linestyle=':', linewidth=1, label=f'Median: {median_correlation_TR:.3f}')
        plt.axvline(x=mean_correlation_TR, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_correlation_TR:.3f}')
        plt.axvspan(mean_correlation_TR - std_correlation_TR, mean_correlation_TR + std_correlation_TR, color='red', alpha=0.05, label=f'std: {std_correlation_TR:.3f}')
        plt.xlabel('Correlation Coefficient')
        plt.ylabel('Density')
        plt.title(f'Distribution of Correlations per TR (N = {TR})')
        plt.legend() 
        plt.subplot(122)
        plt.hist(mse_TR, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(mse_TR, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_TR, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_TR:.3f}')
        plt.axvline(x=mean_mse_TR, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_TR:.3f}')
        plt.axvspan(mean_mse_TR - std_mse_TR, mean_mse_TR + std_mse_TR, color='red', alpha=0.05, label=f'std: {std_mse_TR:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE per TR (N = {TR})')
        plt.legend()
        # Save the TR metrics plot if requested
        if save_plots and save_path:
            tr_path = save_path.replace('.png', '_TR_metrics.png')
            plt.savefig(tr_path, bbox_inches='tight', dpi=300)
        plt.show()
    
        plt.figure(figsize=(22, 7))
        plt.subplot(2, 1, 1)
        plt.title(f'Best TR (Corr({best_TR}) = {np.max(correlations_TR):.3f})')
        plt.plot(label[best_TR, :], label = 'ground truth')
        plt.plot(prediction[best_TR, :], label = 'prediction')
        plt.legend()
        plt.subplot(2, 1, 2)
        plt.title(f'Worst TR (Corr({worst_TR}) = {np.min(correlations_TR):.3f})')
        plt.plot(label[worst_TR, :], label = 'ground truth')
        plt.plot(prediction[worst_TR, :], label = 'prediction')
        plt.legend()
        # Save the TR comparison plot if requested
        if save_plots and save_path:
            tr_voxel_path = save_path.replace('.png', '_TR_voxels.png')
            plt.savefig(tr_voxel_path, bbox_inches='tight', dpi=300)
        plt.show()
#        plt.show()

    if performance_dict:
        return performance_dict

def plot_scatter_metrics(metrics, name, display_plots=True):
    colors = {'Correlations': 'green',
              'MSE': 'red',
              'SSIM': 'blue'}
    
    # Convert dictionary to DataFrame
    df = pd.DataFrame([
        {'Movie': movie, 'Value': value}
        for movie, values in metrics.items()
        for value in values
    ])

    # Retrieve statistics for the 'all' movie
    first_values = np.array(metrics['all'])
    median_value = np.median(first_values)
    mean_value = np.mean(first_values)
    std_dev_value = np.std(first_values)

    if display_plots:
        movie_abreviations = {'Sintel': '1-St',
                          'Payload': '2-Pld',
                          'TearsOfSteel': '3-TOS',
                          'Superhero': '4-Sh',
                          'BigBuckBunny': '5-BBB',
                          'FirstBite': '6-FB',
                          'BetweenViewings': '7-BW',
                          'AfterTheRain': '8-ATR',
                          'TheSecretNumber': '9-TSN',
                          'Chatter': '10-Cht',
                          'Spaceman': '11-Spm',
                          'LessonLearned': '12-LL',
                          'YouAgain': '*13-YA',
                          'ToClaireFromSonny': '14-TCFS'}
        for key, value in movie_abreviations.items():
            df['Movie'] = df['Movie'].replace(key, value)
    
        
        categories = df['Movie'].unique()
        
        fig, ax = plt.subplots(figsize=(15, 9))

        # Create violin plot for the categories
        sns.violinplot(x='Movie', y='Value', data=df,
                       ax=ax, density_norm='width', fill=True,
                       linecolor='k', linewidth=2,
                       color='k', alpha=0.07,
                       inner_kws=dict(box_width=8, whis_width=2, color="0.7"))

        jitter, size = 0.08, 1
        if first_values.shape[0] > 3000:
            jitter, size = 0.04, 0.5
        # Plot scatter points for each category
        for idx, category in enumerate(categories):
            y = df[df['Movie'] == category]['Value']
            x = np.random.normal(idx, jitter, size=len(y))  # Add jitter
    
            # Assign color based on the category
            if category == 'all':
                color = colors.get(name, '0.5')  # Use the color corresponding to 'name'
            else:
                color = 'k'  # Default grey color for other categories
            
            ax.scatter(x, y, alpha=0.5, s=size, color=color)

        # Statistical lines for the 'all' movie
        ax.axhline(mean_value, color='r', linestyle='--', linewidth=3, label=f'Mean: {mean_value:.3f}')
        ax.axhline(median_value, color='g', linestyle=':', linewidth=3,label=f'Median: {median_value:.3f}')
        ax.axhspan(mean_value - std_dev_value, mean_value + std_dev_value, color='red', alpha=0.05, label=f'Std: {std_dev_value:.3f}')
        
        # Set x-ticks and labels
        ax.set_xticks(np.arange(len(categories)))
        ax.set_xticklabels(categories, rotation=30, fontsize=25, ha='center')
        
        # Customize plot
        ax.set_title(f'{name} Distribution by Movie', fontsize=20, y=1.1)
        ax.set_ylabel(name, fontsize=33)
        ax.set_xlabel('Movie', fontsize=33)
        ax.tick_params(axis='y', labelsize=25, rotation=60)
        ax.legend(fontsize=29, loc='upper center', bbox_to_anchor=(0.5, 1.12),
              fancybox=True, shadow=True, ncol=3)
        ax.grid(color='grey', linestyle=':', linewidth=0.8)
        plt.tight_layout()
        plt.show()
        
    return np.round(mean_value, 3), np.round(median_value, 3)

# def plot_scatter_metrics(metrics, name, display_plots=True):
#     colors = {'Correlations': 'green',
#               'MSE': 'red',
#               'SSIM': 'blue'}
    
#     # Convert dictionary to DataFrame
#     df = pd.DataFrame([
#         {'Movie': movie, 'Value': value}
#         for movie, values in metrics.items()
#         for value in values
#     ])

#     # Retrieve statistics for the 'all' movie
#     first_values = np.array(metrics['all'])
#     median_value = np.median(first_values)
#     mean_value = np.mean(first_values)
#     std_dev_value = np.std(first_values)

#     #np.save(f'encoder_15364_2_{name}', first_values)

#     if display_plots:
#         movie_abreviations = {'Sintel': '1-St',
#                           'Payload': '2-Pld',
#                           'TearsOfSteel': '3-TOS',
#                           'Superhero': '4-Sh',
#                           'BigBuckBunny': '5-BBB',
#                           'FirstBite': '6-FB',
#                           'BetweenViewings': '7-BW',
#                           'AfterTheRain': '8-ATR',
#                           'TheSecretNumber': '9-TSN',
#                           'Chatter': '10-Cht',
#                           'Spaceman': '11-Spm',
#                           'LessonLearned': '12-LL',
#                           'YouAgain': '*13-YA',
#                           'ToClaireFromSonny': '14-TCFS'}
#         for key, value in movie_abreviations.items():
#             df['Movie'] = df['Movie'].replace(key, value)
    
        
#         categories = df['Movie'].unique()
#         positions = np.arange(len(categories))
        
#         # Prepare data for boxplot
#         box_data = [df[df['Movie'] == category]['Value'] for category in categories]
        
#         fig, ax = plt.subplots(figsize=(15, 10))
        
#         # Plot the boxplot shifted slightly to the left
#         box_positions = positions - 0.1
#         bp = ax.boxplot(box_data, positions=box_positions, widths=0.1, patch_artist=True, showfliers=False)
        
#         # Customize boxplot appearance
#         for box in bp['boxes']:
#             box.set(facecolor='none', linewidth=1)
#         for whisker in bp['whiskers']:
#             whisker.set(color='black', linewidth=1)
#         for cap in bp['caps']:
#             cap.set(color='black', linewidth=1)
#         for median in bp['medians']:
#             median.set(color='black', linewidth=1)
        
#         # Plot the scatter points shifted slightly to the right
#         strip_positions = positions + 0.2
#         for idx, category in enumerate(categories):
#             y = df[df['Movie'] == category]['Value']
#             x = np.random.normal(strip_positions[idx], 0.08, size=len(y))  # Add jitter
    
#             # Assign color based on the category
#             if category == 'all':
#                 color = colors.get(name, '0.5')  # Use the color corresponding to 'name'
#             else:
#                 color = 'k'  # Default grey color for other categories
            
#             ax.scatter(x, y, alpha=0.5, s=1, color=color)
        
#         # Set x-ticks and labels
#         ax.set_xticks(positions)
#         ax.set_xticklabels(categories, rotation=30, fontsize=25, ha='center')
        
#         # Statistical lines for the 'all' movie
#         ax.axhline(mean_value, color='r', linestyle='--', linewidth=3, label=f'Mean: {mean_value:.3f}')
#         ax.axhline(median_value, color='g', linestyle=':', linewidth=3,label=f'Median: {median_value:.3f}')
#         ax.axhspan(mean_value - std_dev_value, mean_value + std_dev_value, color='red', alpha=0.08, label=f'Std: {std_dev_value:.3f}')
        
#         # Customize plot
#         ax.set_title(f'{name} Distribution by Movie', fontsize=20, y=1.1)
#         ax.set_ylabel(name, fontsize=24)
#         ax.set_xlabel('Movie', fontsize=24)
#         ax.tick_params(axis='y', labelsize=25, rotation=60)
#         ax.legend(fontsize=26, loc='upper center', bbox_to_anchor=(0.5, 1.09),
#               fancybox=True, shadow=True, ncol=3)
#         ax.grid(color='grey', linestyle=':', linewidth=0.8)
#         plt.tight_layout()
#         plt.show()

#     return np.round(mean_value, 3), np.round(median_value, 3)



def plot_decoder_predictions(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, save_path_prefix=None, model_name="", temporal=False, 
#individual_metrics=None
psim_and_tv = False
):
#def plot_decoder_predictions(predictions, videos, performance_dict = None, display_plots=True):
    """
    Display comparison plots between original videos and their corresponding predictions.

    Parameters:
        predictions (dict): A dictionary containing predicted frames, indexed by video names.
        videos (dict): A dictionary containing original video frames, indexed by video names.

    This function iterates through each video key, retrieves frames from both the original and predicted data,
    and plots them side by side for visual comparison. It supports different intervals for videos with a large
    number of frames, adjusting the subplot layout accordingly.
    """
    print("Calculating losses for plots...")

    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        import os
        os.makedirs(save_path_prefix, exist_ok=True)

    total_ssim, total_mse = [], []
    all_ssim = {'all': None}
    all_mse = {'all': None}

    for key in videos.keys():
        prediction = predictions[key]
        if temporal:
            video = videos[key] #since in preparing temporal data we already took the middle frame
        else:
            video = videos[key][..., 15]
        N = video.shape[0]



        ssim_frames = np.array([
            ssim(torch.from_numpy(video[i]).unsqueeze(0), torch.from_numpy(prediction[i]).unsqueeze(0), data_range=1, size_average=True).item()
            for i in range(N) 
        ])
        mean_ssim_frames = np.mean(ssim_frames)
        median_ssim_frames = np.median(ssim_frames)
        std_ssim_frames = np.std(ssim_frames)
        all_ssim[key] = ssim_frames

        mse_frames = np.array([
            F.mse_loss(normalize(torch.from_numpy(video[i])), normalize(torch.from_numpy(prediction[i]))).numpy()
            for i in range(N) 
        ])
        mean_mse_frames = np.mean(mse_frames)
        median_mse_frames = np.median(mse_frames)
        std_mse_frames = np.std(mse_frames)
        all_mse[key] = mse_frames

        if psim_and_tv:
            from Video_dataset.models_new_2 import VGG16Features
            def short_psim(prediction, label):
                """
                Calculate perceptual similarity between prediction and video.
                """
                #print("prediction[0] shape =", prediction[0].shape)
                #print("label[0] shape =", label[0].shape)
                label_features = VGG16Features().eval()(label[0])
                prediction_features = VGG16Features().eval()(prediction[0])
                
        
                c = []
                for a, b in zip(prediction_features, label_features):
                    cos = (1 - F.cosine_similarity(a, b, dim=1)).mean()
                    c.append(cos)
                loss = sum(c)
                return loss
            #new code (psim and tv)
            
            psim_frames = np.array([
                short_psim(prediction=torch.from_numpy(video[i]).unsqueeze(0), label=torch.from_numpy(prediction[i]).unsqueeze(0))
                for i in range(N) 
            ])
            mean_psim_frames = np.mean(psim_frames)
            median_psim_frames = np.median(psim_frames)
            std_psim_frames = np.std(psim_frames)
            #all_psim[key] = psim_frames

            def short_tv(prediction):
                from imports import TotalVariation
                N, C, H, W = prediction[0].shape
                tv = TotalVariation()
                loss = tv(prediction[0]) / (N*C*H*W)
                
                #loss = TotalVariation(prediction[0]) / (N*C*H*W)
                return loss
            
            tv_frames = np.array([
                short_tv(torch.from_numpy(prediction[i]).unsqueeze(0))
                for i in range(N) 
            ])
            mean_tv_frames = np.mean(tv_frames)
            median_tv_frames = np.median(tv_frames)
            std_tv_frames = np.std(tv_frames)
            #all_psim[key] = psim_frames

        total_ssim = np.concatenate((total_ssim, ssim_frames))
        total_mse = np.concatenate((total_mse, mse_frames))

        if display_plots:
            print(f'### {key} ({N}) ###')
            plt.figure(figsize=(17, 3))
            plt.subplot(121)
            plt.hist(ssim_frames, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
            sns.kdeplot(ssim_frames, fill=True, color='b', label='KDE')
            plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
            plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
            plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
            plt.xlabel('SSIM')
            plt.ylabel('Density')
            plt.title(f'Distribution of Structural Similarity per Frame (N = {N})')
            plt.legend() 
            plt.subplot(122)
            plt.hist(mse_frames, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
            sns.kdeplot(mse_frames, fill=True, color='r', label='KDE')
            plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
            plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
            plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
            plt.xlabel('MSE')
            plt.ylabel('Density')
            plt.title(f'Distribution of MSE per Frame (N = {N})')
            plt.legend() 
            # Save the loss histogram plot if requested
            if save_plots and save_path_prefix:
                #since im not storing models here anymore now i need to put the path to other place in model_name, but i dont want that in image name
                model_name = model_name.split("/")[-1]

                losses_path = f"{save_path_prefix}losses_{key}_{model_name}.png"
                plt.savefig(losses_path, bbox_inches='tight', dpi=300)
#            plt.show()
            plt.show()

            if psim_and_tv:
            #new code
                print(f'### {key} ({N}) ###')
                plt.figure(figsize=(17, 3))
                plt.subplot(121)
                plt.hist(psim_frames, bins=30, alpha=0.5, color='green', density=True, label='Histogram')
                sns.kdeplot(psim_frames, fill=True, color='b', label='KDE')
                plt.axvline(x=median_psim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_psim_frames:.3f}')
                plt.axvline(x=mean_psim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_psim_frames:.3f}')
                plt.axvspan(mean_psim_frames - std_psim_frames, mean_psim_frames + std_psim_frames, color='red', alpha=0.05, label=f'std: {std_psim_frames:.3f}')
                plt.xlabel('PSIM')
                plt.ylabel('Density')
                plt.title(f'Distribution of Perceptual Similarity per Frame (N = {N})')
                plt.legend() 
                plt.subplot(122)
                plt.hist(tv_frames, bins=30, alpha=0.5, color='purple', density=True, label='Histogram')
                sns.kdeplot(tv_frames, fill=True, color='r', label='KDE')
                plt.axvline(x=median_tv_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_tv_frames:.3f}')
                plt.axvline(x=mean_tv_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_tv_frames:.3f}')
                plt.axvspan(mean_tv_frames - std_tv_frames, mean_tv_frames + std_tv_frames, color='red', alpha=0.05, label=f'std: {std_tv_frames:.3f}')
                plt.xlabel('TV')
                plt.ylabel('Density')
                plt.title(f'Distribution of TV per Frame (N = {N})')
                plt.legend() 
                # Save the loss histogram plot if requested
                if save_plots and save_path_prefix:
                    losses_path = f"{save_path_prefix}losses_{key}_{model_name}.png"
                    plt.savefig(losses_path, bbox_inches='tight', dpi=300)
    #            plt.show()
                plt.show()


            #new code 20/5/2025 (not useful anymore)
            '''if individual_metrics and key in individual_metrics:
                print("individual_metrics[key] dict tree:")
                print_dict_tree(individual_metrics[key])
                for metric_name, values in individual_metrics[key].items():
                    if metric_name not in ['ssim', 'mse']:  # Skip these as they're already plotted
                        plt.figure(figsize=(10, 3))
                        
                        print("values shape", values.shape)
                        plt.hist(values, bins=30, alpha=0.5, color='green', density=True, label='Histogram')
                        
                        mean_val = np.mean(values)
                        median_val = np.median(values)
                        std_val = np.std(values)
                        
                        plt.axvline(x=median_val, color='r', linestyle=':', linewidth=1, label=f'Median: {median_val:.3f}')
                        plt.axvline(x=mean_val, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_val:.3f}')
                        plt.axvspan(mean_val - std_val, mean_val + std_val, color='red', alpha=0.05, label=f'std: {std_val:.3f}')
                        
                        plt.xlabel(metric_name)
                        plt.ylabel('Density')
                        plt.title(f'Distribution of {metric_name} per Frame (N = {len(values)})')
                        plt.legend()
                        
                        if save_plots and save_path_prefix:
                            metric_path = f"{save_path_prefix}{metric_name}_{key}_{model_name}.png"
                            plt.savefig(metric_path, bbox_inches='tight', dpi=300)
                        plt.show()'''


            
            if prediction.shape[0] > 400:
                indices = np.arange(0, prediction.shape[0], 15)
            else:
                indices = np.arange(0, prediction.shape[0], 3)
            print(f'Display {len(indices)} frames:', indices)
            col = int(indices.shape[0] / 3) + 1
            k = 0
    
            if len(indices) < 20:
                plt.figure(figsize=(8, 10))
            elif len(indices) >= 20 and len(indices) < 40:
                plt.figure(figsize=(10, 15))
            else:
                plt.figure(figsize=(15, 20))
                
            for i in indices:
                k += 1
                plt.subplot(col,6,k)
                if temporal:
                    middle_time = video[i].shape[1] // 2
                    plt.imshow(np.transpose(normalize(video[i][middle_time, :, :, :]), (1, 2, 0)))
                else:
                    plt.imshow(np.transpose(normalize(video[i]), (1, 2, 0)))
                plt.axis('off')
                k += 1
                plt.subplot(col,6,k)
                if temporal:
                    middle_time = prediction[i].shape[1] // 2
                    plt.imshow(np.transpose(normalize(prediction[i][middle_time, :, :, :]), (1, 2, 0)))
                else:
                    plt.imshow(np.transpose(normalize(prediction[i]), (1, 2, 0)))
                plt.axis('off')

            # Save the image comparison plot if requested
            if save_plots and save_path_prefix:
                img_path = f"{save_path_prefix}{key}_{model_name}.png"
                plt.savefig(img_path, bbox_inches='tight', dpi=300)
            plt.show()
            del prediction, video, indices
#            plt.show()
#            del prediction, video, indices

    mean_ssim_frames = np.mean(total_ssim)
    median_ssim_frames = np.median(total_ssim)
    std_ssim_frames = np.std(total_ssim)

    mean_mse_frames = np.mean(total_mse)
    median_mse_frames = np.median(total_mse)
    std_mse_frames = np.std(total_mse)

    all_ssim['all'] = total_ssim
    all_mse['all'] = total_mse

    if display_plots:
        print(f'### all ({total_ssim.shape[0]}) ###')
        plt.figure(figsize=(17, 3))
        plt.subplot(121)
        plt.hist(total_ssim, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_ssim, fill=True, color='b', label='KDE')
        plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
        plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
        plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
        plt.xlabel('SSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Structural Similarity per Frame (N = {total_ssim.shape[0]})')
        plt.legend() 
        plt.subplot(122)
        plt.hist(total_mse, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(total_mse, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
        plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
        plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE per Frame (N = {total_mse.shape[0]})')
        plt.legend()
        # Save the all-videos metrics plot if requested
        if save_plots and save_path_prefix:
            all_path = f"{save_path_prefix}losses_all_{model_name}.png"
            plt.savefig(all_path, bbox_inches='tight', dpi=300)
#        plt.show()
        plt.show()

    if performance_dict:
        performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_ssim, 'SSIM', display_plots)
        performance_dict['mean_mse_D'], performance_dict['median_mse_D'] = plot_scatter_metrics(all_mse, 'MSE', display_plots)
        return performance_dict



def plot_all_predictions(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, save_path_prefix=None, model_name=""):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version includes extensive debugging and a more direct approach to handling multiple inputs.
    """
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        import os
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # If we don't have common keys, we need to find a match
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        # Take the first video key as reference (assuming all predictions relate to same video data)
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        
        # Map all prediction keys to this video key
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        # Create a mapping from prediction keys to video keys
        key_mapping = {}
        for pred_key in predictions.keys():
            # If the key exists directly in videos, use it
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                # Otherwise, try to find a matching video key
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                
                # If no match found and we have videos, use the first video key
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Initialize metrics
    total_ssim, total_mse = [], []
    all_ssim = {'all': None}
    all_mse = {'all': None}
    
    # ===== PART 1: Calculate metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        N = video.shape[0]
        
        # Calculate SSIM
        try:
            ssim_frames = np.array([
                ssim(torch.from_numpy(video[i]).unsqueeze(0), 
                     torch.from_numpy(prediction[i]).unsqueeze(0), 
                     data_range=1, size_average=True).item()
                for i in range(N) 
            ])
            mean_ssim = np.mean(ssim_frames)
            print(f"SSIM for {pred_key}: Mean = {mean_ssim:.3f}")
            all_ssim[pred_key] = ssim_frames
            total_ssim = np.concatenate((total_ssim, ssim_frames))
        except Exception as e:
            print(f"Error calculating SSIM for {pred_key}: {e}")
            ssim_frames = np.zeros(N)
            all_ssim[pred_key] = ssim_frames
        
        # Calculate MSE
        try:
            mse_frames = np.array([
                F.mse_loss(normalize(torch.from_numpy(video[i])), 
                          normalize(torch.from_numpy(prediction[i]))).numpy()
                for i in range(N) 
            ])
            mean_mse = np.mean(mse_frames)
            print(f"MSE for {pred_key}: Mean = {mean_mse:.3f}")
            all_mse[pred_key] = mse_frames
            total_mse = np.concatenate((total_mse, mse_frames))
        except Exception as e:
            print(f"Error calculating MSE for {pred_key}: {e}")
            mse_frames = np.zeros(N)
            all_mse[pred_key] = mse_frames
    
    # ===== PART 2: Plot metrics for each prediction =====
    if display_plots and total_ssim.size > 0:
        # Plot metrics across all predictions
        plt.figure(figsize=(15, 5))
        
        # Plot SSIM
        plt.subplot(1, 2, 1)
        for pred_key in key_mapping.keys():
            if pred_key in all_ssim:
                plt.plot(all_ssim[pred_key], label=f"{pred_key} (mean={np.mean(all_ssim[pred_key]):.3f})")
        plt.xlabel("Frame Index")
        plt.ylabel("SSIM")
        plt.title("SSIM Comparison")
        plt.legend()
        
        # Plot MSE
        plt.subplot(1, 2, 2)
        for pred_key in key_mapping.keys():
            if pred_key in all_mse:
                plt.plot(all_mse[pred_key], label=f"{pred_key} (mean={np.mean(all_mse[pred_key]):.3f})")
        plt.xlabel("Frame Index")
        plt.ylabel("MSE")
        plt.title("MSE Comparison")
        plt.legend()
        
        plt.tight_layout()
        
        # Save metrics plot if requested
        if save_plots and save_path_prefix:
            metrics_path = f"{save_path_prefix}metrics_comparison_{model_name}.png"
            plt.savefig(metrics_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 3: Plot frame comparisons =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # For large videos, sample frames
        if N > 10:
            indices = np.linspace(0, N-1, 10, dtype=int)
        else:
            indices = np.arange(N)
        
        print(f"Plotting {len(indices)} sample frames")
        
        # Number of rows and columns
        n_rows = len(indices)
        n_cols = 1 + len(key_mapping)  # Original + each prediction
        
        # Create a figure
        plt.figure(figsize=(4 * n_cols, 4 * n_rows))
        
        # Plot each frame
        for i, frame_idx in enumerate(indices):
            # Plot original frame
            ax = plt.subplot(n_rows, n_cols, i * n_cols + 1)
            plt.imshow(np.transpose(normalize(ref_video[frame_idx]), (1, 2, 0)))
            plt.title(f"Original Frame {frame_idx}")
            plt.axis('off')
            
            # Plot each prediction
            for j, (pred_key, video_key) in enumerate(key_mapping.items()):
                try:
                    pred_frame = predictions[pred_key][frame_idx]
                    ax = plt.subplot(n_rows, n_cols, i * n_cols + j + 2)
                    plt.imshow(np.transpose(normalize(pred_frame), (1, 2, 0)))
                    
                    # Add metrics to title if available
                    ssim_val = all_ssim[pred_key][frame_idx] if pred_key in all_ssim else 0
                    mse_val = all_mse[pred_key][frame_idx] if pred_key in all_mse else 0
                    plt.title(f"{pred_key}\nSSIM: {ssim_val:.3f}, MSE: {mse_val:.3f}")
                    plt.axis('off')
                except Exception as e:
                    print(f"Error plotting prediction {pred_key} for frame {frame_idx}: {e}")
        
        plt.tight_layout()
        
        # Save frames comparison plot if requested
        if save_plots and save_path_prefix:
            frames_path = f"{save_path_prefix}frames_comparison_{model_name}.png"
            plt.savefig(frames_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 4: Plot overall distributions =====
    if display_plots and total_ssim.size > 0:
        # Calculate overall metrics
        mean_ssim_frames = np.mean(total_ssim)
        median_ssim_frames = np.median(total_ssim)
        std_ssim_frames = np.std(total_ssim)
        
        mean_mse_frames = np.mean(total_mse)
        median_mse_frames = np.median(total_mse)
        std_mse_frames = np.std(total_mse)
        
        all_ssim['all'] = total_ssim
        all_mse['all'] = total_mse
        
        # Plot overall distributions
        plt.figure(figsize=(15, 5))
        
        # SSIM distribution
        plt.subplot(1, 2, 1)
        plt.hist(total_ssim, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_ssim, fill=True, color='b', label='KDE')
        plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
        plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
        plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
        plt.xlabel('SSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Structural Similarity (N = {total_ssim.shape[0]})')
        plt.legend()
        
        # MSE distribution
        plt.subplot(1, 2, 2)
        plt.hist(total_mse, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(total_mse, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
        plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
        plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE (N = {total_mse.shape[0]})')
        plt.legend()
        
        plt.tight_layout()
        
        # Save distribution plot if requested
        if save_plots and save_path_prefix:
            dist_path = f"{save_path_prefix}distributions_{model_name}.png"
            plt.savefig(dist_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # Update performance dictionary if provided
    if performance_dict:
        try:
            performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_ssim, 'SSIM', display_plots)
            performance_dict['mean_mse_D'], performance_dict['median_mse_D'] = plot_scatter_metrics(all_mse, 'MSE', display_plots)
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    return performance_dict




def plot_all_predictions2(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, save_path_prefix=None, model_name="", device="cuda" if torch.cuda.is_available() else "cpu"):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version includes extensive debugging and a more direct approach to handling multiple inputs.
    Uses perceptual similarity (PSIM) instead of MSE.
    """
    # Define perceptual similarity loss function here
    def perceptual_sim_loss(prediction, label):
        # Extract VGG features and compute loss (example using one layer)
        prediction_features = vgg16_features(prediction)
        label_features = vgg16_features(label)
    
        c = []
        for a, b in zip(prediction_features, label_features):
            cos = (1 - F.cosine_similarity(a, b, dim=1)).mean()
            c.append(cos)
        loss = sum(c)
        return loss
    
    # Define VGG16 features extraction function
    def vgg16_features(x):
        """
        Extract features from pretrained VGG16 model
        """
        # Import required modules for VGG16
        from torchvision import models
        import torch.nn as nn
        
        # Initialize the model if it doesn't exist
        if not hasattr(vgg16_features, 'model'):
            # Load pretrained VGG16 model
            vgg16_features.model = models.vgg16(pretrained=True).features.eval().to(device)
            
            # Freeze parameters
            for param in vgg16_features.model.parameters():
                param.requires_grad = False
                
            # Define which layers to extract features from (common choices for perceptual loss)
            vgg16_features.layers = [3, 8, 15, 22]  # After ReLU layers in VGG16
        
        # Process the input
        features = []
        x = x.to(device).float()  # Ensure input is float and on correct device
        
        # Get features from specified layers
        for i, layer in enumerate(vgg16_features.model):
            x = layer(x)
            if i in vgg16_features.layers:
                features.append(x)
                
        return features
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        import os
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # If we don't have common keys, we need to find a match
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        # Take the first video key as reference (assuming all predictions relate to same video data)
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        
        # Map all prediction keys to this video key
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        # Create a mapping from prediction keys to video keys
        key_mapping = {}
        for pred_key in predictions.keys():
            # If the key exists directly in videos, use it
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                # Otherwise, try to find a matching video key
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                
                # If no match found and we have videos, use the first video key
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Initialize metrics
    total_ssim, total_psim = [], []
    all_ssim = {'all': None}
    all_psim = {'all': None}
    
    # ===== PART 1: Calculate metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        N = video.shape[0]
        
        # Calculate SSIM
        try:
            ssim_frames = np.array([
                ssim(torch.from_numpy(video[i]).unsqueeze(0), 
                     torch.from_numpy(prediction[i]).unsqueeze(0), 
                     data_range=1, size_average=True).item()
                for i in range(N) 
            ])
            mean_ssim = np.mean(ssim_frames)
            print(f"SSIM for {pred_key}: Mean = {mean_ssim:.3f}")
            all_ssim[pred_key] = ssim_frames
            total_ssim = np.concatenate((total_ssim, ssim_frames))
        except Exception as e:
            print(f"Error calculating SSIM for {pred_key}: {e}")
            ssim_frames = np.zeros(N)
            all_ssim[pred_key] = ssim_frames
        
        # Calculate PSIM (Perceptual Similarity) instead of MSE
        try:
            # Make sure inputs are properly normalized for VGG
            psim_frames = np.array([
                perceptual_sim_loss(
                    normalize(torch.from_numpy(prediction[i]).unsqueeze(0)), 
                    normalize(torch.from_numpy(video[i]).unsqueeze(0))
                ).item()
                for i in range(N) 
            ])
            mean_psim = np.mean(psim_frames)
            print(f"PSIM for {pred_key}: Mean = {mean_psim:.3f}")
            all_psim[pred_key] = psim_frames
            total_psim = np.concatenate((total_psim, psim_frames))
        except Exception as e:
            print(f"Error calculating PSIM for {pred_key}: {e}")
            psim_frames = np.zeros(N)
            all_psim[pred_key] = psim_frames
    
    # ===== PART 2: Plot metrics for each prediction =====
    if display_plots and total_ssim.size > 0:
        # Plot metrics across all predictions
        plt.figure(figsize=(15, 5))
        
        # Plot SSIM
        plt.subplot(1, 2, 1)
        for pred_key in key_mapping.keys():
            if pred_key in all_ssim:
                plt.plot(all_ssim[pred_key], label=f"{pred_key} (mean={np.mean(all_ssim[pred_key]):.3f})")
        plt.xlabel("Frame Index")
        plt.ylabel("SSIM")
        plt.title("SSIM Comparison")
        plt.legend()
        
        # Plot PSIM
        plt.subplot(1, 2, 2)
        for pred_key in key_mapping.keys():
            if pred_key in all_psim:
                plt.plot(all_psim[pred_key], label=f"{pred_key} (mean={np.mean(all_psim[pred_key]):.3f})")
        plt.xlabel("Frame Index")
        plt.ylabel("PSIM")
        plt.title("Perceptual Similarity Comparison")
        plt.legend()
        
        plt.tight_layout()
        
        # Save metrics plot if requested
        if save_plots and save_path_prefix:
            metrics_path = f"{save_path_prefix}metrics_comparison_{model_name}.png"
            plt.savefig(metrics_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 3: Plot frame comparisons =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # For large videos, sample frames
        if N > 10:
            indices = np.linspace(0, N-1, 10, dtype=int)
        else:
            indices = np.arange(N)
        
        print(f"Plotting {len(indices)} sample frames")
        
        # Number of rows and columns
        n_rows = len(indices)
        n_cols = 1 + len(key_mapping)  # Original + each prediction
        
        # Create a figure
        plt.figure(figsize=(4 * n_cols, 4 * n_rows))
        
        # Plot each frame
        for i, frame_idx in enumerate(indices):
            # Plot original frame
            ax = plt.subplot(n_rows, n_cols, i * n_cols + 1)
            plt.imshow(np.transpose(normalize(ref_video[frame_idx]), (1, 2, 0)))
            plt.title(f"Original Frame {frame_idx}")
            plt.axis('off')
            
            # Plot each prediction
            for j, (pred_key, video_key) in enumerate(key_mapping.items()):
                try:
                    pred_frame = predictions[pred_key][frame_idx]
                    ax = plt.subplot(n_rows, n_cols, i * n_cols + j + 2)
                    plt.imshow(np.transpose(normalize(pred_frame), (1, 2, 0)))
                    
                    # Add metrics to title if available
                    ssim_val = all_ssim[pred_key][frame_idx] if pred_key in all_ssim else 0
                    psim_val = all_psim[pred_key][frame_idx] if pred_key in all_psim else 0
                    plt.title(f"{pred_key}\nSSIM: {ssim_val:.3f}, PSIM: {psim_val:.3f}")
                    plt.axis('off')
                except Exception as e:
                    print(f"Error plotting prediction {pred_key} for frame {frame_idx}: {e}")
        
        plt.tight_layout()
        
        # Save frames comparison plot if requested
        if save_plots and save_path_prefix:
            frames_path = f"{save_path_prefix}frames_comparison_{model_name}.png"
            plt.savefig(frames_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 4: Plot overall distributions =====
    if display_plots and total_ssim.size > 0:
        # Calculate overall metrics
        mean_ssim_frames = np.mean(total_ssim)
        median_ssim_frames = np.median(total_ssim)
        std_ssim_frames = np.std(total_ssim)
        
        mean_psim_frames = np.mean(total_psim)
        median_psim_frames = np.median(total_psim)
        std_psim_frames = np.std(total_psim)
        
        all_ssim['all'] = total_ssim
        all_psim['all'] = total_psim
        
        # Plot overall distributions
        plt.figure(figsize=(15, 5))
        
        # SSIM distribution
        plt.subplot(1, 2, 1)
        plt.hist(total_ssim, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_ssim, fill=True, color='b', label='KDE')
        plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
        plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
        plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
        plt.xlabel('SSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Structural Similarity (N = {total_ssim.shape[0]})')
        plt.legend()
        
        # PSIM distribution
        plt.subplot(1, 2, 2)
        plt.hist(total_psim, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(total_psim, fill=True, color='r', label='KDE')
        plt.axvline(x=median_psim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_psim_frames:.3f}')
        plt.axvline(x=mean_psim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_psim_frames:.3f}')
        plt.axvspan(mean_psim_frames - std_psim_frames, mean_psim_frames + std_psim_frames, color='red', alpha=0.05, label=f'std: {std_psim_frames:.3f}')
        plt.xlabel('PSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Perceptual Similarity (N = {total_psim.shape[0]})')
        plt.legend()
        
        plt.tight_layout()
        
        # Save distribution plot if requested
        if save_plots and save_path_prefix:
            dist_path = f"{save_path_prefix}distributions_{model_name}.png"
            plt.savefig(dist_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # Update performance dictionary if provided
    if performance_dict:
        try:
            performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_ssim, 'SSIM', display_plots)
            performance_dict['mean_psim_D'], performance_dict['median_psim_D'] = plot_scatter_metrics(all_psim, 'PSIM', display_plots)
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    return performance_dict



def plot_all_predictions3(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, save_path_prefix=None, model_name="", device="cuda" if torch.cuda.is_available() else "cpu"):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version includes extensive debugging and a more direct approach to handling multiple inputs.
    Uses perceptual similarity (PSIM) instead of MSE.
    """
    # Define perceptual similarity loss function here
    def perceptual_sim_loss(prediction, label):
        # Extract VGG features and compute loss (example using one layer)
        prediction_features = vgg16_features(prediction)
        label_features = vgg16_features(label)
    
        c = []
        for a, b in zip(prediction_features, label_features):
            cos = (1 - F.cosine_similarity(a, b, dim=1)).mean()
            c.append(cos)
        loss = sum(c)
        return loss
    
    # Define VGG16 features extraction function
    def vgg16_features(x):
        """
        Extract features from pretrained VGG16 model
        """
        # Import required modules for VGG16
        from torchvision import models
        import torch.nn as nn
        
        # Initialize the model if it doesn't exist
        if not hasattr(vgg16_features, 'model'):
            # Load pretrained VGG16 model
            vgg16_features.model = models.vgg16(pretrained=True).features.eval().to(device)
            
            # Freeze parameters
            for param in vgg16_features.model.parameters():
                param.requires_grad = False
                
            # Define which layers to extract features from (common choices for perceptual loss)
            vgg16_features.layers = [3, 8, 15, 22]  # After ReLU layers in VGG16
        
        # Process the input
        features = []
        x = x.to(device).float()  # Ensure input is float and on correct device
        
        # Get features from specified layers
        for i, layer in enumerate(vgg16_features.model):
            x = layer(x)
            if i in vgg16_features.layers:
                features.append(x)
                
        return features
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        import os
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # If we don't have common keys, we need to find a match
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        # Take the first video key as reference (assuming all predictions relate to same video data)
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        
        # Map all prediction keys to this video key
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        # Create a mapping from prediction keys to video keys
        key_mapping = {}
        for pred_key in predictions.keys():
            # If the key exists directly in videos, use it
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                # Otherwise, try to find a matching video key
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                
                # If no match found and we have videos, use the first video key
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Initialize metrics
    total_ssim, total_psim = [], []
    all_ssim = {'all': None}
    all_psim = {'all': None}
    
    # ===== PART 1: Calculate metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        N = video.shape[0]
        
        # Calculate SSIM
        try:
            ssim_frames = np.array([
                ssim(torch.from_numpy(video[i]).unsqueeze(0), 
                     torch.from_numpy(prediction[i]).unsqueeze(0), 
                     data_range=1, size_average=True).item()
                for i in range(N) 
            ])
            mean_ssim = np.mean(ssim_frames)
            print(f"SSIM for {pred_key}: Mean = {mean_ssim:.3f}")
            all_ssim[pred_key] = ssim_frames
            total_ssim = np.concatenate((total_ssim, ssim_frames))
        except Exception as e:
            print(f"Error calculating SSIM for {pred_key}: {e}")
            ssim_frames = np.zeros(N)
            all_ssim[pred_key] = ssim_frames
        
        # Calculate PSIM (Perceptual Similarity) instead of MSE
        try:
            # Make sure inputs are properly normalized for VGG
            psim_frames = np.array([
                perceptual_sim_loss(
                    normalize(torch.from_numpy(prediction[i]).unsqueeze(0)), 
                    normalize(torch.from_numpy(video[i]).unsqueeze(0))
                ).item()
                for i in range(N) 
            ])
            mean_psim = np.mean(psim_frames)
            print(f"PSIM for {pred_key}: Mean = {mean_psim:.3f}")
            all_psim[pred_key] = psim_frames
            total_psim = np.concatenate((total_psim, psim_frames))
        except Exception as e:
            print(f"Error calculating PSIM for {pred_key}: {e}")
            psim_frames = np.zeros(N)
            all_psim[pred_key] = psim_frames
    
    # ===== PART 2: Plot metrics for each prediction =====
    if display_plots and total_ssim.size > 0:
        # Plot metrics across all predictions
        plt.figure(figsize=(15, 5))
        
        # Plot SSIM
        plt.subplot(1, 2, 1)
        for pred_key in key_mapping.keys():
            if pred_key in all_ssim:
                plt.plot(all_ssim[pred_key], label=f"{pred_key} (mean={np.mean(all_ssim[pred_key]):.3f})")
        plt.xlabel("Frame Index")
        plt.ylabel("SSIM")
        plt.title("SSIM Comparison")
        plt.legend()
        
        # Plot PSIM
        plt.subplot(1, 2, 2)
        for pred_key in key_mapping.keys():
            if pred_key in all_psim:
                plt.plot(all_psim[pred_key], label=f"{pred_key} (mean={np.mean(all_psim[pred_key]):.3f})")
        plt.xlabel("Frame Index")
        plt.ylabel("PSIM")
        plt.title("Perceptual Similarity Comparison")
        plt.legend()
        
        plt.tight_layout()
        
        # Save metrics plot if requested
        if save_plots and save_path_prefix:
            metrics_path = f"{save_path_prefix}metrics_comparison_{model_name}.png"
            plt.savefig(metrics_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 3: Plot frame comparisons =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # For large videos, sample frames
        if N > 10:
            indices = np.linspace(0, N-1, 10, dtype=int)
        else:
            indices = np.arange(N)
        
        print(f"Plotting {len(indices)} sample frames")
        
        # Number of rows and columns
        n_rows = len(indices)
        n_cols = 1 + len(key_mapping)  # Original + each prediction
        
        # Create a figure
        plt.figure(figsize=(4 * n_cols, 4 * n_rows))
        
        # REORDERING: Create an ordered list of prediction keys with "original_combined" as the second column
        ordered_keys = []
        for key in key_mapping.keys():
            if key != "original_combined":
                ordered_keys.append(key)
        
        # If "original_combined" exists, insert it at the beginning of the list
        if "original_combined" in key_mapping:
            ordered_keys.insert(0, "original_combined")
        
        # Plot each frame
        for i, frame_idx in enumerate(indices):
            # Plot original frame
            ax = plt.subplot(n_rows, n_cols, i * n_cols + 1)
            plt.imshow(np.transpose(normalize(ref_video[frame_idx]), (1, 2, 0)))
            plt.title(f"Original Frame {frame_idx}")
            plt.axis('off')
            
            # Plot each prediction in the reordered sequence
            for j, pred_key in enumerate(ordered_keys):
                try:
                    video_key = key_mapping[pred_key]
                    pred_frame = predictions[pred_key][frame_idx]
                    ax = plt.subplot(n_rows, n_cols, i * n_cols + j + 2)
                    plt.imshow(np.transpose(normalize(pred_frame), (1, 2, 0)))
                    
                    # Add metrics to title if available
                    ssim_val = all_ssim[pred_key][frame_idx] if pred_key in all_ssim else 0
                    psim_val = all_psim[pred_key][frame_idx] if pred_key in all_psim else 0
                    plt.title(f"{pred_key}\nSSIM: {ssim_val:.3f}, PSIM: {psim_val:.3f}")
                    plt.axis('off')
                except Exception as e:
                    print(f"Error plotting prediction {pred_key} for frame {frame_idx}: {e}")
        
        plt.tight_layout()
        
        # Save frames comparison plot if requested
        if save_plots and save_path_prefix:
            frames_path = f"{save_path_prefix}frames_comparison_{model_name}.png"
            plt.savefig(frames_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 4: Plot overall distributions =====
    if display_plots and total_ssim.size > 0:
        # Calculate overall metrics
        mean_ssim_frames = np.mean(total_ssim)
        median_ssim_frames = np.median(total_ssim)
        std_ssim_frames = np.std(total_ssim)
        
        mean_psim_frames = np.mean(total_psim)
        median_psim_frames = np.median(total_psim)
        std_psim_frames = np.std(total_psim)
        
        all_ssim['all'] = total_ssim
        all_psim['all'] = total_psim
        
        # Plot overall distributions
        plt.figure(figsize=(15, 5))
        
        # SSIM distribution
        plt.subplot(1, 2, 1)
        plt.hist(total_ssim, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_ssim, fill=True, color='b', label='KDE')
        plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
        plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
        plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
        plt.xlabel('SSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Structural Similarity (N = {total_ssim.shape[0]})')
        plt.legend()
        
        # PSIM distribution
        plt.subplot(1, 2, 2)
        plt.hist(total_psim, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(total_psim, fill=True, color='r', label='KDE')
        plt.axvline(x=median_psim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_psim_frames:.3f}')
        plt.axvline(x=mean_psim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_psim_frames:.3f}')
        plt.axvspan(mean_psim_frames - std_psim_frames, mean_psim_frames + std_psim_frames, color='red', alpha=0.05, label=f'std: {std_psim_frames:.3f}')
        plt.xlabel('PSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Perceptual Similarity (N = {total_psim.shape[0]})')
        plt.legend()
        
        plt.tight_layout()
        
        # Save distribution plot if requested
        if save_plots and save_path_prefix:
            dist_path = f"{save_path_prefix}distributions_{model_name}.png"
            plt.savefig(dist_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # Update performance dictionary if provided
    if performance_dict:
        try:
            performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_ssim, 'SSIM', display_plots)
            performance_dict['mean_psim_D'], performance_dict['median_psim_D'] = plot_scatter_metrics(all_psim, 'PSIM', display_plots)
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    return performance_dict




def plot_all_predictions4(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, save_path_prefix=None, model_name="", device="cuda" if torch.cuda.is_available() else "cpu"):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version includes extensive debugging and a more direct approach to handling multiple inputs.
    Uses only SSIM for metrics.
    """
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        import os
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # If we don't have common keys, we need to find a match
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        # Take the first video key as reference (assuming all predictions relate to same video data)
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        
        # Map all prediction keys to this video key
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        # Create a mapping from prediction keys to video keys
        key_mapping = {}
        for pred_key in predictions.keys():
            # If the key exists directly in videos, use it
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                # Otherwise, try to find a matching video key
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                
                # If no match found and we have videos, use the first video key
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Initialize metrics
    total_ssim = []
    all_ssim = {'all': None}
    
    # ===== PART 1: Calculate metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        N = video.shape[0]
        
        # Calculate SSIM
        try:
            ssim_frames = np.array([
                ssim(torch.from_numpy(video[i]).unsqueeze(0), 
                     torch.from_numpy(prediction[i]).unsqueeze(0), 
                     data_range=1, size_average=True).item()
                for i in range(N) 
            ])
            mean_ssim = np.mean(ssim_frames)
            print(f"SSIM for {pred_key}: Mean = {mean_ssim:.3f}")
            all_ssim[pred_key] = ssim_frames
            total_ssim = np.concatenate((total_ssim, ssim_frames))
        except Exception as e:
            print(f"Error calculating SSIM for {pred_key}: {e}")
            ssim_frames = np.zeros(N)
            all_ssim[pred_key] = ssim_frames
    
    # ===== PART 2: Plot metrics for each prediction =====
    if display_plots and total_ssim.size > 0:
        # Plot metrics across all predictions
        plt.figure(figsize=(10, 5))
        
        # Plot SSIM
        for pred_key in key_mapping.keys():
            if pred_key in all_ssim:
                plt.plot(all_ssim[pred_key], label=f"{pred_key} (mean={np.mean(all_ssim[pred_key]):.3f})")
        plt.xlabel("Frame Index")
        plt.ylabel("SSIM")
        plt.title("SSIM Comparison")
        plt.legend()
        
        plt.tight_layout()
        
        # Save metrics plot if requested
        if save_plots and save_path_prefix:
            metrics_path = f"{save_path_prefix}metrics_comparison_{model_name}.png"
            plt.savefig(metrics_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 3: Plot frame comparisons =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # For large videos, sample frames
        if N > 10:
            indices = np.linspace(0, N-1, 10, dtype=int)
        else:
            indices = np.arange(N)
        
        print(f"Plotting {len(indices)} sample frames")
        
        # Number of rows and columns
        n_rows = len(indices)
        n_cols = 1 + len(key_mapping)  # Original + each prediction
        
        # Create a figure
        plt.figure(figsize=(4 * n_cols, 4 * n_rows))
        
        # REORDERING: Create an ordered list of prediction keys with "original_combined" as the second column
        ordered_keys = []
        for key in key_mapping.keys():
            if key != "original_combined":
                ordered_keys.append(key)
        
        # If "original_combined" exists, insert it at the beginning of the list
        if "original_combined" in key_mapping:
            ordered_keys.insert(0, "original_combined")
        
        # Plot each frame
        for i, frame_idx in enumerate(indices):
            # Plot original frame
            ax = plt.subplot(n_rows, n_cols, i * n_cols + 1)
            plt.imshow(np.transpose(normalize(ref_video[frame_idx]), (1, 2, 0)))
            plt.title(f"Original Frame {frame_idx}")
            plt.axis('off')
            
            # Plot each prediction in the reordered sequence
            for j, pred_key in enumerate(ordered_keys):
                try:
                    video_key = key_mapping[pred_key]
                    pred_frame = predictions[pred_key][frame_idx]
                    ax = plt.subplot(n_rows, n_cols, i * n_cols + j + 2)
                    plt.imshow(np.transpose(normalize(pred_frame), (1, 2, 0)))
                    
                    # Add metrics to title if available
                    ssim_val = all_ssim[pred_key][frame_idx] if pred_key in all_ssim else 0
                    plt.title(f"{pred_key}\nSSIM: {ssim_val:.3f}")
                    plt.axis('off')
                except Exception as e:
                    print(f"Error plotting prediction {pred_key} for frame {frame_idx}: {e}")
        
        plt.tight_layout()
        
        # Save frames comparison plot if requested
        if save_plots and save_path_prefix:
            frames_path = f"{save_path_prefix}frames_comparison_{model_name}.png"
            plt.savefig(frames_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 4: Plot overall distributions =====
    if display_plots and total_ssim.size > 0:
        # Calculate overall metrics
        mean_ssim_frames = np.mean(total_ssim)
        median_ssim_frames = np.median(total_ssim)
        std_ssim_frames = np.std(total_ssim)
        
        all_ssim['all'] = total_ssim
        
        # Plot overall distributions
        plt.figure(figsize=(10, 5))
        
        # SSIM distribution
        plt.hist(total_ssim, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_ssim, fill=True, color='b', label='KDE')
        plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
        plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
        plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
        plt.xlabel('SSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Structural Similarity (N = {total_ssim.shape[0]})')
        plt.legend()
        
        plt.tight_layout()
        
        # Save distribution plot if requested
        if save_plots and save_path_prefix:
            dist_path = f"{save_path_prefix}distributions_{model_name}.png"
            plt.savefig(dist_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # Update performance dictionary if provided
    if performance_dict:
        try:
            performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_ssim, 'SSIM', display_plots)
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    return performance_dict




def plot_all_predictions5(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, 
                 save_path_prefix=None, model_name="", device="cuda" if torch.cuda.is_available() else "cpu",
                 metric="ssim", mean_flag=False):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version includes extensive debugging and a more direct approach to handling multiple inputs.
    NEW: Added support for selecting between SSIM and TV loss as evaluation metrics.
    
    Parameters:
    -----------
    predictions : dict
        Dictionary of prediction arrays
    videos : dict
        Dictionary of ground truth video arrays
    performance_dict : dict, optional
        Dictionary to store performance metrics
    display_plots : bool
        Whether to display the plots
    save_plots : bool
        Whether to save the plots
    save_path_prefix : str, optional
        Path prefix for saving plots
    model_name : str
        Name of the model for saving plots
    device : str
        Device to use for computations
    metric : str
        Metric to use for evaluation: "ssim" or "tv" (Total Variation)
    mean_flag : bool
        Whether to return mean metrics or not
    """
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        import os
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    print(f"Using metric: {metric}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # If we don't have common keys, we need to find a match
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        # Take the first video key as reference (assuming all predictions relate to same video data)
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        
        # Map all prediction keys to this video key
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        # Create a mapping from prediction keys to video keys
        key_mapping = {}
        for pred_key in predictions.keys():
            # If the key exists directly in videos, use it
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                # Otherwise, try to find a matching video key
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                
                # If no match found and we have videos, use the first video key
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Initialize metrics
    total_metric_values = []
    all_metric_values = {'all': None}
    metric_name = "SSIM" if metric == "ssim" else "TV Loss"
    
    # Create a TV loss calculator if needed
    if metric == "tv":
        tv_calculator = TotalVariation().to(device)
    
    # ===== PART 1: Calculate metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        N = video.shape[0]
        
        # Calculate metric (SSIM or TV Loss)
        try:
            if metric == "ssim":
                # Calculate SSIM
                metric_frames = np.array([
                    ssim(torch.from_numpy(video[i]).unsqueeze(0), 
                         torch.from_numpy(prediction[i]).unsqueeze(0), 
                         data_range=1, size_average=True).item()
                    for i in range(N) 
                ])
                # For SSIM, higher is better
                mean_metric = np.mean(metric_frames)
                print(f"SSIM for {pred_key}: Mean = {mean_metric:.3f}")
            else:
                # Calculate TV Loss
                metric_frames = np.array([
                    tv_calculator(torch.from_numpy(prediction[i]).unsqueeze(0).to(device)).cpu().item() / 
                    (prediction[i].size)  # Normalize by number of elements
                    for i in range(N)
                ])
                # For TV Loss, lower is better
                mean_metric = np.mean(metric_frames)
                print(f"TV Loss for {pred_key}: Mean = {mean_metric:.3f}")
                
            all_metric_values[pred_key] = metric_frames
            total_metric_values = np.concatenate((total_metric_values, metric_frames))
        except Exception as e:
            print(f"Error calculating {metric_name} for {pred_key}: {e}")
            metric_frames = np.zeros(N)
            all_metric_values[pred_key] = metric_frames
    
    # ===== PART 2: Plot metrics for each prediction =====
    if display_plots and len(total_metric_values) > 0:
        # Plot metrics across all predictions
        plt.figure(figsize=(10, 5))
        
        # Plot metric
        for pred_key in key_mapping.keys():
            if pred_key in all_metric_values:
                plt.plot(all_metric_values[pred_key], label=f"{pred_key} (mean={np.mean(all_metric_values[pred_key]):.3f})")
        plt.xlabel("Frame Index")
        plt.ylabel(metric_name)
        plt.title(f"{metric_name} Comparison")
        plt.legend()
        
        plt.tight_layout()
        
        # Save metrics plot if requested
        if save_plots and save_path_prefix:
            metrics_path = f"{save_path_prefix}metrics_{metric}_comparison_{model_name}.png"
            plt.savefig(metrics_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 3: Plot frame comparisons =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # For large videos, sample frames
        if N > 10:
            indices = np.linspace(0, N-1, 10, dtype=int)
        else:
            indices = np.arange(N)
        
        print(f"Plotting {len(indices)} sample frames")
        
        # Number of rows and columns
        n_rows = len(indices)
        n_cols = 1 + len(key_mapping)  # Original + each prediction
        
        # Create a figure
        plt.figure(figsize=(4 * n_cols, 4 * n_rows))
        
        # REORDERING: Create an ordered list of prediction keys with "original_combined" as the second column
        ordered_keys = []
        for key in key_mapping.keys():
            if key != "original_combined":
                ordered_keys.append(key)
        
        # If "original_combined" exists, insert it at the beginning of the list
        if "original_combined" in key_mapping:
            ordered_keys.insert(0, "original_combined")
        
        # Plot each frame
        for i, frame_idx in enumerate(indices):
            # Plot original frame
            ax = plt.subplot(n_rows, n_cols, i * n_cols + 1)
            plt.imshow(np.transpose(normalize(ref_video[frame_idx]), (1, 2, 0)))
            plt.title(f"Original Frame {frame_idx}")
            plt.axis('off')
            
            # Plot each prediction in the reordered sequence
            for j, pred_key in enumerate(ordered_keys):
                try:
                    video_key = key_mapping[pred_key]
                    pred_frame = predictions[pred_key][frame_idx]
                    ax = plt.subplot(n_rows, n_cols, i * n_cols + j + 2)
                    plt.imshow(np.transpose(normalize(pred_frame), (1, 2, 0)))
                    
                    # Add metrics to title if available
                    metric_val = all_metric_values[pred_key][frame_idx] if pred_key in all_metric_values else 0
                    plt.title(f"{pred_key}\n{metric_name}: {metric_val:.3f}")
                    plt.axis('off')
                except Exception as e:
                    print(f"Error plotting prediction {pred_key} for frame {frame_idx}: {e}")
        
        plt.tight_layout()
        
        # Save frames comparison plot if requested
        if save_plots and save_path_prefix:
            frames_path = f"{save_path_prefix}frames_comparison_{metric}_{model_name}.png"
            plt.savefig(frames_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 4: Plot overall distributions =====
    if display_plots and len(total_metric_values) > 0:
        # Calculate overall metrics
        mean_metric = np.mean(total_metric_values)
        median_metric = np.median(total_metric_values)
        std_metric = np.std(total_metric_values)
        
        all_metric_values['all'] = total_metric_values
        
        # Plot overall distributions
        plt.figure(figsize=(10, 5))
        
        # Metric distribution
        plt.hist(total_metric_values, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_metric_values, fill=True, color='b', label='KDE')
        plt.axvline(x=median_metric, color='r', linestyle=':', linewidth=1, label=f'Median: {median_metric:.3f}')
        plt.axvline(x=mean_metric, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_metric:.3f}')
        plt.axvspan(mean_metric - std_metric, mean_metric + std_metric, color='red', alpha=0.05, label=f'std: {std_metric:.3f}')
        plt.xlabel(metric_name)
        plt.ylabel('Density')
        plt.title(f'Distribution of {metric_name} (N = {len(total_metric_values)})')
        plt.legend()
        
        plt.tight_layout()
        
        # Save distribution plot if requested
        if save_plots and save_path_prefix:
            dist_path = f"{save_path_prefix}distributions_{metric}_{model_name}.png"
            plt.savefig(dist_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # Update performance dictionary if provided
    if performance_dict:
        try:
            if metric == "ssim":
                performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_metric_values, 'SSIM', display_plots)
            else:
                performance_dict['mean_tv_D'], performance_dict['median_tv_D'] = plot_scatter_metrics(all_metric_values, 'TV Loss', display_plots)
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    if mean_flag:
        return mean_metric
    
    return performance_dict



'''
def plot_all_predictions6(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, 
                 save_path_prefix=None, model_name="", device="cuda" if torch.cuda.is_available() else "cpu",
                 metric="ssim", mean_flag=False, zone_type="quadrants"):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version computes SSIM (or TV loss) by zones (quadrants or center/background).
    An updated version of plot_all_predictions5 with zone-based metrics.
    
    Parameters:
    -----------
    predictions : dict
        Dictionary of prediction arrays
    videos : dict
        Dictionary of ground truth video arrays
    performance_dict : dict, optional
        Dictionary to store performance metrics
    display_plots : bool
        Whether to display the plots
    save_plots : bool
        Whether to save the plots
    save_path_prefix : str, optional
        Path prefix for saving plots
    model_name : str
        Name of the model for saving plots
    device : str
        Device to use for computations
    metric : str
        Metric to use for evaluation: "ssim" or "tv" (Total Variation)
    mean_flag : bool
        Whether to return mean metrics or not
    zone_type : str
        Type of zones: "quadrants" or "center_bg"
    """
    import numpy as np
    import matplotlib.pyplot as plt
    import seaborn as sns
    import torch
    import os
    from matplotlib.patches import Rectangle
    from matplotlib.colors import LinearSegmentedColormap
    
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    print(f"Using metric: {metric}")
    print(f"Using zone type: {zone_type}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # If we don't have common keys, we need to find a match
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        # Take the first video key as reference (assuming all predictions relate to same video data)
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        
        # Map all prediction keys to this video key
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        # Create a mapping from prediction keys to video keys
        key_mapping = {}
        for pred_key in predictions.keys():
            # If the key exists directly in videos, use it
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                # Otherwise, try to find a matching video key
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                
                # If no match found and we have videos, use the first video key
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Helper function to split frame into zones
    def split_into_zones(frame, zone_type="quadrants", center_ratio=0.5):
        """
        Split a frame into zones.
        
        Parameters:
        -----------
        frame : torch.Tensor or np.ndarray
            Frame to split
        zone_type : str
            Type of zones: "quadrants" or "center_bg"
        center_ratio : float
            Ratio of the center region size to the full image (for "center_bg" only)
            
        Returns:
        --------
        zones : dict
            Dictionary of zones with their corresponding slices
        """
        if isinstance(frame, torch.Tensor):
            C, H, W = frame.shape
        else:
            C, H, W = frame.shape
            
        zones = {}
        
        if zone_type == "quadrants":
            # Split into 4 quadrants
            h_mid = H // 2
            w_mid = W // 2
            
            zones["top_left"] = (slice(None), slice(0, h_mid), slice(0, w_mid))
            zones["top_right"] = (slice(None), slice(0, h_mid), slice(w_mid, W))
            zones["bottom_left"] = (slice(None), slice(h_mid, H), slice(0, w_mid))
            zones["bottom_right"] = (slice(None), slice(h_mid, H), slice(w_mid, W))
            
        elif zone_type == "center_bg":
            # Split into center and background
            h_center = int(H * center_ratio)
            w_center = int(W * center_ratio)
            
            h_start = (H - h_center) // 2
            h_end = h_start + h_center
            w_start = (W - w_center) // 2
            w_end = w_start + w_center
            
            zones["center"] = (slice(None), slice(h_start, h_end), slice(w_start, w_end))
            
            # Background is everything except the center (we'll handle this differently)
            # We'll use a mask approach for visualization
            center_mask = np.zeros((H, W), dtype=bool)
            center_mask[h_start:h_end, w_start:w_end] = True
            
            zones["background"] = {"mask": ~center_mask, 
                                   "bounds": (h_start, h_end, w_start, w_end)}
            
        else:
            raise ValueError(f"Unknown zone type: {zone_type}")
            
        return zones
    
    # Helper function to calculate zone metrics
    def calculate_zone_metrics(orig_frame, pred_frame, zones, metric="ssim", device=device):
        """
        Calculate metrics for each zone.
        
        Parameters:
        -----------
        orig_frame : np.ndarray
            Original frame
        pred_frame : np.ndarray
            Predicted frame
        zones : dict
            Dictionary of zones with their corresponding slices
        metric : str
            Metric to use: "ssim" or "tv"
        device : str
            Device to use for computations
            
        Returns:
        --------
        zone_metrics : dict
            Dictionary of metrics for each zone
        """
        from pytorch_msssim import ssim
        
        zone_metrics = {}
        
        # Convert to torch tensors if needed
        if not isinstance(orig_frame, torch.Tensor):
            orig_tensor = torch.from_numpy(orig_frame).unsqueeze(0)
        else:
            orig_tensor = orig_frame.unsqueeze(0)
            
        if not isinstance(pred_frame, torch.Tensor):
            pred_tensor = torch.from_numpy(pred_frame).unsqueeze(0)
        else:
            pred_tensor = pred_frame.unsqueeze(0)
        
        for zone_name, zone_slice in zones.items():
            # Special handling for background in center_bg mode
            if isinstance(zone_slice, dict):  # Background in center_bg mode
                mask = zone_slice["mask"]
                
                # Create masked tensors - this is more complex
                # We'll create copies where we zero out everything except our zone
                orig_zone = orig_tensor.clone()
                pred_zone = pred_tensor.clone()
                
                # Apply mask to all channels
                for c in range(orig_zone.shape[1]):  # For each channel
                    orig_zone[0, c][~mask] = 0
                    pred_zone[0, c][~mask] = 0
                
                # Calculate metric on the masked tensors
                if metric == "ssim":
                    zone_metrics[zone_name] = ssim(orig_zone, pred_zone, data_range=1, size_average=True).item()
                else:
                    # TV Loss calculation for masked region - simplified for this example
                    tv_loss = torch.abs(pred_zone[:,:,1:,:] - pred_zone[:,:,:-1,:]).sum() + \
                              torch.abs(pred_zone[:,:,:,1:] - pred_zone[:,:,:,:-1]).sum()
                    # Normalize by number of pixels in the zone
                    zone_metrics[zone_name] = tv_loss.item() / mask.sum()
                
            else:  # Normal zones
                # Get the zone data
                orig_zone = orig_tensor[0][zone_slice].unsqueeze(0)
                pred_zone = pred_tensor[0][zone_slice].unsqueeze(0)
                
                if metric == "ssim":
                    zone_metrics[zone_name] = ssim(orig_zone, pred_zone, data_range=1, size_average=True).item()
                else:
                    # TV Loss calculation
                    tv_loss = torch.abs(pred_zone[:,:,1:,:] - pred_zone[:,:,:-1,:]).sum() + \
                              torch.abs(pred_zone[:,:,:,1:] - pred_zone[:,:,:,:-1]).sum()
                    # Normalize by number of pixels in the zone
                    zone_metrics[zone_name] = tv_loss.item() / (orig_zone.shape[2] * orig_zone.shape[3])
        
        return zone_metrics
    
    # Helper function for normalizing images for display
    def normalize(img):
        """Normalize image for display"""
        if isinstance(img, torch.Tensor):
            img = img.detach().cpu().numpy()
        
        img = img.copy()
        if img.min() < 0:
            img = (img + 1) / 2  # [-1, 1] -> [0, 1]
        return np.clip(img, 0, 1)
    
    # Modified function to visualize zones with metrics
    def visualize_zones_with_metrics(frame, zone_metrics, zone_type, ax=None, title=None):
        """
        Visualize frame with zone boundaries and metrics.
        
        Parameters:
        -----------
        frame : np.ndarray
            Frame to visualize, should be in [C, H, W] format
        zone_metrics : dict
            Dictionary of metrics for each zone
        zone_type : str
            Type of zones: "quadrants" or "center_bg"
        ax : matplotlib.axes.Axes, optional
            Axes to plot on
        title : str, optional
            Title of the plot
        """
        if ax is None:
            _, ax = plt.subplots(figsize=(8, 8))
        
        # Display the image
        ax.imshow(np.transpose(normalize(frame), (1, 2, 0)))
        
        C, H, W = frame.shape
        
        # Draw zone boundaries and add metrics
        if zone_type == "quadrants":
            h_mid = H // 2
            w_mid = W // 2
            
            # Draw horizontal and vertical lines
            ax.axhline(y=h_mid, color='white', linestyle='-', linewidth=1)
            ax.axvline(x=w_mid, color='white', linestyle='-', linewidth=1)
            
            # Add metrics to each quadrant
            # Top left
            ax.text(w_mid/4, h_mid/4, f"{zone_metrics.get('top_left', 0):.3f}", 
                    color='white', fontsize=10, ha='center', va='center',
                    bbox=dict(facecolor='black', alpha=0.5))
            
            # Top right
            ax.text(w_mid + w_mid/2, h_mid/4, f"{zone_metrics.get('top_right', 0):.3f}", 
                    color='white', fontsize=10, ha='center', va='center',
                    bbox=dict(facecolor='black', alpha=0.5))
            
            # Bottom left
            ax.text(w_mid/4, h_mid + h_mid/2, f"{zone_metrics.get('bottom_left', 0):.3f}", 
                    color='white', fontsize=10, ha='center', va='center',
                    bbox=dict(facecolor='black', alpha=0.5))
            
            # Bottom right
            ax.text(w_mid + w_mid/2, h_mid + h_mid/2, f"{zone_metrics.get('bottom_right', 0):.3f}", 
                    color='white', fontsize=10, ha='center', va='center',
                    bbox=dict(facecolor='black', alpha=0.5))
            
        elif zone_type == "center_bg":
            # Get center boundaries
            bg_data = zones["background"]
            h_start, h_end, w_start, w_end = bg_data["bounds"]
            
            # Draw rectangle around center
            rect = Rectangle((w_start, h_start), w_end - w_start, h_end - h_start,
                            linewidth=1, edgecolor='white', facecolor='none')
            ax.add_patch(rect)
            
            # Add metrics to center and background
            # Center
            ax.text((w_start + w_end) / 2, (h_start + h_end) / 2, f"{zone_metrics.get('center', 0):.3f}", 
                    color='white', fontsize=10, ha='center', va='center',
                    bbox=dict(facecolor='black', alpha=0.5))
            
            # Background (in each corner)
            ax.text(w_start/2, h_start/2, f"BG: {zone_metrics.get('background', 0):.3f}", 
                    color='white', fontsize=10, ha='center', va='center',
                    bbox=dict(facecolor='black', alpha=0.5))
        
        ax.set_title(title or "")
        ax.axis('off')
        
        return ax
    
    # Initialize metrics
    all_zone_metrics = {}  # Will store metrics for each prediction key
    metric_name = "SSIM" if metric == "ssim" else "TV Loss"
    
    # Create a TV loss calculator if needed
    if metric == "tv":
        tv_calculator = TotalVariation().to(device)
    
    # ===== PART 1: Calculate zone metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        N = video.shape[0]
        
        # Store metrics for all frames in this prediction
        pred_metrics = []
        
        # Calculate metrics for each frame
        for i in range(N):
            # Get zones for this frame
            zones = split_into_zones(video[i], zone_type=zone_type)
            
            # Calculate metrics for each zone
            try:
                zone_metrics = calculate_zone_metrics(video[i], prediction[i], zones, metric=metric)
                pred_metrics.append(zone_metrics)
            except Exception as e:
                print(f"Error calculating zone metrics for {pred_key}, frame {i}: {e}")
                # Create empty metrics
                if zone_type == "quadrants":
                    pred_metrics.append({
                        "top_left": 0, "top_right": 0, 
                        "bottom_left": 0, "bottom_right": 0
                    })
                else:
                    pred_metrics.append({"center": 0, "background": 0})
        
        # Store metrics for this prediction
        all_zone_metrics[pred_key] = pred_metrics
        
        # Print average metrics for this prediction
        print(f"\nAverage {metric_name} for {pred_key} by zone:")
        
        # Calculate and print mean metrics across frames for each zone
        if len(pred_metrics) > 0:
            zone_names = list(pred_metrics[0].keys())
            
            for zone in zone_names:
                zone_values = [metrics[zone] for metrics in pred_metrics]
                mean_zone = np.mean(zone_values)
                print(f"  - {zone}: {mean_zone:.3f}")
    
    # ===== PART 2: Plot zone metrics for each prediction =====
    if display_plots and len(all_zone_metrics) > 0:
        # For each prediction, plot the zone metrics across frames
        for pred_key, pred_metrics in all_zone_metrics.items():
            if len(pred_metrics) == 0:
                continue
                
            zone_names = list(pred_metrics[0].keys())
            
            plt.figure(figsize=(12, 6))
            
            for zone in zone_names:
                # Get metrics for this zone across all frames
                zone_values = [metrics[zone] for metrics in pred_metrics]
                plt.plot(zone_values, label=f"{zone} (mean={np.mean(zone_values):.3f})")
            
            plt.xlabel("Frame Index")
            plt.ylabel(metric_name)
            plt.title(f"{metric_name} by Zone for {pred_key}")
            plt.legend()
            plt.tight_layout()
            
            # Save metrics plot if requested
            if save_plots and save_path_prefix:
                metrics_path = f"{save_path_prefix}zone_metrics_{metric}_{zone_type}_{pred_key}.png"
                plt.savefig(metrics_path, bbox_inches='tight', dpi=300)
            
            plt.show()
    
    # ===== PART 3: Plot frame comparisons with zone visualization =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # For large videos, sample frames
        if N > 5:
            indices = np.linspace(0, N-1, 5, dtype=int)
        else:
            indices = np.arange(N)
        
        print(f"\nPlotting {len(indices)} sample frames with zones")
        
        # Number of rows and columns
        n_rows = len(indices)
        n_cols = 1 + len(key_mapping)  # Original + each prediction
        
        # Create a figure
        plt.figure(figsize=(5 * n_cols, 5 * n_rows))
        
        # REORDERING: Create an ordered list of prediction keys with "original_combined" first
        ordered_keys = []
        for key in key_mapping.keys():
            if key != "original_combined":
                ordered_keys.append(key)
        
        # If "original_combined" exists, insert it at the beginning of the list
        if "original_combined" in key_mapping:
            ordered_keys.insert(0, "original_combined")
        
        # For each frame index
        for i, frame_idx in enumerate(indices):
            # First, get zones for the original frame
            ref_frame = ref_video[frame_idx]
            zones = split_into_zones(ref_frame, zone_type=zone_type)
            
            # Plot original frame with zones but no metrics
            ax = plt.subplot(n_rows, n_cols, i * n_cols + 1)
            ax.imshow(np.transpose(normalize(ref_frame), (1, 2, 0)))
            
            # Draw zone boundaries
            if zone_type == "quadrants":
                C, H, W = ref_frame.shape
                h_mid = H // 2
                w_mid = W // 2
                ax.axhline(y=h_mid, color='white', linestyle='-', linewidth=1)
                ax.axvline(x=w_mid, color='white', linestyle='-', linewidth=1)
            elif zone_type == "center_bg":
                bg_data = zones["background"]
                h_start, h_end, w_start, w_end = bg_data["bounds"]
                rect = Rectangle((w_start, h_start), w_end - w_start, h_end - h_start,
                              linewidth=1, edgecolor='white', facecolor='none')
                ax.add_patch(rect)
            
            ax.set_title(f"Original Frame {frame_idx}")
            ax.axis('off')
            
            # Plot each prediction with zone metrics
            for j, pred_key in enumerate(ordered_keys):
                try:
                    video_key = key_mapping[pred_key]
                    pred_frame = predictions[pred_key][frame_idx]
                    
                    # Get zone metrics for this prediction
                    zone_metrics = all_zone_metrics[pred_key][frame_idx]
                    
                    # Plot with zone visualization
                    ax = plt.subplot(n_rows, n_cols, i * n_cols + j + 2)
                    visualize_zones_with_metrics(
                        pred_frame, zone_metrics, zone_type, ax=ax,
                        title=f"{pred_key} (Frame {frame_idx})"
                    )
                    
                except Exception as e:
                    print(f"Error plotting prediction {pred_key} for frame {frame_idx}: {e}")
        
        plt.tight_layout()
        
        # Save frames comparison plot if requested
        if save_plots and save_path_prefix:
            frames_path = f"{save_path_prefix}zone_frames_{metric}_{zone_type}_{model_name}.png"
            plt.savefig(frames_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 4: Plot overall zone statistics =====
    if display_plots and len(all_zone_metrics) > 0:
        # Create a single plot with all zone statistics
        plt.figure(figsize=(12, 8))
        
        # Get all zone names from the first prediction
        first_pred_key = list(all_zone_metrics.keys())[0]
        if len(all_zone_metrics[first_pred_key]) > 0:
            zone_names = list(all_zone_metrics[first_pred_key][0].keys())
            
            # Number of zones and predictions
            n_zones = len(zone_names)
            n_preds = len(all_zone_metrics)
            
            # Create a bar plot of mean metrics for each zone for each prediction
            bar_width = 0.8 / n_preds
            index = np.arange(n_zones)
            
            for i, (pred_key, pred_metrics) in enumerate(all_zone_metrics.items()):
                # Calculate mean for each zone
                zone_means = []
                for zone in zone_names:
                    zone_values = [metrics[zone] for metrics in pred_metrics]
                    zone_means.append(np.mean(zone_values))
                
                # Plot bars
                plt.bar(index + i * bar_width, zone_means, bar_width,
                        label=pred_key, alpha=0.7)
            
            plt.xlabel("Zone")
            plt.ylabel(metric_name)
            plt.title(f"Average {metric_name} by Zone")
            plt.xticks(index + bar_width * (n_preds - 1) / 2, zone_names)
            plt.legend()
            plt.tight_layout()
            
            # Save zone statistics plot if requested
            if save_plots and save_path_prefix:
                stats_path = f"{save_path_prefix}zone_stats_{metric}_{zone_type}_{model_name}.png"
                plt.savefig(stats_path, bbox_inches='tight', dpi=300)
            
            plt.show()
    
    # Update performance dictionary if provided
    if performance_dict:
        try:
            # Calculate overall metrics across all zones
            for pred_key, pred_metrics in all_zone_metrics.items():
                for zone in zone_names:
                    zone_values = [metrics[zone] for metrics in pred_metrics]
                    zone_mean = np.mean(zone_values)
                    zone_median = np.median(zone_values)
                    
                    # Add to performance dict
                    if metric == "ssim":
                        performance_dict[f'mean_ssim_{zone}_D'] = zone_mean
                        performance_dict[f'median_ssim_{zone}_D'] = zone_median
                    else:
                        performance_dict[f'mean_tv_{zone}_D'] = zone_mean
                        performance_dict[f'median_tv_{zone}_D'] = zone_median
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    # Calculate overall mean metric
    overall_mean = 0
    if len(all_zone_metrics) > 0:
        # Average across all predictions and all zones
        all_values = []
        for pred_metrics in all_zone_metrics.values():
            for metrics in pred_metrics:
                all_values.extend(list(metrics.values()))
        
        if all_values:
            overall_mean = np.mean(all_values)
    
    if mean_flag:
        return overall_mean
    
    return performance_dict
'''

def split_into_zones(frame, zone_type="quadrants", center_ratio=0.5):
    """
    Split a frame into zones.
    
    Parameters:
    -----------
    frame : torch.Tensor or np.ndarray
        Frame to split
    zone_type : str or int
        Type of zones: 
        - "quadrants" for 2x2 grid
        - "center_bg" for center and background
        - integer n for n×n grid (e.g., 4 creates a 4×4 grid with 16 zones)
    center_ratio : float
        Ratio of the center region size to the full image (for "center_bg" only)
        
    Returns:
    --------
    zones : dict
        Dictionary of zones with their corresponding slices
    """
    if isinstance(frame, torch.Tensor):
        C, H, W = frame.shape
    else:
        C, H, W = frame.shape
        
    zones = {}
    
    if zone_type == "quadrants":
        # Split into 4 quadrants (2×2 grid)
        h_mid = H // 2
        w_mid = W // 2
        
        zones["top_left"] = (slice(None), slice(0, h_mid), slice(0, w_mid))
        zones["top_right"] = (slice(None), slice(0, h_mid), slice(w_mid, W))
        zones["bottom_left"] = (slice(None), slice(h_mid, H), slice(0, w_mid))
        zones["bottom_right"] = (slice(None), slice(h_mid, H), slice(w_mid, W))
        
    elif zone_type == "center_bg":
        # Split into center and background
        h_center = int(H * center_ratio)
        w_center = int(W * center_ratio)
        
        h_start = (H - h_center) // 2
        h_end = h_start + h_center
        w_start = (W - w_center) // 2
        w_end = w_start + w_center
        
        zones["center"] = (slice(None), slice(h_start, h_end), slice(w_start, w_end))
        
        # Background is everything except the center (we'll handle this differently)
        # We'll use a mask approach for visualization
        center_mask = np.zeros((H, W), dtype=bool)
        center_mask[h_start:h_end, w_start:w_end] = True
        
        zones["background"] = {"mask": ~center_mask, 
                               "bounds": (h_start, h_end, w_start, w_end)}
    
    elif isinstance(zone_type, int) and zone_type > 0:
        # Create an n×n grid where n = zone_type
        n = zone_type
        
        # Calculate heights of each section
        h_sections = [i * H // n for i in range(n+1)]
        w_sections = [i * W // n for i in range(n+1)]
        
        # Create zones for each grid cell
        for i in range(n):
            for j in range(n):
                zone_name = f"grid_{i}_{j}"  # Row_Column naming
                zones[zone_name] = (
                    slice(None),
                    slice(h_sections[i], h_sections[i+1]),
                    slice(w_sections[j], w_sections[j+1])
                )
                
    else:
        raise ValueError(f"Unknown zone type: {zone_type}")
        
    return zones


def visualize_zones_with_metrics(frame, zone_metrics, zone_type, ax=None, title=None):
    """
    Visualize frame with zone boundaries and metrics.
    
    Parameters:
    -----------
    frame : np.ndarray
        Frame to visualize, should be in [C, H, W] format
    zone_metrics : dict
        Dictionary of metrics for each zone
    zone_type : str or int
        Type of zones: "quadrants", "center_bg", or integer for grid
    ax : matplotlib.axes.Axes, optional
        Axes to plot on
    title : str, optional
        Title of the plot
    """
    import numpy as np
    import matplotlib.pyplot as plt
    from matplotlib.patches import Rectangle
    
    if ax is None:
        _, ax = plt.subplots(figsize=(8, 8))
    
    # Helper function for normalizing images
    def normalize(img):
        if isinstance(img, torch.Tensor):
            img = img.detach().cpu().numpy()
        
        img = img.copy()
        if img.min() < 0:
            img = (img + 1) / 2  # [-1, 1] -> [0, 1]
        return np.clip(img, 0, 1)
    
    # Display the image
    ax.imshow(np.transpose(normalize(frame), (1, 2, 0)))
    
    C, H, W = frame.shape
    
    # Draw zone boundaries and add metrics
    if zone_type == "quadrants":
        h_mid = H // 2
        w_mid = W // 2
        
        # Draw horizontal and vertical lines
        ax.axhline(y=h_mid, color='white', linestyle='-', linewidth=1)
        ax.axvline(x=w_mid, color='white', linestyle='-', linewidth=1)
        
        # Add metrics to each quadrant
        # Top left
        ax.text(w_mid/4, h_mid/4, f"{zone_metrics.get('top_left', 0):.3f}", 
                color='white', fontsize=10, ha='center', va='center',
                bbox=dict(facecolor='black', alpha=0.5))
        
        # Top right
        ax.text(w_mid + w_mid/2, h_mid/4, f"{zone_metrics.get('top_right', 0):.3f}", 
                color='white', fontsize=10, ha='center', va='center',
                bbox=dict(facecolor='black', alpha=0.5))
        
        # Bottom left
        ax.text(w_mid/4, h_mid + h_mid/2, f"{zone_metrics.get('bottom_left', 0):.3f}", 
                color='white', fontsize=10, ha='center', va='center',
                bbox=dict(facecolor='black', alpha=0.5))
        
        # Bottom right
        ax.text(w_mid + w_mid/2, h_mid + h_mid/2, f"{zone_metrics.get('bottom_right', 0):.3f}", 
                color='white', fontsize=10, ha='center', va='center',
                bbox=dict(facecolor='black', alpha=0.5))
        
    elif zone_type == "center_bg":
        # Get center boundaries from zone_metrics keys
        # Find the background key
        bg_key = [k for k in zone_metrics.keys() if k == 'background'][0]
        center_key = [k for k in zone_metrics.keys() if k == 'center'][0]
        
        # These coordinates would need to be calculated or passed separately
        h_center = int(H * 0.5)  # Using default center_ratio
        w_center = int(W * 0.5)
        
        h_start = (H - h_center) // 2
        h_end = h_start + h_center
        w_start = (W - w_center) // 2
        w_end = w_start + w_center
        
        # Draw rectangle around center
        rect = Rectangle((w_start, h_start), w_end - w_start, h_end - h_start,
                        linewidth=1, edgecolor='white', facecolor='none')
        ax.add_patch(rect)
        
        # Add metrics to center and background
        # Center
        ax.text((w_start + w_end) / 2, (h_start + h_end) / 2, f"{zone_metrics.get(center_key, 0):.3f}", 
                color='white', fontsize=10, ha='center', va='center',
                bbox=dict(facecolor='black', alpha=0.5))
        
        # Background (in corner)
        ax.text(w_start/2, h_start/2, f"BG: {zone_metrics.get(bg_key, 0):.3f}", 
                color='white', fontsize=10, ha='center', va='center',
                bbox=dict(facecolor='black', alpha=0.5))
    
    elif isinstance(zone_type, int) and zone_type > 0:
        # Grid visualization
        n = zone_type
        
        # Calculate heights of each section
        h_sections = [i * H // n for i in range(n+1)]
        w_sections = [i * W // n for i in range(n+1)]
        
        # Draw grid lines
        for i in range(1, n):
            ax.axhline(y=h_sections[i], color='white', linestyle='-', linewidth=1)
            ax.axvline(x=w_sections[i], color='white', linestyle='-', linewidth=1)
        
        # Add metrics to each grid cell
        font_size = max(8, min(10, 14 - n))  # Adjust font size based on grid size
        
        for i in range(n):
            for j in range(n):
                zone_name = f"grid_{i}_{j}"
                if zone_name in zone_metrics:
                    # Calculate center of this zone
                    y_center = (h_sections[i] + h_sections[i+1]) / 2
                    x_center = (w_sections[j] + w_sections[j+1]) / 2
                    
                    ax.text(x_center, y_center, f"{zone_metrics.get(zone_name, 0):.3f}", 
                           color='white', fontsize=font_size, ha='center', va='center',
                           bbox=dict(facecolor='black', alpha=0.5))
    
    ax.set_title(title or "")
    ax.axis('off')
    
    return ax

'''
def plot_all_predictions7(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, 
                 save_path_prefix=None, model_name="", device="cuda" if torch.cuda.is_available() else "cpu",
                 metric="ssim", mean_flag=False, zone_type="quadrants"):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version computes SSIM (or TV loss) by zones (quadrants, center/background, or n×n grid).
    
    Parameters:
    -----------
    predictions : dict
        Dictionary of prediction arrays
    videos : dict
        Dictionary of ground truth video arrays
    performance_dict : dict, optional
        Dictionary to store performance metrics
    display_plots : bool
        Whether to display the plots
    save_plots : bool
        Whether to save the plots
    save_path_prefix : str, optional
        Path prefix for saving plots
    model_name : str
        Name of the model for saving plots
    device : str
        Device to use for computations
    metric : str
        Metric to use for evaluation: "ssim" or "tv" (Total Variation)
    mean_flag : bool
        Whether to return mean metrics or not
    zone_type : str or int
        Type of zones: 
        - "quadrants" for 2×2 grid
        - "center_bg" for center and background
        - integer n for n×n grid (e.g., 4 creates a 4×4 grid with 16 zones)
    """
    import numpy as np
    import matplotlib.pyplot as plt
    import torch
    import os
    from matplotlib.patches import Rectangle
    from matplotlib.colors import LinearSegmentedColormap
    
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    print(f"Using metric: {metric}")
    
    if isinstance(zone_type, int):
        print(f"Using {zone_type}×{zone_type} grid zones ({zone_type*zone_type} total zones)")
    else:
        print(f"Using zone type: {zone_type}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # If we don't have common keys, we need to find a match
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        # Take the first video key as reference (assuming all predictions relate to same video data)
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        
        # Map all prediction keys to this video key
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        # Create a mapping from prediction keys to video keys
        key_mapping = {}
        for pred_key in predictions.keys():
            # If the key exists directly in videos, use it
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                # Otherwise, try to find a matching video key
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                
                # If no match found and we have videos, use the first video key
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Helper function to calculate zone metrics
    def calculate_zone_metrics(orig_frame, pred_frame, zones, metric="ssim", device=device):
        """
        Calculate metrics for each zone.
        
        Parameters:
        -----------
        orig_frame : np.ndarray
            Original frame
        pred_frame : np.ndarray
            Predicted frame
        zones : dict
            Dictionary of zones with their corresponding slices
        metric : str
            Metric to use: "ssim" or "tv"
        device : str
            Device to use for computations
            
        Returns:
        --------
        zone_metrics : dict
            Dictionary of metrics for each zone
        """
        from pytorch_msssim import ssim
        
        zone_metrics = {}
        
        # Convert to torch tensors if needed
        if not isinstance(orig_frame, torch.Tensor):
            orig_tensor = torch.from_numpy(orig_frame).unsqueeze(0)
        else:
            orig_tensor = orig_frame.unsqueeze(0)
            
        if not isinstance(pred_frame, torch.Tensor):
            pred_tensor = torch.from_numpy(pred_frame).unsqueeze(0)
        else:
            pred_tensor = pred_frame.unsqueeze(0)
        
        for zone_name, zone_slice in zones.items():
            # Special handling for background in center_bg mode
            if isinstance(zone_slice, dict):  # Background in center_bg mode
                mask = zone_slice["mask"]
                
                # Create masked tensors - this is more complex
                # We'll create copies where we zero out everything except our zone
                orig_zone = orig_tensor.clone()
                pred_zone = pred_tensor.clone()
                
                # Apply mask to all channels
                for c in range(orig_zone.shape[1]):  # For each channel
                    orig_zone[0, c][~mask] = 0
                    pred_zone[0, c][~mask] = 0
                
                # Calculate metric on the masked tensors
                if metric == "ssim":
                    zone_metrics[zone_name] = ssim(orig_zone, pred_zone, data_range=1, size_average=True).item()
                else:
                    # TV Loss calculation for masked region - simplified for this example
                    tv_loss = torch.abs(pred_zone[:,:,1:,:] - pred_zone[:,:,:-1,:]).sum() + \
                              torch.abs(pred_zone[:,:,:,1:] - pred_zone[:,:,:,:-1]).sum()
                    # Normalize by number of pixels in the zone
                    zone_metrics[zone_name] = tv_loss.item() / mask.sum()
                
            else:  # Normal zones
                # Get the zone data
                orig_zone = orig_tensor[0][zone_slice].unsqueeze(0)
                pred_zone = pred_tensor[0][zone_slice].unsqueeze(0)
                
                if metric == "ssim":
                    zone_metrics[zone_name] = ssim(orig_zone, pred_zone, data_range=1, size_average=True).item()
                else:
                    # TV Loss calculation
                    tv_loss = torch.abs(pred_zone[:,:,1:,:] - pred_zone[:,:,:-1,:]).sum() + \
                              torch.abs(pred_zone[:,:,:,1:] - pred_zone[:,:,:,:-1]).sum()
                    # Normalize by number of pixels in the zone
                    zone_metrics[zone_name] = tv_loss.item() / (orig_zone.shape[2] * orig_zone.shape[3])
        
        return zone_metrics
    
    # Helper function for normalizing images for display
    def normalize(img):
        """Normalize image for display"""
        if isinstance(img, torch.Tensor):
            img = img.detach().cpu().numpy()
        
        img = img.copy()
        if img.min() < 0:
            img = (img + 1) / 2  # [-1, 1] -> [0, 1]
        return np.clip(img, 0, 1)
    
    # Initialize metrics
    all_zone_metrics = {}  # Will store metrics for each prediction key
    metric_name = "SSIM" if metric == "ssim" else "TV Loss"
    
    # Create a TV loss calculator if needed
    if metric == "tv":
        tv_calculator = TotalVariation().to(device)
    
    # ===== PART 1: Calculate zone metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        N = video.shape[0]
        
        # Store metrics for all frames in this prediction
        pred_metrics = []
        
        # Calculate metrics for each frame
        for i in range(N):
            # Get zones for this frame
            zones = split_into_zones(video[i], zone_type=zone_type)
            
            # Calculate metrics for each zone
            try:
                zone_metrics = calculate_zone_metrics(video[i], prediction[i], zones, metric=metric)
                pred_metrics.append(zone_metrics)
            except Exception as e:
                print(f"Error calculating zone metrics for {pred_key}, frame {i}: {e}")
                # Create empty metrics
                if zone_type == "quadrants":
                    pred_metrics.append({
                        "top_left": 0, "top_right": 0, 
                        "bottom_left": 0, "bottom_right": 0
                    })
                elif zone_type == "center_bg":
                    pred_metrics.append({"center": 0, "background": 0})
                elif isinstance(zone_type, int):
                    empty_metrics = {}
                    for ii in range(zone_type):
                        for jj in range(zone_type):
                            empty_metrics[f"grid_{ii}_{jj}"] = 0
                    pred_metrics.append(empty_metrics)
        
        # Store metrics for this prediction
        all_zone_metrics[pred_key] = pred_metrics
        
        # Print average metrics for this prediction
        print(f"\nAverage {metric_name} for {pred_key} by zone:")
        
        # Calculate and print mean metrics across frames for each zone
        if len(pred_metrics) > 0:
            zone_names = list(pred_metrics[0].keys())
            
            for zone in zone_names:
                zone_values = [metrics[zone] for metrics in pred_metrics]
                mean_zone = np.mean(zone_values)
                print(f"  - {zone}: {mean_zone:.3f}")
    
    # ===== PART 2: Plot zone metrics for each prediction =====
    if display_plots and len(all_zone_metrics) > 0:
        # For each prediction, plot the zone metrics across frames
        for pred_key, pred_metrics in all_zone_metrics.items():
            if len(pred_metrics) == 0:
                continue
                
            zone_names = list(pred_metrics[0].keys())
            
            # If there are too many zones, plot only a subset or aggregate
            if len(zone_names) > 10:
                # For a grid, we might want to aggregate by rows or columns
                if isinstance(zone_type, int):
                    plt.figure(figsize=(15, 10))
                    
                    # Create row averages
                    for i in range(zone_type):
                        row_zones = [f"grid_{i}_{j}" for j in range(zone_type)]
                        row_values = []
                        
                        for frame_idx in range(len(pred_metrics)):
                            # Average the values for this row in this frame
                            row_avg = np.mean([pred_metrics[frame_idx][zone] for zone in row_zones])
                            row_values.append(row_avg)
                        
                        plt.plot(row_values, label=f"Row {i} (mean={np.mean(row_values):.3f})")
                    
                    plt.xlabel("Frame Index")
                    plt.ylabel(f"Row Average {metric_name}")
                    plt.title(f"Row-Averaged {metric_name} for {pred_key}")
                    plt.legend()
                    plt.tight_layout()
                    
                    # Also create a second plot for column averages
                    plt.figure(figsize=(15, 10))
                    
                    for j in range(zone_type):
                        col_zones = [f"grid_{i}_{j}" for i in range(zone_type)]
                        col_values = []
                        
                        for frame_idx in range(len(pred_metrics)):
                            # Average the values for this column in this frame
                            col_avg = np.mean([pred_metrics[frame_idx][zone] for zone in col_zones])
                            col_values.append(col_avg)
                        
                        plt.plot(col_values, label=f"Column {j} (mean={np.mean(col_values):.3f})")
                    
                    plt.xlabel("Frame Index")
                    plt.ylabel(f"Column Average {metric_name}")
                    plt.title(f"Column-Averaged {metric_name} for {pred_key}")
                    plt.legend()
                    plt.tight_layout()
                else:
                    # For other zone types with many zones, just plot all
                    plt.figure(figsize=(15, 10))
                    
                    for zone in zone_names:
                        # Get metrics for this zone across all frames
                        zone_values = [metrics[zone] for metrics in pred_metrics]
                        plt.plot(zone_values, label=f"{zone} (mean={np.mean(zone_values):.3f})")
                    
                    plt.xlabel("Frame Index")
                    plt.ylabel(metric_name)
                    plt.title(f"{metric_name} by Zone for {pred_key}")
                    plt.legend()
                    plt.tight_layout()
            else:
                # For fewer zones, plot all normally
                plt.figure(figsize=(12, 6))
                
                for zone in zone_names:
                    # Get metrics for this zone across all frames
                    zone_values = [metrics[zone] for metrics in pred_metrics]
                    plt.plot(zone_values, label=f"{zone} (mean={np.mean(zone_values):.3f})")
                
                plt.xlabel("Frame Index")
                plt.ylabel(metric_name)
                plt.title(f"{metric_name} by Zone for {pred_key}")
                plt.legend()
                plt.tight_layout()
            
            # Save metrics plot if requested
            if save_plots and save_path_prefix:
                metrics_path = f"{save_path_prefix}zone_metrics_{metric}_{zone_type}_{pred_key}.png"
                plt.savefig(metrics_path, bbox_inches='tight', dpi=300)
            
            plt.show()
    
    # ===== PART 3: Plot frame comparisons with zone visualization =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # For large videos, sample frames
        if N > 5:
            indices = np.linspace(0, N-1, 5, dtype=int)
        else:
            indices = np.arange(N)
        
        print(f"\nPlotting {len(indices)} sample frames with zones")
        
        # Number of rows and columns
        n_rows = len(indices)
        n_cols = 1 + len(key_mapping)  # Original + each prediction
        
        # Create a figure
        plt.figure(figsize=(5 * n_cols, 5 * n_rows))
        
        # REORDERING: Create an ordered list of prediction keys with "original_combined" first
        ordered_keys = []
        for key in key_mapping.keys():
            if key != "original_combined":
                ordered_keys.append(key)
        
        # If "original_combined" exists, insert it at the beginning of the list
        if "original_combined" in key_mapping:
            ordered_keys.insert(0, "original_combined")
        
        # For each frame index
        for i, frame_idx in enumerate(indices):
            # First, get zones for the original frame
            ref_frame = ref_video[frame_idx]
            zones = split_into_zones(ref_frame, zone_type=zone_type)
            
            # Plot original frame with zones
            ax = plt.subplot(n_rows, n_cols, i * n_cols + 1)
            
            # Create empty metrics dictionary for visualization only
            empty_metrics = {}
            if isinstance(zone_type, int):
                # For grid, create empty metrics for each cell
                for ii in range(zone_type):
                    for jj in range(zone_type):
                        empty_metrics[f"grid_{ii}_{jj}"] = 0
            elif zone_type == "quadrants":
                empty_metrics = {"top_left": 0, "top_right": 0, "bottom_left": 0, "bottom_right": 0}
            elif zone_type == "center_bg":
                empty_metrics = {"center": 0, "background": 0}
            
            # Visualize original frame with zones
            visualize_zones_with_metrics(ref_frame, empty_metrics, zone_type, ax=ax, title=f"Original Frame {frame_idx}")
            
            # Plot each prediction with zone metrics
            for j, pred_key in enumerate(ordered_keys):
                try:
                    video_key = key_mapping[pred_key]
                    pred_frame = predictions[pred_key][frame_idx]
                    
                    # Get zone metrics for this prediction
                    zone_metrics = all_zone_metrics[pred_key][frame_idx]
                    
                    # Plot with zone visualization
                    ax = plt.subplot(n_rows, n_cols, i * n_cols + j + 2)
                    visualize_zones_with_metrics(
                        pred_frame, zone_metrics, zone_type, ax=ax,
                        title=f"{pred_key} (Frame {frame_idx})"
                    )
                    
                except Exception as e:
                    print(f"Error plotting prediction {pred_key} for frame {frame_idx}: {e}")
        
        plt.tight_layout()
        
        # Save frames comparison plot if requested
        if save_plots and save_path_prefix:
            zone_type_str = zone_type if isinstance(zone_type, str) else f"grid{zone_type}x{zone_type}"
            frames_path = f"{save_path_prefix}zone_frames_{metric}_{zone_type_str}_{model_name}.png"
            plt.savefig(frames_path, bbox_inches='tight', dpi=300)
        
        plt.show()
    
    # ===== PART 4: Plot overall zone statistics (for grid, plot as heatmap) =====
    if display_plots and len(all_zone_metrics) > 0:
        # Create a single plot with all zone statistics
        first_pred_key = list(all_zone_metrics.keys())[0]
        
        if len(all_zone_metrics[first_pred_key]) > 0:
            zone_names = list(all_zone_metrics[first_pred_key][0].keys())
            
            # For grid zones, create a heatmap
            if isinstance(zone_type, int):
                for pred_key, pred_metrics in all_zone_metrics.items():
                    plt.figure(figsize=(10, 8))
                    
                    # Create a grid to hold average values
                    grid_avg = np.zeros((zone_type, zone_type))
                    
                    # Calculate average for each grid cell
                    for i in range(zone_type):
                        for j in range(zone_type):
                            zone_name = f"grid_{i}_{j}"
                            zone_values = [metrics[zone_name] for metrics in pred_metrics]
                            grid_avg[i, j] = np.mean(zone_values)
                    
                    # Create heatmap
                    plt.imshow(grid_avg, cmap='viridis')
                    plt.colorbar(label=metric_name)
                    
                    # Add text annotations
                    for i in range(zone_type):
                        for j in range(zone_type):
                            plt.text(j, i, f"{grid_avg[i, j]:.3f}", 
                                    ha="center", va="center", color="white",
                                    fontsize=10 if zone_type < 5 else 8)
                    
                    plt.title(f"{metric_name} by Zone for {pred_key}")
                    plt.xlabel("Column")
                    plt.ylabel("Row")
                    plt.tight_layout()
                    
                    if save_plots and save_path_prefix:
                        heatmap_path = f"{save_path_prefix}zone_heatmap_{metric}_grid{zone_type}x{zone_type}_{pred_key}.png"
                        plt.savefig(heatmap_path, bbox_inches='tight', dpi=300)
                    
                    plt.show()
            else:
                # For other zone types, create a bar chart
                plt.figure(figsize=(12, 8))
                
                # Number of zones and predictions
                n_zones = len(zone_names)
                n_preds = len(all_zone_metrics)
                
                # Create a bar plot of mean metrics for each zone for each prediction
                bar_width = 0.8 / n_preds
                index = np.arange(n_zones)
                
                for i, (pred_key, pred_metrics) in enumerate(all_zone_metrics.items()):
                    # Calculate mean for each zone
                    zone_means = []
                    for zone in zone_names:
                        zone_values = [metrics[zone] for metrics in pred_metrics]
                        zone_means.append(np.mean(zone_values))
                    
                    # Plot bars
                    plt.bar(index + i * bar_width, zone_means, bar_width,
                            label=pred_key, alpha=0.7)
                
                plt.xlabel("Zone")
                plt.ylabel(metric_name)
                plt.title(f"Average {metric_name} by Zone")
                plt.xticks(index + bar_width * (n_preds - 1) / 2, zone_names)
                plt.legend()
                plt.tight_layout()
                
                # Save zone statistics plot if requested
                if save_plots and save_path_prefix:
                    stats_path = f"{save_path_prefix}zone_stats_{metric}_{zone_type}_{model_name}.png"
                    plt.savefig(stats_path, bbox_inches='tight', dpi=300)
                
                plt.show()
    
    # Update performance dictionary if provided
    if performance_dict is not None:
        try:
            # Calculate overall metrics across all zones
            for pred_key, pred_metrics in all_zone_metrics.items():
                if len(pred_metrics) > 0:
                    zone_names = list(pred_metrics[0].keys())
                    
                    for zone in zone_names:
                        zone_values = [metrics[zone] for metrics in pred_metrics]
                        zone_mean = np.mean(zone_values)
                        zone_median = np.median(zone_values)
                        
                        # Add to performance dict
                        if metric == "ssim":
                            performance_dict[f'mean_ssim_{zone}_D'] = zone_mean
                            performance_dict[f'median_ssim_{zone}_D'] = zone_median
                        else:
                            performance_dict[f'mean_tv_{zone}_D'] = zone_mean
                            performance_dict[f'median_tv_{zone}_D'] = zone_median
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    # Calculate overall mean metric
    overall_mean = 0
    if len(all_zone_metrics) > 0:
        # Average across all predictions and all zones
        all_values = []
        for pred_metrics in all_zone_metrics.values():
            for metrics in pred_metrics:
                all_values.extend(list(metrics.values()))
        
        if all_values:
            overall_mean = np.mean(all_values)
    
    if mean_flag:
        return overall_mean
    
    return performance_dict
'''


#this was the plot_all_predictions7 and it was working well to show the loss, the new one just shows difference from baseline
def plot_all_predictions72(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, 
                 save_path_prefix=None, model_name="", device="cuda" if torch.cuda.is_available() else "cpu",
                 metric="ssim", mean_flag=False, zone_type="quadrants", max_frames=None):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version shows: Original image, reconstructed image, and heatmap of loss values.
    
    Parameters:
    -----------
    predictions : dict
        Dictionary of prediction arrays
    videos : dict
        Dictionary of ground truth video arrays
    performance_dict : dict, optional
        Dictionary to store performance metrics
    display_plots : bool
        Whether to display the plots
    save_plots : bool
        Whether to save the plots
    save_path_prefix : str, optional
        Path prefix for saving plots
    model_name : str
        Name of the model for saving plots
    device : str
        Device to use for computations
    metric : str
        Metric to use for evaluation: "ssim" or "tv" (Total Variation)
    mean_flag : bool
        Whether to return mean metrics or not
    zone_type : str or int
        Type of zones: 
        - "quadrants" for 2×2 grid
        - "center_bg" for center and background
        - integer n for n×n grid (e.g., 4 creates a 4×4 grid with 16 zones)
    max_frames : int, optional
        Maximum number of frames to plot. If None, all frames will be plotted.
    """
    import numpy as np
    import matplotlib.pyplot as plt
    import torch
    import os
    from matplotlib.patches import Rectangle
    from matplotlib.colors import LinearSegmentedColormap
    
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    print(f"Using metric: {metric}")
    
    if isinstance(zone_type, int):
        print(f"Using {zone_type}×{zone_type} grid zones ({zone_type*zone_type} total zones)")
    else:
        print(f"Using zone type: {zone_type}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # Create key mapping between predictions and videos
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        key_mapping = {}
        for pred_key in predictions.keys():
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Helper function to split frame into zones
    def split_into_zones(frame, zone_type="quadrants", center_ratio=0.5):
        """
        Split a frame into zones.
        """
        if isinstance(frame, torch.Tensor):
            C, H, W = frame.shape
        else:
            C, H, W = frame.shape
            
        zones = {}
        
        if zone_type == "quadrants":
            # Split into 4 quadrants (2×2 grid)
            h_mid = H // 2
            w_mid = W // 2
            
            zones["top_left"] = (slice(None), slice(0, h_mid), slice(0, w_mid))
            zones["top_right"] = (slice(None), slice(0, h_mid), slice(w_mid, W))
            zones["bottom_left"] = (slice(None), slice(h_mid, H), slice(0, w_mid))
            zones["bottom_right"] = (slice(None), slice(h_mid, H), slice(w_mid, W))
            
        elif zone_type == "center_bg":
            # Split into center and background
            h_center = int(H * center_ratio)
            w_center = int(W * center_ratio)
            
            h_start = (H - h_center) // 2
            h_end = h_start + h_center
            w_start = (W - w_center) // 2
            w_end = w_start + w_center
            
            zones["center"] = (slice(None), slice(h_start, h_end), slice(w_start, w_end))
            
            # Background is everything except the center
            center_mask = np.zeros((H, W), dtype=bool)
            center_mask[h_start:h_end, w_start:w_end] = True
            
            zones["background"] = {"mask": ~center_mask, 
                                   "bounds": (h_start, h_end, w_start, w_end)}
            
        elif isinstance(zone_type, int) and zone_type > 0:
            # Create an n×n grid where n = zone_type
            n = zone_type
            
            # Calculate heights of each section
            h_sections = [i * H // n for i in range(n+1)]
            w_sections = [i * W // n for i in range(n+1)]
            
            # Create zones for each grid cell
            for i in range(n):
                for j in range(n):
                    zone_name = f"grid_{i}_{j}"  # Row_Column naming
                    zones[zone_name] = (
                        slice(None),
                        slice(h_sections[i], h_sections[i+1]),
                        slice(w_sections[j], w_sections[j+1])
                    )
                    
        else:
            raise ValueError(f"Unknown zone type: {zone_type}")
            
        return zones
    
    # Helper function to calculate zone metrics
    def calculate_zone_metrics(orig_frame, pred_frame, zones, metric="ssim", device=device):
        """
        Calculate metrics for each zone.
        """
        from pytorch_msssim import ssim
        
        zone_metrics = {}
        
        # Convert to torch tensors if needed
        if not isinstance(orig_frame, torch.Tensor):
            orig_tensor = torch.from_numpy(orig_frame).unsqueeze(0)
        else:
            orig_tensor = orig_frame.unsqueeze(0)
            
        if not isinstance(pred_frame, torch.Tensor):
            pred_tensor = torch.from_numpy(pred_frame).unsqueeze(0)
        else:
            pred_tensor = pred_frame.unsqueeze(0)
        
        for zone_name, zone_slice in zones.items():
            # Special handling for background in center_bg mode
            if isinstance(zone_slice, dict):  # Background in center_bg mode
                mask = zone_slice["mask"]
                
                orig_zone = orig_tensor.clone()
                pred_zone = pred_tensor.clone()
                
                # Apply mask to all channels
                for c in range(orig_zone.shape[1]):  # For each channel
                    orig_zone[0, c][~mask] = 0
                    pred_zone[0, c][~mask] = 0
                
                # Calculate metric 
                if metric == "ssim":
                    zone_metrics[zone_name] = ssim(orig_zone, pred_zone, data_range=1, size_average=True).item()
                else:
                    # TV Loss calculation for masked region
                    tv_loss = torch.abs(pred_zone[:,:,1:,:] - pred_zone[:,:,:-1,:]).sum() + \
                              torch.abs(pred_zone[:,:,:,1:] - pred_zone[:,:,:,:-1]).sum()
                    # Normalize by number of pixels in the zone
                    zone_metrics[zone_name] = tv_loss.item() / mask.sum()
                
            else:  # Normal zones
                # Get the zone data
                orig_zone = orig_tensor[0][zone_slice].unsqueeze(0)
                pred_zone = pred_tensor[0][zone_slice].unsqueeze(0)
                
                if metric == "ssim":
                    zone_metrics[zone_name] = ssim(orig_zone, pred_zone, data_range=1, size_average=True).item()
                else:
                    # TV Loss calculation
                    tv_loss = torch.abs(pred_zone[:,:,1:,:] - pred_zone[:,:,:-1,:]).sum() + \
                              torch.abs(pred_zone[:,:,:,1:] - pred_zone[:,:,:,:-1]).sum()
                    # Normalize by number of pixels in the zone
                    zone_metrics[zone_name] = tv_loss.item() / (orig_zone.shape[2] * orig_zone.shape[3])
        
        return zone_metrics
    
    # Helper function for normalizing images for display
    def normalize(img):
        """Normalize image for display"""
        if isinstance(img, torch.Tensor):
            img = img.detach().cpu().numpy()
        
        img = img.copy()
        if img.min() < 0:
            img = (img + 1) / 2  # [-1, 1] -> [0, 1]
        return np.clip(img, 0, 1)
    
    # Initialize metrics
    all_zone_metrics = {}  # Will store metrics for each prediction key
    metric_name = "SSIM" if metric == "ssim" else "TV Loss"
    
    # Create a TV loss calculator if needed
    if metric == "tv":
        tv_calculator = TotalVariation().to(device)
    
    # ===== PART 1: Calculate zone metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        N = video.shape[0]
        
        # Store metrics for all frames in this prediction
        pred_metrics = []
        
        # Calculate metrics for each frame
        for i in range(N):
            # Get zones for this frame
            zones = split_into_zones(video[i], zone_type=zone_type)
            
            # Calculate metrics for each zone
            try:
                zone_metrics = calculate_zone_metrics(video[i], prediction[i], zones, metric=metric)
                pred_metrics.append(zone_metrics)
            except Exception as e:
                print(f"Error calculating zone metrics for {pred_key}, frame {i}: {e}")
                # Create empty metrics
                if zone_type == "quadrants":
                    pred_metrics.append({
                        "top_left": 0, "top_right": 0, 
                        "bottom_left": 0, "bottom_right": 0
                    })
                elif zone_type == "center_bg":
                    pred_metrics.append({"center": 0, "background": 0})
                elif isinstance(zone_type, int):
                    empty_metrics = {}
                    for ii in range(zone_type):
                        for jj in range(zone_type):
                            empty_metrics[f"grid_{ii}_{jj}"] = 0
                    pred_metrics.append(empty_metrics)
        
        # Store metrics for this prediction
        all_zone_metrics[pred_key] = pred_metrics
        
        # Print average metrics for this prediction
        print(f"\nAverage {metric_name} for {pred_key} by zone:")
        
        # Calculate and print mean metrics across frames for each zone
        if len(pred_metrics) > 0:
            zone_names = list(pred_metrics[0].keys())
            
            for zone in zone_names:
                zone_values = [metrics[zone] for metrics in pred_metrics]
                mean_zone = np.mean(zone_values)
                print(f"  - {zone}: {mean_zone:.3f}")
    
    # ===== PART 2: Plot the original image, reconstructed image, and heatmap side by side =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # Determine the frames to plot
        if max_frames is not None and max_frames < N:
            # Evenly sample frames if max_frames is specified
            indices = np.linspace(0, N-1, max_frames, dtype=int)
        else:
            # Plot all frames
            indices = np.arange(N)
        
        print(f"\nPlotting {len(indices)} frames with Original, Reconstruction, and Heatmap")
        
        # REORDERING: Create an ordered list of prediction keys with "original_combined" first
        ordered_keys = []
        for key in key_mapping.keys():
            if key != "original_combined":
                ordered_keys.append(key)
        
        # If "original_combined" exists, insert it at the beginning of the list
        if "original_combined" in key_mapping:
            ordered_keys.insert(0, "original_combined")
        
        # For each frame index
        for frame_idx in indices:
            # Plot original reference frame
            ref_frame = ref_video[frame_idx]
            
            # For each prediction
            for pred_key in ordered_keys:
                try:
                    video_key = key_mapping[pred_key]
                    pred_frame = predictions[pred_key][frame_idx]
                    
                    # Get zone metrics for this prediction
                    zone_metrics = all_zone_metrics[pred_key][frame_idx]
                    
                    # Create a figure with three subplots: Original, Prediction, Heatmap
                    fig, axes = plt.subplots(1, 3, figsize=(15, 5), 
                                           gridspec_kw={'width_ratios': [1, 1, 1]})  # Equal width for all panels
                    
                    # 1. Original Frame (left)
                    axes[0].imshow(np.transpose(normalize(ref_frame), (1, 2, 0)))
                    axes[0].set_title(f"Original Frame {frame_idx}")
                    axes[0].axis('off')
                    
                    # 2. Reconstructed Frame (middle)
                    axes[1].imshow(np.transpose(normalize(pred_frame), (1, 2, 0)))
                    axes[1].set_title(f"{pred_key} (Frame {frame_idx})")
                    axes[1].axis('off')
                    
                    # 3. Heatmap of metrics (right)
                    if isinstance(zone_type, int):
                        # For n×n grid, create a grid to display metrics
                        grid_values = np.zeros((zone_type, zone_type))
                        
                        for i in range(zone_type):
                            for j in range(zone_type):
                                zone_name = f"grid_{i}_{j}"
                                grid_values[i, j] = zone_metrics.get(zone_name, 0)
                        
                        # Show the heatmap
                        im = axes[2].imshow(grid_values, cmap='viridis', interpolation='nearest')
                        
                        # Add text labels with adjustable font size
                        fontsize = max(6, min(10, 16 - zone_type))  # Scale font size based on grid density
                        for i in range(zone_type):
                            for j in range(zone_type):
                                text = axes[2].text(j, i, f"{grid_values[i, j]:.3f}",
                                                 ha="center", va="center", color="white",
                                                 fontsize=fontsize, fontweight='bold')
                        
                        # Add colorbar
                        cbar = plt.colorbar(im, ax=axes[2])
                        cbar.set_label(metric_name)
                        
                        axes[2].set_title(f"Zone {metric_name}")
                        
                    elif zone_type == "quadrants":
                        # For quadrants, create a 2×2 heatmap
                        quadrant_values = np.zeros((2, 2))
                        quadrant_values[0, 0] = zone_metrics.get("top_left", 0)
                        quadrant_values[0, 1] = zone_metrics.get("top_right", 0)
                        quadrant_values[1, 0] = zone_metrics.get("bottom_left", 0)
                        quadrant_values[1, 1] = zone_metrics.get("bottom_right", 0)
                        
                        # Show the heatmap
                        im = axes[2].imshow(quadrant_values, cmap='viridis', interpolation='nearest')
                        
                        # Add text labels
                        axes[2].text(0, 0, f"{quadrant_values[0, 0]:.3f}", ha="center", va="center", color="white")
                        axes[2].text(1, 0, f"{quadrant_values[0, 1]:.3f}", ha="center", va="center", color="white")
                        axes[2].text(0, 1, f"{quadrant_values[1, 0]:.3f}", ha="center", va="center", color="white")
                        axes[2].text(1, 1, f"{quadrant_values[1, 1]:.3f}", ha="center", va="center", color="white")
                        
                        # Add colorbar
                        cbar = plt.colorbar(im, ax=axes[2])
                        cbar.set_label(metric_name)
                        
                        axes[2].set_title(f"Quadrant {metric_name}")
                        
                    elif zone_type == "center_bg":
                        # For center/background, special visualization
                        center_val = zone_metrics.get("center", 0)
                        bg_val = zone_metrics.get("background", 0)
                        
                        # Create a mask-based visualization
                        mask = np.zeros((3, 3), dtype=bool)
                        mask[1, 1] = True  # Center is True, background is False
                        
                        # Create values array where center has one value, background another
                        values = np.ones((3, 3)) * bg_val
                        values[1, 1] = center_val
                        
                        # Show the heatmap
                        im = axes[2].imshow(values, cmap='viridis', interpolation='nearest')
                        
                        # Add text labels
                        axes[2].text(1, 1, f"Center\n{center_val:.3f}", ha="center", va="center", color="white")
                        axes[2].text(0, 0, f"BG\n{bg_val:.3f}", ha="center", va="center", color="white")
                        
                        # Add colorbar
                        cbar = plt.colorbar(im, ax=axes[2])
                        cbar.set_label(metric_name)
                        
                        axes[2].set_title(f"Center/Background {metric_name}")
                    
                    plt.tight_layout()
                    
                    # Save figure if requested
                    if save_plots and save_path_prefix:
                        zone_type_str = zone_type if isinstance(zone_type, str) else f"grid{zone_type}x{zone_type}"
                        fig_path = f"{save_path_prefix}{pred_key}_frame{frame_idx}_{metric}_{zone_type_str}.png"
                        plt.savefig(fig_path, bbox_inches='tight', dpi=300)
                    
                    plt.show()
                    
                except Exception as e:
                    print(f"Error creating visualization for {pred_key} (Frame {frame_idx}): {e}")
    
    # Calculate overall mean metric
    overall_mean = 0
    if len(all_zone_metrics) > 0:
        # Average across all predictions and all zones
        all_values = []
        for pred_metrics in all_zone_metrics.values():
            for metrics in pred_metrics:
                all_values.extend(list(metrics.values()))
        
        if all_values:
            overall_mean = np.mean(all_values)
    
    # Update performance dictionary if provided
    if performance_dict is not None:
        try:
            # Calculate overall metrics across all zones
            for pred_key, pred_metrics in all_zone_metrics.items():
                if len(pred_metrics) > 0:
                    zone_names = list(pred_metrics[0].keys())
                    
                    for zone in zone_names:
                        zone_values = [metrics[zone] for metrics in pred_metrics]
                        zone_mean = np.mean(zone_values)
                        zone_median = np.median(zone_values)
                        
                        # Add to performance dict
                        if metric == "ssim":
                            performance_dict[f'mean_ssim_{zone}_D'] = zone_mean
                            performance_dict[f'median_ssim_{zone}_D'] = zone_median
                        else:
                            performance_dict[f'mean_tv_{zone}_D'] = zone_mean
                            performance_dict[f'median_tv_{zone}_D'] = zone_median
            
            # Add overall mean metric
            if metric == "ssim":
                performance_dict['mean_ssim_D'] = overall_mean
            else:
                performance_dict['mean_tv_D'] = overall_mean
                
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    if mean_flag:
        return overall_mean
    
    return performance_dict

# Example usage:
# plot_all_predictions6(predictions, videos, metric="ssim", zone_type="quadrants")
# plot_all_predictions6(predictions, videos, metric="ssim", zone_type="center_bg")

# Example function to calculate total variation loss (similar to what was in the original function)
class TotalVariation(torch.nn.Module):
    def __init__(self):
        super(TotalVariation, self).__init__()
    
    def forward(self, x):
        """Calculate total variation loss"""
        tv_h = torch.abs(x[:, :, 1:, :] - x[:, :, :-1, :]).sum()
        tv_w = torch.abs(x[:, :, :, 1:] - x[:, :, :, :-1]).sum()
        return tv_h + tv_w



def plot_all_predictions7(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, 
                 save_path_prefix=None, model_name="", device="cuda" if torch.cuda.is_available() else "cpu",
                 metric="ssim", mean_flag=False, zone_type="quadrants", max_frames=None, baseline_predictions=None):
    """
    Display comparison plots between original videos and multiple corresponding predictions.
    This version shows: Original image, reconstructed image, and heatmap of loss values or differences from baseline.
    
    Parameters:
    -----------
    predictions : dict
        Dictionary of prediction arrays
    videos : dict
        Dictionary of ground truth video arrays
    performance_dict : dict, optional
        Dictionary to store performance metrics
    display_plots : bool
        Whether to display the plots
    save_plots : bool
        Whether to save the plots
    save_path_prefix : str, optional
        Path prefix for saving plots
    model_name : str
        Name of the model for saving plots
    device : str
        Device to use for computations
    metric : str
        Metric to use for evaluation: "ssim" or "tv" (Total Variation)
    mean_flag : bool
        Whether to return mean metrics or not
    zone_type : str or int
        Type of zones: 
        - "quadrants" for 2×2 grid
        - "center_bg" for center and background
        - integer n for n×n grid (e.g., 4 creates a 4×4 grid with 16 zones)
    max_frames : int, optional
        Maximum number of frames to plot. If None, all frames will be plotted.
    baseline_predictions : dict, optional
        Dictionary of baseline prediction arrays (without perturbation)
    """
    import numpy as np
    import matplotlib.pyplot as plt
    import torch
    import os
    from matplotlib.patches import Rectangle
    from matplotlib.colors import LinearSegmentedColormap, Normalize
    
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        os.makedirs(save_path_prefix, exist_ok=True)

    # Debug information about inputs
    print("\n=== DEBUG INFO ===")
    print(f"Predictions dictionary contains {len(predictions)} keys: {list(predictions.keys())}")
    print(f"Videos dictionary contains {len(videos)} keys: {list(videos.keys())}")
    print(f"Using metric: {metric}")
    print(f"Baseline predictions provided: {baseline_predictions is not None}")
    
    if isinstance(zone_type, int):
        print(f"Using {zone_type}×{zone_type} grid zones ({zone_type*zone_type} total zones)")
    else:
        print(f"Using zone type: {zone_type}")
    
    # Find the overlapping keys between predictions and videos
    common_keys = [key for key in predictions.keys() if key in videos]
    print(f"Common keys in both dictionaries: {common_keys}")
    
    # Create key mapping between predictions and videos
    if not common_keys and len(videos) > 0:
        print("No common keys found. Trying to match prediction keys to video keys.")
        ref_video_key = list(videos.keys())[0]
        print(f"Using {ref_video_key} as reference video for all predictions")
        key_mapping = {pred_key: ref_video_key for pred_key in predictions.keys()}
    else:
        key_mapping = {}
        for pred_key in predictions.keys():
            if pred_key in videos:
                key_mapping[pred_key] = pred_key
            else:
                matched = False
                for video_key in videos.keys():
                    if video_key in pred_key:
                        key_mapping[pred_key] = video_key
                        matched = True
                        break
                if not matched and len(videos) > 0:
                    key_mapping[pred_key] = list(videos.keys())[0]
    
    print(f"Key mapping from prediction keys to video keys: {key_mapping}")
    
    # Helper function to split frame into zones
    def split_into_zones(frame, zone_type="quadrants", center_ratio=0.5):
        """
        Split a frame into zones.
        """
        if isinstance(frame, torch.Tensor):
            C, H, W = frame.shape
        else:
            C, H, W = frame.shape
            
        zones = {}
        
        if zone_type == "quadrants":
            # Split into 4 quadrants (2×2 grid)
            h_mid = H // 2
            w_mid = W // 2
            
            zones["top_left"] = (slice(None), slice(0, h_mid), slice(0, w_mid))
            zones["top_right"] = (slice(None), slice(0, h_mid), slice(w_mid, W))
            zones["bottom_left"] = (slice(None), slice(h_mid, H), slice(0, w_mid))
            zones["bottom_right"] = (slice(None), slice(h_mid, H), slice(w_mid, W))
            
        elif zone_type == "center_bg":
            # Split into center and background
            h_center = int(H * center_ratio)
            w_center = int(W * center_ratio)
            
            h_start = (H - h_center) // 2
            h_end = h_start + h_center
            w_start = (W - w_center) // 2
            w_end = w_start + w_center
            
            zones["center"] = (slice(None), slice(h_start, h_end), slice(w_start, w_end))
            
            # Background is everything except the center
            center_mask = np.zeros((H, W), dtype=bool)
            center_mask[h_start:h_end, w_start:w_end] = True
            
            zones["background"] = {"mask": ~center_mask, 
                                   "bounds": (h_start, h_end, w_start, w_end)}
            
        elif isinstance(zone_type, int) and zone_type > 0:
            # Create an n×n grid where n = zone_type
            n = zone_type
            
            # Calculate heights of each section
            h_sections = [i * H // n for i in range(n+1)]
            w_sections = [i * W // n for i in range(n+1)]
            
            # Create zones for each grid cell
            for i in range(n):
                for j in range(n):
                    zone_name = f"grid_{i}_{j}"  # Row_Column naming
                    zones[zone_name] = (
                        slice(None),
                        slice(h_sections[i], h_sections[i+1]),
                        slice(w_sections[j], w_sections[j+1])
                    )
                    
        else:
            raise ValueError(f"Unknown zone type: {zone_type}")
            
        return zones
    
    # Helper function to calculate zone metrics
    def calculate_zone_metrics(orig_frame, pred_frame, baseline_frame=None, zones=None, metric="ssim", device=device):
        """
        Calculate metrics for each zone.
        If baseline_frame is provided, calculate the difference: baseline_metrics - pred_metrics
        """
        from pytorch_msssim import ssim
        
        # If zones not provided, calculate them
        if zones is None:
            zones = split_into_zones(orig_frame, zone_type=zone_type)
        
        zone_metrics = {}
        
        # Convert to torch tensors if needed
        if not isinstance(orig_frame, torch.Tensor):
            orig_tensor = torch.from_numpy(orig_frame).unsqueeze(0)
        else:
            orig_tensor = orig_frame.unsqueeze(0)
            
        if not isinstance(pred_frame, torch.Tensor):
            pred_tensor = torch.from_numpy(pred_frame).unsqueeze(0)
        else:
            pred_tensor = pred_frame.unsqueeze(0)
        
        # Process baseline frame if provided    
        if baseline_frame is not None:
            if not isinstance(baseline_frame, torch.Tensor):
                baseline_tensor = torch.from_numpy(baseline_frame).unsqueeze(0)
            else:
                baseline_tensor = baseline_frame.unsqueeze(0)
        
        # Calculate metrics for each zone
        for zone_name, zone_slice in zones.items():
            # Special handling for background in center_bg mode
            if isinstance(zone_slice, dict):  # Background in center_bg mode
                mask = zone_slice["mask"]
                
                orig_zone = orig_tensor.clone()
                pred_zone = pred_tensor.clone()
                
                # Apply mask to all channels
                for c in range(orig_zone.shape[1]):  # For each channel
                    orig_zone[0, c][~mask] = 0
                    pred_zone[0, c][~mask] = 0
                
                # Calculate metric for prediction
                if metric == "ssim":
                    pred_metric = ssim(orig_zone, pred_zone, data_range=1, size_average=True).item()
                else:
                    # TV Loss calculation for masked region
                    tv_loss = torch.abs(pred_zone[:,:,1:,:] - pred_zone[:,:,:-1,:]).sum() + \
                              torch.abs(pred_zone[:,:,:,1:] - pred_zone[:,:,:,:-1]).sum()
                    # Normalize by number of pixels in the zone
                    pred_metric = tv_loss.item() / mask.sum()
                
                # If baseline provided, calculate baseline metric and difference
                if baseline_frame is not None:
                    baseline_zone = baseline_tensor.clone()
                    for c in range(baseline_zone.shape[1]):
                        baseline_zone[0, c][~mask] = 0
                        
                    if metric == "ssim":
                        base_metric = ssim(orig_zone, baseline_zone, data_range=1, size_average=True).item()
                        # For SSIM, higher is better, so baseline - perturbed shows how much we lost
                        # (negative value means perturbation improved SSIM)
                        zone_metrics[zone_name] = base_metric - pred_metric
                    else:
                        # TV Loss
                        tv_loss = torch.abs(baseline_zone[:,:,1:,:] - baseline_zone[:,:,:-1,:]).sum() + \
                                 torch.abs(baseline_zone[:,:,:,1:] - baseline_zone[:,:,:,:-1]).sum()
                        base_metric = tv_loss.item() / mask.sum()
                        # For TV loss, lower is better, so perturbed - baseline shows how much we lost
                        # (positive value means perturbation worsened TV loss)
                        zone_metrics[zone_name] = pred_metric - base_metric
                else:
                    # No baseline, just use the prediction metric
                    zone_metrics[zone_name] = pred_metric
                
            else:  # Normal zones
                # Get the zone data
                orig_zone = orig_tensor[0][zone_slice].unsqueeze(0)
                pred_zone = pred_tensor[0][zone_slice].unsqueeze(0)
                
                # Calculate metric for prediction
                if metric == "ssim":
                    pred_metric = ssim(orig_zone, pred_zone, data_range=1, size_average=True).item()
                else:
                    # TV Loss calculation
                    tv_loss = torch.abs(pred_zone[:,:,1:,:] - pred_zone[:,:,:-1,:]).sum() + \
                              torch.abs(pred_zone[:,:,:,1:] - pred_zone[:,:,:,:-1]).sum()
                    # Normalize by number of pixels in the zone
                    pred_metric = tv_loss.item() / (orig_zone.shape[2] * orig_zone.shape[3])
                
                # If baseline provided, calculate baseline metric and difference
                if baseline_frame is not None:
                    baseline_zone = baseline_tensor[0][zone_slice].unsqueeze(0)
                    
                    if metric == "ssim":
                        base_metric = ssim(orig_zone, baseline_zone, data_range=1, size_average=True).item()
                        # For SSIM, higher is better, so baseline - perturbed shows how much we lost
                        zone_metrics[zone_name] = base_metric - pred_metric
                    else:
                        # TV Loss
                        tv_loss = torch.abs(baseline_zone[:,:,1:,:] - baseline_zone[:,:,:-1,:]).sum() + \
                                 torch.abs(baseline_zone[:,:,:,1:] - baseline_zone[:,:,:,:-1]).sum()
                        base_metric = tv_loss.item() / (baseline_zone.shape[2] * baseline_zone.shape[3])
                        # For TV loss, lower is better, so perturbed - baseline shows how much we lost
                        zone_metrics[zone_name] = pred_metric - base_metric
                else:
                    # No baseline, just use the prediction metric
                    zone_metrics[zone_name] = pred_metric
        
        return zone_metrics
    
    # Helper function for normalizing images for display
    def normalize(img):
        """Normalize image for display"""
        if isinstance(img, torch.Tensor):
            img = img.detach().cpu().numpy()
        
        img = img.copy()
        if img.min() < 0:
            img = (img + 1) / 2  # [-1, 1] -> [0, 1]
        return np.clip(img, 0, 1)
    
    # Initialize metrics
    all_zone_metrics = {}  # Will store metrics for each prediction key
    
    # Get appropriate metric name
    if baseline_predictions is not None:
        metric_description = "Difference from Baseline" 
        if metric == "ssim":
            metric_name = "SSIM Difference"
        else:
            metric_name = "TV Loss Difference"
    else:
        metric_name = "SSIM" if metric == "ssim" else "TV Loss"
        metric_description = metric_name
    
    # Create a TV loss calculator if needed
    if metric == "tv":
        tv_calculator = TotalVariation().to(device)
    
    # ===== PART 1: Calculate zone metrics for each prediction =====
    for pred_key, video_key in key_mapping.items():
        prediction = predictions[pred_key]
        video = videos[video_key][..., 15]  # Middle frame
        
        # Get baseline prediction if available
        baseline_pred = None
        if baseline_predictions is not None:
            # Find the appropriate key in baseline predictions
            baseline_keys = list(baseline_predictions.keys())
            if pred_key in baseline_predictions:
                baseline_pred = baseline_predictions[pred_key]
            elif len(baseline_keys) > 0:
                # If exact key not found, use first available baseline key
                baseline_pred = baseline_predictions[baseline_keys[0]]
                print(f"Using {baseline_keys[0]} as baseline for {pred_key}")
        
        # Check shapes
        print(f"Prediction {pred_key} shape: {prediction.shape}")
        print(f"Video {video_key} shape: {video.shape}")
        if baseline_pred is not None:
            print(f"Baseline prediction shape: {baseline_pred.shape}")
        
        # Ensure prediction and video have compatible shapes
        if prediction.shape[0] != video.shape[0]:
            print(f"Warning: Shape mismatch for {pred_key} vs {video_key}. Skipping.")
            continue
        
        # If baseline exists, ensure it has compatible shape
        if baseline_pred is not None and baseline_pred.shape[0] != prediction.shape[0]:
            print(f"Warning: Baseline shape {baseline_pred.shape} doesn't match prediction shape {prediction.shape}. Ignoring baseline.")
            baseline_pred = None
        
        N = video.shape[0]
        
        # Store metrics for all frames in this prediction
        pred_metrics = []
        
        # Calculate metrics for each frame
        for i in range(N):
            # Get zones for this frame
            zones = split_into_zones(video[i], zone_type=zone_type)
            
            # Calculate metrics for each zone
            try:
                # If baseline exists, calculate difference metrics
                if baseline_pred is not None:
                    zone_metrics = calculate_zone_metrics(
                        video[i], prediction[i], baseline_frame=baseline_pred[i], zones=zones, metric=metric
                    )
                else:
                    zone_metrics = calculate_zone_metrics(
                        video[i], prediction[i], zones=zones, metric=metric
                    )
                pred_metrics.append(zone_metrics)
            except Exception as e:
                print(f"Error calculating zone metrics for {pred_key}, frame {i}: {e}")
                # Create empty metrics
                if zone_type == "quadrants":
                    pred_metrics.append({
                        "top_left": 0, "top_right": 0, 
                        "bottom_left": 0, "bottom_right": 0
                    })
                elif zone_type == "center_bg":
                    pred_metrics.append({"center": 0, "background": 0})
                elif isinstance(zone_type, int):
                    empty_metrics = {}
                    for ii in range(zone_type):
                        for jj in range(zone_type):
                            empty_metrics[f"grid_{ii}_{jj}"] = 0
                    pred_metrics.append(empty_metrics)
        
        # Store metrics for this prediction
        all_zone_metrics[pred_key] = pred_metrics
        
        # Print average metrics for this prediction
        print(f"\nAverage {metric_description} for {pred_key} by zone:")
        
        # Calculate and print mean metrics across frames for each zone
        if len(pred_metrics) > 0:
            zone_names = list(pred_metrics[0].keys())
            
            for zone in zone_names:
                zone_values = [metrics[zone] for metrics in pred_metrics]
                mean_zone = np.mean(zone_values)
                print(f"  - {zone}: {mean_zone:.4f}")
    
    # ===== PART 2: Plot the original image, reconstructed image, and heatmap side by side =====
    if display_plots and len(key_mapping) > 0:
        # Get a reference video key and shape
        ref_video_key = list(videos.keys())[0]
        ref_video = videos[ref_video_key][..., 15]
        N = ref_video.shape[0]
        
        # Determine the frames to plot
        if max_frames is not None and max_frames < N:
            # Evenly sample frames if max_frames is specified
            indices = np.linspace(0, N-1, max_frames, dtype=int)
        else:
            # Plot all frames
            indices = np.arange(N)
        
        print(f"\nPlotting {len(indices)} frames with Original, Reconstruction, and Heatmap")
        
        # REORDERING: Create an ordered list of prediction keys with "original_combined" first
        ordered_keys = []
        for key in key_mapping.keys():
            if key != "original_combined":
                ordered_keys.append(key)
        
        # If "original_combined" exists, insert it at the beginning of the list
        if "original_combined" in key_mapping:
            ordered_keys.insert(0, "original_combined")
        
        # For each frame index
        for frame_idx in indices:
            # Plot original reference frame
            ref_frame = ref_video[frame_idx]
            
            # For each prediction
            for pred_key in ordered_keys:
                try:
                    video_key = key_mapping[pred_key]
                    pred_frame = predictions[pred_key][frame_idx]
                    
                    # Get baseline frame if available
                    baseline_frame = None
                    if baseline_predictions is not None:
                        if pred_key in baseline_predictions:
                            baseline_frame = baseline_predictions[pred_key][frame_idx]
                        elif len(baseline_predictions) > 0:
                            # Use first available baseline prediction
                            first_key = list(baseline_predictions.keys())[0]
                            baseline_frame = baseline_predictions[first_key][frame_idx]
                    
                    # Get zone metrics for this prediction
                    zone_metrics = all_zone_metrics[pred_key][frame_idx]
                    
                    # Determine if we're showing differences or absolute values
                    is_difference = baseline_frame is not None
                    
                    # Create a figure with three subplots: Original, Prediction, Heatmap
                    fig, axes = plt.subplots(1, 3, figsize=(15, 5), 
                                           gridspec_kw={'width_ratios': [1, 1, 1]})  # Equal width for all panels
                    
                    # 1. Original Frame (left)
                    axes[0].imshow(np.transpose(normalize(ref_frame), (1, 2, 0)))
                    axes[0].set_title(f"Original Frame {frame_idx}")
                    axes[0].axis('off')
                    
                    # 2. Reconstructed Frame (middle)
                    axes[1].imshow(np.transpose(normalize(pred_frame), (1, 2, 0)))
                    title = f"{pred_key} (Frame {frame_idx})"
                    if baseline_frame is not None:
                        title += " (Perturbed)"
                    axes[1].set_title(title)
                    axes[1].axis('off')
                    
                    # 3. Heatmap of metrics (right)
                    # Determine colormap based on if we're showing differences
                    cmap_name = 'coolwarm' if is_difference else 'viridis'
                    
                    # Also determine normalization based on if we're showing differences
                    if is_difference:
                        # For differences, use symmetric normalization around zero
                        values = list(zone_metrics.values())
                        max_abs = max(abs(min(values)), abs(max(values))) if values else 1.0
                        norm = Normalize(vmin=-max_abs, vmax=max_abs)
                    else:
                        # For absolute values, use standard normalization
                        norm = None  # Let matplotlib handle it
                    
                    if isinstance(zone_type, int):
                        # For n×n grid, create a grid to display metrics
                        grid_values = np.zeros((zone_type, zone_type))
                        
                        for i in range(zone_type):
                            for j in range(zone_type):
                                zone_name = f"grid_{i}_{j}"
                                grid_values[i, j] = zone_metrics.get(zone_name, 0)
                        
                        # Show the heatmap
                        im = axes[2].imshow(grid_values, cmap=cmap_name, interpolation='nearest', norm=norm)
                        
                        # Add text labels with adjustable font size
                        fontsize = max(6, min(10, 16 - zone_type))  # Scale font size based on grid density
                        for i in range(zone_type):
                            for j in range(zone_type):
                                # Format the value based on magnitude
                                val = grid_values[i, j]
                                if abs(val) >= 0.01:
                                    text = f"{val:.3f}"
                                else:
                                    text = f"{val:.1e}"
                                    
                                axes[2].text(j, i, text,
                                           ha="center", va="center", color="white",
                                           fontsize=fontsize, fontweight='bold')
                        
                        # Add colorbar
                        cbar = plt.colorbar(im, ax=axes[2])
                        cbar.set_label(metric_name)
                        
                        axes[2].set_title(f"Zone {metric_description}")
                        
                    elif zone_type == "quadrants":
                        # For quadrants, create a 2×2 heatmap
                        quadrant_values = np.zeros((2, 2))
                        quadrant_values[0, 0] = zone_metrics.get("top_left", 0)
                        quadrant_values[0, 1] = zone_metrics.get("top_right", 0)
                        quadrant_values[1, 0] = zone_metrics.get("bottom_left", 0)
                        quadrant_values[1, 1] = zone_metrics.get("bottom_right", 0)
                        
                        # Show the heatmap
                        im = axes[2].imshow(quadrant_values, cmap=cmap_name, interpolation='nearest', norm=norm)
                        
                        # Add text labels
                        for i, j in [(0,0), (0,1), (1,0), (1,1)]:
                            val = quadrant_values[i, j]
                            if abs(val) >= 0.01:
                                text = f"{val:.3f}"
                            else:
                                text = f"{val:.1e}"
                            axes[2].text(j, i, text, ha="center", va="center", color="white", fontsize=10)
                        
                        # Add colorbar
                        cbar = plt.colorbar(im, ax=axes[2])
                        cbar.set_label(metric_name)
                        
                        axes[2].set_title(f"Quadrant {metric_description}")
                        
                    elif zone_type == "center_bg":
                        # For center/background, special visualization
                        center_val = zone_metrics.get("center", 0)
                        bg_val = zone_metrics.get("background", 0)
                        
                        # Create a mask-based visualization
                        mask = np.zeros((3, 3), dtype=bool)
                        mask[1, 1] = True  # Center is True, background is False
                        
                        # Create values array where center has one value, background another
                        values = np.ones((3, 3)) * bg_val
                        values[1, 1] = center_val
                        
                        # Show the heatmap
                        im = axes[2].imshow(values, cmap=cmap_name, interpolation='nearest', norm=norm)
                        
                        # Format values for display
                        if abs(center_val) >= 0.01:
                            center_text = f"Center\n{center_val:.3f}"
                        else:
                            center_text = f"Center\n{center_val:.1e}"
                            
                        if abs(bg_val) >= 0.01:
                            bg_text = f"BG\n{bg_val:.3f}"
                        else:
                            bg_text = f"BG\n{bg_val:.1e}"
                        
                        # Add text labels
                        axes[2].text(1, 1, center_text, ha="center", va="center", color="white", fontsize=10)
                        axes[2].text(0, 0, bg_text, ha="center", va="center", color="white", fontsize=10)
                        
                        # Add colorbar
                        cbar = plt.colorbar(im, ax=axes[2])
                        cbar.set_label(metric_name)
                        
                        axes[2].set_title(f"Center/Background {metric_description}")
                    
                    plt.tight_layout()
                    
                    # Save figure if requested
                    if save_plots and save_path_prefix:
                        zone_type_str = zone_type if isinstance(zone_type, str) else f"grid{zone_type}x{zone_type}"
                        type_str = "diff" if is_difference else "abs"
                        fig_path = f"{save_path_prefix}{pred_key}_frame{frame_idx}_{metric}_{zone_type_str}_{type_str}.png"
                        plt.savefig(fig_path, bbox_inches='tight', dpi=300)
                    
                    plt.show()
                    
                except Exception as e:
                    print(f"Error creating visualization for {pred_key} (Frame {frame_idx}): {e}")
    
    # Calculate overall mean metric
    overall_mean = 0
    if len(all_zone_metrics) > 0:
        # Average across all predictions and all zones
        all_values = []
        for pred_metrics in all_zone_metrics.values():
            for metrics in pred_metrics:
                all_values.extend(list(metrics.values()))
        
        if all_values:
            overall_mean = np.mean(all_values)
    
    # Update performance dictionary if provided
    if performance_dict is not None:
        try:
            # Calculate overall metrics across all zones
            for pred_key, pred_metrics in all_zone_metrics.items():
                if len(pred_metrics) > 0:
                    zone_names = list(pred_metrics[0].keys())
                    
                    for zone in zone_names:
                        zone_values = [metrics[zone] for metrics in pred_metrics]
                        zone_mean = np.mean(zone_values)
                        zone_median = np.median(zone_values)
                        
                        # Add to performance dict
                        if baseline_predictions is not None:
                            # This is a difference metric
                            if metric == "ssim":
                                performance_dict[f'diff_ssim_{zone}_D'] = zone_mean
                                performance_dict[f'median_diff_ssim_{zone}_D'] = zone_median
                            else:
                                performance_dict[f'diff_tv_{zone}_D'] = zone_mean
                                performance_dict[f'median_diff_tv_{zone}_D'] = zone_median
                        else:
                            # This is an absolute metric
                            if metric == "ssim":
                                performance_dict[f'mean_ssim_{zone}_D'] = zone_mean
                                performance_dict[f'median_ssim_{zone}_D'] = zone_median
                            else:
                                performance_dict[f'mean_tv_{zone}_D'] = zone_mean
                                performance_dict[f'median_tv_{zone}_D'] = zone_median
            
            # Add overall mean metric
            if baseline_predictions is not None:
                # This is a difference metric
                if metric == "ssim":
                    performance_dict['diff_ssim_D'] = overall_mean
                else:
                    performance_dict['diff_tv_D'] = overall_mean
            else:
                # This is an absolute metric
                if metric == "ssim":
                    performance_dict['mean_ssim_D'] = overall_mean
                else:
                    performance_dict['mean_tv_D'] = overall_mean
                
        except Exception as e:
            print(f"Error updating performance dictionary: {e}")
    
    if mean_flag:
        return overall_mean
    
    return performance_dict


'''
#to plot all frames not just one every 3 or 15 frames
def plot_all_predictions(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, save_path_prefix=None, model_name=""):
#def plot_decoder_predictions(predictions, videos, performance_dict = None, display_plots=True):
    """
    Display comparison plots between original videos and their corresponding predictions.

    Parameters:
        predictions (dict): A dictionary containing predicted frames, indexed by video names.
        videos (dict): A dictionary containing original video frames, indexed by video names.

    This function iterates through each video key, retrieves frames from both the original and predicted data,
    and plots them side by side for visual comparison. It supports different intervals for videos with a large
    number of frames, adjusting the subplot layout accordingly.
    """
    # Create output directory if it doesn't exist
    if save_plots and save_path_prefix:
        import os
        os.makedirs(save_path_prefix, exist_ok=True)

    total_ssim, total_mse = [], []
    all_ssim = {'all': None}
    all_mse = {'all': None}

    for key in videos.keys():
        prediction = predictions[key]
        video = videos[key][..., 15]
        N = video.shape[0]

        ssim_frames = np.array([
            ssim(torch.from_numpy(video[i]).unsqueeze(0), torch.from_numpy(prediction[i]).unsqueeze(0), data_range=1, size_average=True).item()
            for i in range(N) 
        ])
        mean_ssim_frames = np.mean(ssim_frames)
        median_ssim_frames = np.median(ssim_frames)
        std_ssim_frames = np.std(ssim_frames)
        all_ssim[key] = ssim_frames

        mse_frames = np.array([
            F.mse_loss(normalize(torch.from_numpy(video[i])), normalize(torch.from_numpy(prediction[i]))).numpy()
            for i in range(N) 
        ])
        mean_mse_frames = np.mean(mse_frames)
        median_mse_frames = np.median(mse_frames)
        std_mse_frames = np.std(mse_frames)
        all_mse[key] = mse_frames

        total_ssim = np.concatenate((total_ssim, ssim_frames))
        total_mse = np.concatenate((total_mse, mse_frames))

        if display_plots:
            print(f'### {key} ({N}) ###')
            plt.figure(figsize=(17, 3))
            plt.subplot(121)
            plt.hist(ssim_frames, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
            sns.kdeplot(ssim_frames, fill=True, color='b', label='KDE')
            plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
            plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
            plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
            plt.xlabel('SSIM')
            plt.ylabel('Density')
            plt.title(f'Distribution of Structural Similarity per Frame (N = {N})')
            plt.legend() 
            plt.subplot(122)
            plt.hist(mse_frames, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
            sns.kdeplot(mse_frames, fill=True, color='r', label='KDE')
            plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
            plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
            plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
            plt.xlabel('MSE')
            plt.ylabel('Density')
            plt.title(f'Distribution of MSE per Frame (N = {N})')
            plt.legend() 
            # Save the loss histogram plot if requested
            if save_plots and save_path_prefix:
                losses_path = f"{save_path_prefix}losses_{key}_{model_name}.png"
                plt.savefig(losses_path, bbox_inches='tight', dpi=300)
#            plt.show()
            plt.show()
            
            indices = np.arange(0, prediction.shape[0], 1)  # Show every frame
            print(f'Display {len(indices)} frames:', indices)
            col = int(indices.shape[0] / 3) + 1
            k = 0
    
            if len(indices) < 20:
                plt.figure(figsize=(8, 10))
            elif len(indices) >= 20 and len(indices) < 40:
                plt.figure(figsize=(10, 15))
            else:
                plt.figure(figsize=(15, 20))
                
            for i in indices:
                k += 1
                plt.subplot(col,6,k)
                plt.imshow(np.transpose(normalize(video[i]), (1, 2, 0)))
                plt.axis('off')
                k += 1
                plt.subplot(col,6,k)
                plt.imshow(np.transpose(normalize(prediction[i]), (1, 2, 0)))
                plt.axis('off')

            # Save the image comparison plot if requested
            if save_plots and save_path_prefix:
                img_path = f"{save_path_prefix}{key}_{model_name}.png"
                plt.savefig(img_path, bbox_inches='tight', dpi=300)
            plt.show()
            del prediction, video, indices
#            plt.show()
#            del prediction, video, indices

    mean_ssim_frames = np.mean(total_ssim)
    median_ssim_frames = np.median(total_ssim)
    std_ssim_frames = np.std(total_ssim)

    mean_mse_frames = np.mean(total_mse)
    median_mse_frames = np.median(total_mse)
    std_mse_frames = np.std(total_mse)

    all_ssim['all'] = total_ssim
    all_mse['all'] = total_mse

    if display_plots:
        print(f'### all ({total_ssim.shape[0]}) ###')
        plt.figure(figsize=(17, 3))
        plt.subplot(121)
        plt.hist(total_ssim, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_ssim, fill=True, color='b', label='KDE')
        plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
        plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
        plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
        plt.xlabel('SSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Structural Similarity per Frame (N = {total_ssim.shape[0]})')
        plt.legend() 
        plt.subplot(122)
        plt.hist(total_mse, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(total_mse, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
        plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
        plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE per Frame (N = {total_mse.shape[0]})')
        plt.legend()
        # Save the all-videos metrics plot if requested
        if save_plots and save_path_prefix:
            all_path = f"{save_path_prefix}losses_all_{model_name}.png"
            plt.savefig(all_path, bbox_inches='tight', dpi=300)
#        plt.show()
        plt.show()

    if performance_dict:
        performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_ssim, 'SSIM', display_plots)
        performance_dict['mean_mse_D'], performance_dict['median_mse_D'] = plot_scatter_metrics(all_mse, 'MSE', display_plots)
        return performance_dict

'''



#used with
#results['test_performance'] = plot_decoder_predictions(results['decoder_predictions'], labels_dict, results['test_performance'], display_plots, save_plots=save_plots, save_path_prefix='outputs/' if save_plots else None, model_name=model_name)

def simple_plot_decoder_predictions(predictions, videos, performance_dict=None, display_plots=True, save_plots=False, save_path_prefix=None, model_name=""):

    total_ssim, total_mse = [], []
    all_ssim = {'all': None}
    all_mse = {'all': None}

    #video = videos[..., 15]
    #total_elements = video.size
    #random_indices = np.random.choice(total_elements, size=30, replace=False)
    #multi_indices = np.unravel_index(random_indices, video.shape)
    #video = video[multi_indices]
    #nah I need to do this in the test_model function because I need the predictions to be just of this as well

    for key in videos.keys():               #i could change this to use with training data to get just like 20 random frames off the whole set
        prediction = predictions[key]
        video = videos[key][..., 15]    #middle frame 
        #video becomes only one frame (middle frame) for each chunk of 32 frames

        #trainset -> videos (shape: (4321, 3, 112, 112, 32))
        #I could do video = videos[..., 15] and use only 30 random chunks from it (30 random elements)  (or do that first and then video = videos[..., 15])
        #at a later stage I could change it to present instead of 30 random elements, the 20 elements with the best ssim and the 20 elements with worst ssim or something, though it would take a lot more to run probably because I would need to compute ssim for all frames, if only I could store that in the training it would be quicker
        N = video.shape[0]

        ssim_frames = np.array([
            ssim(torch.from_numpy(video[i]).unsqueeze(0), torch.from_numpy(prediction[i]).unsqueeze(0), data_range=1, size_average=True).item()
            for i in range(N) 
        ])
        mean_ssim_frames = np.mean(ssim_frames)
        median_ssim_frames = np.median(ssim_frames)
        std_ssim_frames = np.std(ssim_frames)
        all_ssim[key] = ssim_frames

        mse_frames = np.array([
            F.mse_loss(normalize(torch.from_numpy(video[i])), normalize(torch.from_numpy(prediction[i]))).numpy()
            for i in range(N) 
        ])
        mean_mse_frames = np.mean(mse_frames)
        median_mse_frames = np.median(mse_frames)
        std_mse_frames = np.std(mse_frames)
        all_mse[key] = mse_frames

        total_ssim = np.concatenate((total_ssim, ssim_frames))
        total_mse = np.concatenate((total_mse, mse_frames))

        if display_plots:
            print(f'### {key} ({N}) ###')
            plt.figure(figsize=(17, 3))
            plt.subplot(121)
            plt.hist(ssim_frames, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
            sns.kdeplot(ssim_frames, fill=True, color='b', label='KDE')
            plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
            plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
            plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
            plt.xlabel('SSIM')
            plt.ylabel('Density')
            plt.title(f'Distribution of Structural Similarity per Frame (N = {N})')
            plt.legend() 
            plt.subplot(122)
            plt.hist(mse_frames, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
            sns.kdeplot(mse_frames, fill=True, color='r', label='KDE')
            plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
            plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
            plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
            plt.xlabel('MSE')
            plt.ylabel('Density')
            plt.title(f'Distribution of MSE per Frame (N = {N})')
            plt.legend() 
            # Save the loss histogram plot if requested
            if save_plots and save_path_prefix:
                losses_path = f"{save_path_prefix}losses_{key}_{model_name}.png"
                plt.savefig(losses_path, bbox_inches='tight', dpi=300)
#            plt.show()
            plt.show()
            
            if prediction.shape[0] > 400:
                indices = np.arange(0, prediction.shape[0], 15)
            else:
                indices = np.arange(0, prediction.shape[0], 3)
            print(f'Display {len(indices)} frames:', indices)
            col = int(indices.shape[0] / 3) + 1
            k = 0
    
            if len(indices) < 20:
                plt.figure(figsize=(8, 10))
            elif len(indices) >= 20 and len(indices) < 40:
                plt.figure(figsize=(10, 15))
            else:
                plt.figure(figsize=(15, 20))
                
            for i in indices:
                k += 1
                plt.subplot(col,6,k)
                plt.imshow(np.transpose(normalize(video[i]), (1, 2, 0)))
                plt.axis('off')
                k += 1
                plt.subplot(col,6,k)
                plt.imshow(np.transpose(normalize(prediction[i]), (1, 2, 0)))
                plt.axis('off')

            # Save the image comparison plot if requested
            if save_plots and save_path_prefix:
                img_path = f"{save_path_prefix}{key}_{model_name}.png"
                plt.savefig(img_path, bbox_inches='tight', dpi=300)
            plt.show()
            del prediction, video, indices
#            plt.show()
#            del prediction, video, indices

    mean_ssim_frames = np.mean(total_ssim)
    median_ssim_frames = np.median(total_ssim)
    std_ssim_frames = np.std(total_ssim)

    mean_mse_frames = np.mean(total_mse)
    median_mse_frames = np.median(total_mse)
    std_mse_frames = np.std(total_mse)

    all_ssim['all'] = total_ssim
    all_mse['all'] = total_mse

    if display_plots:
        print(f'### all ({total_ssim.shape[0]}) ###')
        plt.figure(figsize=(17, 3))
        plt.subplot(121)
        plt.hist(total_ssim, bins=30, alpha=0.5, color='blue', density=True, label='Histogram')
        sns.kdeplot(total_ssim, fill=True, color='b', label='KDE')
        plt.axvline(x=median_ssim_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_ssim_frames:.3f}')
        plt.axvline(x=mean_ssim_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_ssim_frames:.3f}')
        plt.axvspan(mean_ssim_frames - std_ssim_frames, mean_ssim_frames + std_ssim_frames, color='red', alpha=0.05, label=f'std: {std_ssim_frames:.3f}')
        plt.xlabel('SSIM')
        plt.ylabel('Density')
        plt.title(f'Distribution of Structural Similarity per Frame (N = {total_ssim.shape[0]})')
        plt.legend() 
        plt.subplot(122)
        plt.hist(total_mse, bins=30, alpha=0.5, color='red', density=True, label='Histogram')
        sns.kdeplot(total_mse, fill=True, color='r', label='KDE')
        plt.axvline(x=median_mse_frames, color='r', linestyle=':', linewidth=1, label=f'Median: {median_mse_frames:.3f}')
        plt.axvline(x=mean_mse_frames, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_mse_frames:.3f}')
        plt.axvspan(mean_mse_frames - std_mse_frames, mean_mse_frames + std_mse_frames, color='red', alpha=0.05, label=f'std: {std_mse_frames:.3f}')
        plt.xlabel('MSE')
        plt.ylabel('Density')
        plt.title(f'Distribution of MSE per Frame (N = {total_mse.shape[0]})')
        plt.legend()
        # Save the all-videos metrics plot if requested
        if save_plots and save_path_prefix:
            all_path = f"{save_path_prefix}losses_all_{model_name}.png"
            plt.savefig(all_path, bbox_inches='tight', dpi=300)
#        plt.show()
        plt.show()

    if performance_dict:
        performance_dict['mean_ssim_D'], performance_dict['median_ssim_D'] = plot_scatter_metrics(all_ssim, 'SSIM', display_plots)
        performance_dict['mean_mse_D'], performance_dict['median_mse_D'] = plot_scatter_metrics(all_mse, 'MSE', display_plots)
        return performance_dict



def plot_saliency_distribution(saliency_values):

    print('### Saliency Mapping to Anatomical Regions ###')
    percentile = 20
    threshold = np.percentile(saliency_values, q=100-percentile)
    print(f'Top {percentile}% saliency values are above {np.round(threshold,3)}')

    # Define the dictionary with indexes as keys and region names as values
    brain_regions = {
        4112: 'ParaHippocampal_R',
        5001: 'Calcarine_L',
        5002: 'Calcarine_R',
        5011: 'Cuneus_L',
        5012: 'Cuneus_R',
        5021: 'Lingual_L',
        5022: 'Lingual_R',
        5101: 'Occipital_Sup_L',
        5102: 'Occipital_Sup_R',
        5201: 'Occipital_Mid_L',
        5202: 'Occipital_Mid_R',
        5301: 'Occipital_Inf_L',
        5302: 'Occipital_Inf_R',
        5401: 'Fusiform_L',
        5402: 'Fusiform_R',
        6101: 'Parietal_Sup_L',
        6102: 'Parietal_Sup_R',
        6201: 'Parietal_Inf_L',
        6222: 'Angular_R',
        6301: 'Precuneus_L',
        6302: 'Precuneus_R',
        8201: 'Temporal_Mid_L',
        8202: 'Temporal_Mid_R',
        8301: 'Temporal_Inf_L',
        8302: 'Temporal_Inf_R',
        9001: 'Cerebelum_Crus1_L',
        9002: 'Cerebelum_Crus1_R',
        9031: 'Cerebelum_4_5_L',
        9032: 'Cerebelum_4_5_R',
        9041: 'Cerebelum_6_L',
        9042: 'Cerebelum_6_R',
        9130: 'Vermis_6'
    }
    
    main_rois = {'Calcarine': 1,
            'Lingual': 2,
            'Occipital_Sup': 3,
            'Occipital_Mid': 4,
            'Occipital_Inf': 5,
            'Fusiform': 6,
            'Temporal_Mid': 7,
            'Temporal_Inf': 8,
            'Other': 9}

    colors = [
        "black",             # Black
        "mediumseagreen",
        "hotpink",
        "lightskyblue",
        "#0072B2",           # Medium blue
        "navy",
        "darkgreen",
        "red",
        "coral",
        "dimgray"
    ]


    brain_path = '/media/miplab-nas-shadow/Data2/Movies_Emo/Preprocessed_data/sub-S01/ses-1/pp_sub-S01_ses-1_YouAgain.feat/filtered_func_data_res_MNI.nii'
    brain_img = nib.load(brain_path)
    template_img = datasets.load_mni152_template()
    template_img = image.resample_to_img(source_img=template_img,
                                        target_img=brain_img,
                                        interpolation='nearest')
    template_data = np.asanyarray(template_img.dataobj)
    
    atlas_path = './aal_SPM12/aal/atlas/AAL.nii'
    atlas_img = nib.load(atlas_path)
    atlas_data = np.asanyarray(atlas_img.dataobj)
    
    mask3d_path_s = f'mask_schaefer1000_{saliency_values.shape[0]}.npy'
    mask3d_path_a = f'mask_schaefer1000_4609.npy'
    #mask3d_path_a = f'mask_schaefer1000_15364.npy'
    mask_3d_s = np.load(mask3d_path_s, mmap_mode = 'r')
    mask_3d_a = np.load(mask3d_path_a, mmap_mode = 'r')
    masked_anatomical = atlas_data * mask_3d_a
    mask_2d = mask_3d_s.reshape(-1,) 
    indices = np.where(mask_2d == 1)[0]

    # Step 1: Create a zero-filled array with the same shape as the original 3D data
    reconstructed_3d = np.zeros(mask_3d_s.shape, dtype=np.float32)
    # Step 2: Flatten the zero-filled array (this step is optional and for clarity)
    reconstructed_flat = reconstructed_3d.flatten()
    # Step 3: Use the stored indices to place the masked_fMRI values into the flattened array
    reconstructed_flat[indices] = saliency_values
    # Step 4: Reshape back to the original 3D shape
    masked_saliency = reconstructed_flat.reshape(mask_3d_s.shape)

    #np.save('saliency_map', masked_saliency)

    # Step 1: Create a zero-filled array with the same shape as the original 3D data
    reconstructed_3d_n = np.zeros(mask_3d_s.shape, dtype=np.float32)
    # Step 2: Flatten the zero-filled array (this step is optional and for clarity)
    reconstructed_flat_n = reconstructed_3d_n.flatten()
    # Step 3: Use the stored indices to place the masked_fMRI values into the flattened array
    saliency_values_n = normalize(saliency_values)
    reconstructed_flat_n[indices] = saliency_values_n
    # Step 4: Reshape back to the original 3D shape
    masked_saliency_n = reconstructed_flat_n.reshape(mask_3d_s.shape)

    labels, counts = np.unique(masked_anatomical, return_counts=True)

    # Create a dictionary to hold the voxel counts for each category
    voxel_counts = {key: 0 for key in main_rois}
    total_voxels = 0

    best_saliency_counts = {key: 0 for key in main_rois}
    total_best_saliency = 0
    
    # Create a dictionary to hold saliency values for each category
    saliency_dict = {key: [] for key in main_rois}

    # Assign counts to categories and collect saliency values
    for label, count in zip(labels, counts):
        if label == 0:
            continue  # Skip background or unlabeled
        roi_name = brain_regions.get(label, 'Other')

        # Determine the category for each roi_name
        matched = False
        for roi in main_rois.keys():
            if roi in roi_name:  # Check for partial matches
                # Get indices where mask == label
                indices = np.where(masked_anatomical == label)
                # Get saliency values at those indices
                saliency_vals = masked_saliency[indices]
                # Append to saliency_dict[roi]
                saliency_dict[roi].extend(saliency_vals.tolist())
                
                voxel_counts[roi] += count
                matched = True
                masked_anatomical[masked_anatomical == label] = main_rois[roi]

                best_saliency_counts[roi] += (saliency_vals > threshold).sum()
                break

        if not matched:
            indices = np.where(masked_anatomical == label)
            saliency_vals = masked_saliency[indices]
            saliency_dict['Other'].extend(saliency_vals.tolist())

            voxel_counts['Other'] += count
            masked_anatomical[masked_anatomical == label] = main_rois['Other']
            best_saliency_counts['Other'] += (saliency_vals > threshold).sum()

        total_voxels += count
        #print(f"({int(label)}) | {roi_name}: {count}")

    #print(f"Total: {total_voxels} voxels")

    # Convert counts to percentages
    roi_percentages = {k: (v / total_voxels * 100) for k, v in voxel_counts.items()}
    best_saliency_percentages = {k: (v / (saliency_values.shape[0] * percentile) * 10000) for k, v in best_saliency_counts.items()}

    print('Region' + ' '*10, '| Number of voxels    | Number of best saliencies')
    for k in voxel_counts.keys():
        print(k, ' '*(15-len(k)), f'| {voxel_counts[k]}', ' '*(18-len(str(voxel_counts[k]))), f'| {best_saliency_counts[k]}')

    print()
    
    print('Region' + ' '*10, '| Total % of voxels   | % of best saliencies')
    for k in voxel_counts.keys():
        print(k, ' '*(15-len(k)), f'| {np.round(roi_percentages[k],1)}%', ' '*(17-len(str(np.round(roi_percentages[k],1)))), f'| {np.round(best_saliency_percentages[k],1)}%')

    df = pd.DataFrame([
        {'Region': region, 'Value': value}
        for region, values in saliency_dict.items()
        for value in values
    ])
    
    categories = df['Region'].unique()
    positions = np.arange(len(categories))
    
    # Prepare data for boxplot
    box_data = [df[df['Region'] == category]['Value'] for category in categories]
    
    fig, ax = plt.subplots(figsize=(10, 5))
    
    # Plot the boxplot shifted slightly to the left
    box_positions = positions - 0.1
    bp = ax.boxplot(box_data, positions=box_positions, widths=0.1, patch_artist=True, showfliers=False)
    
    # Customize boxplot appearance
    for box in bp['boxes']:
        box.set(facecolor='none', linewidth=1)
    for whisker in bp['whiskers']:
        whisker.set(color='black', linewidth=1)
    for cap in bp['caps']:
        cap.set(color='black', linewidth=1)
    for median in bp['medians']:
        median.set(color='black', linewidth=1)
    
    # Plot the scatter points shifted slightly to the right
    strip_positions = positions + 0.1
    for idx, category in enumerate(categories):
        y = df[df['Region'] == category]['Value']
        x = np.random.normal(strip_positions[idx], 0.05, size=len(y))  # Add jitter
        color = colors[idx+1]  # Default grey color for other categories
        
        ax.scatter(x, y, alpha=1, s=1, color=color)

    ax.axhline(threshold, color='r', linestyle='--', linewidth=3, label=f'Top {percentile}% above {np.round(threshold,3)}')

    x_labels = [f'{category}\n({roi_percentages[category]:.1f}%)' for category in categories]
    # Set x-ticks and labels
    ax.set_xticks(positions)
    ax.set_xticklabels(x_labels, rotation=40, fontsize=13)
    # Customize plot
    ax.set_title(f'Saliency Distribution by Region', fontsize=15)
    ax.set_ylabel('Absolute Saliency', fontsize=15)
    ax.tick_params(axis='y', labelsize=15)
    ax.grid(color='grey', linestyle=':', linewidth=0.4)
    plt.tight_layout()
    plt.legend()
    plt.show()

    n_labels = len(main_rois.keys())
    # Create a custom colormap with a color for each ROI plus one for 'Other'
    custom_cmap = mcolors.ListedColormap(colors)
    custom_cmap.set_under(color='black', alpha=0)
    
    # Custom normalization function
    class CustomNormalize(mcolors.Normalize):
        def __init__(self, vmin=None, vmax=None, midpoint=0.4, clip=False):
            self.midpoint = midpoint
            super().__init__(vmin, vmax, clip)
        
        def __call__(self, value, clip=None):
            x, y = [self.vmin, self.midpoint, self.vmax], [0, 0.5, 1]
            return np.ma.masked_array(np.interp(value, x, y))

    n_colors = 256
    color_array = plt.cm.hot(np.linspace(0, 1, n_colors))  # Reds colormap
    color_array[0] = (0, 0, 0, 0)  # Set 0 to transparent black

    colors = [(0, 0, 0, 0), (1, 1, 0, 1)]  # Transparent black, Yellow
    custom_cmap_s = mcolors.ListedColormap(colors)

    # Example: Creating a norm to map values (0 maps to transparent, >0 to yellow)
    bounds = [0, 0.1, 1]  # Define boundaries for 0 and non-zero
    norm = mcolors.BoundaryNorm(bounds, custom_cmap_s.N)

    masked_saliency[masked_saliency < threshold] = 0

    color_bar = True
    for i in range(0, 90, 10):

        plt.figure(figsize=(15, 6))
        
        # Sagittal slices
        ax1 = plt.subplot(1, 3, 1)
        ax1.axis('off')
        plt.imshow(rotate(template_data[i, :, :], 90, reshape=True), cmap='gray')
        img1a = plt.imshow(rotate(masked_anatomical[i, :, :], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[i, :, :], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.title(f'Sagittal slice at x={i}')
        
        # Coronal slices
        ax2 = plt.subplot(1, 3, 2)
        ax2.axis('off')
        plt.imshow(rotate(template_data[:, i, :], 90, reshape=True), cmap='gray')
        img2a = plt.imshow(rotate(masked_anatomical[:, i, :], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[:, i, :], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.title(f'Coronal slice at y={i}')
        
        # Axial slices
        ax3 = plt.subplot(1, 3, 3)
        ax3.axis('off')
        plt.imshow(rotate(template_data[:, :, i], 90, reshape=True), cmap='gray')
        img3a = plt.imshow(rotate(masked_anatomical[:, :, i], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[:, :, i], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.title(f'Axial slice at z={i}')
        
        plt.suptitle('Visual Mask', fontsize=16)
        plt.show()

    
    best_slices = [27, 31, 33]
    for i in best_slices:        # for i in range(90):

        plt.figure(figsize=(6, 6))
        plt.axis('off')
        plt.imshow(rotate(template_data[i, :, :], 90, reshape=True), cmap='gray')
        img1a = plt.imshow(rotate(masked_anatomical[i, :, :], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[i, :, :], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.show()
        print(f'Sagittal slice at x={i}')

        plt.figure(figsize=(6, 6))
        plt.axis('off')
        plt.imshow(rotate(template_data[:, i, :], 90, reshape=True), cmap='gray')
        img2a = plt.imshow(rotate(masked_anatomical[:, i, :], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[:, i, :], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.show()
        print(f'Coronal slice at y={i}')

        plt.figure(figsize=(6, 6))
        plt.axis('off')
        plt.imshow(rotate(template_data[:, :, i], 90, reshape=True), cmap='gray')
        img3a = plt.imshow(rotate(masked_anatomical[:, :, i], 90, reshape=True), cmap=custom_cmap, alpha=0.7, vmin=0.1, vmax=n_labels)
        rotated_saliency = rotate(masked_saliency[:, :, i], 90, reshape=True)
        nonzero_coords = np.argwhere(rotated_saliency > 0)  # Get the coordinates of non-zero values
        saliency_values = rotated_saliency[rotated_saliency > 0]  # Get the corresponding saliency values
        plt.scatter(nonzero_coords[:, 1], nonzero_coords[:, 0], c=saliency_values, cmap=custom_cmap_s, norm=norm, marker='x', s=20)
        plt.show()
        print(f'Axial slice at z={i}')

# def plot_saliency_distribution_old(saliency_values):

#     print('### Saliency Mapping to Anatomical Regions ###')

#     # Define the dictionary with indexes as keys and region names as values
#     brain_regions = {
#         4112: 'ParaHippocampal_R',
#         5001: 'Calcarine_L',
#         5002: 'Calcarine_R',
#         5011: 'Cuneus_L',
#         5012: 'Cuneus_R',
#         5021: 'Lingual_L',
#         5022: 'Lingual_R',
#         5101: 'Occipital_Sup_L',
#         5102: 'Occipital_Sup_R',
#         5201: 'Occipital_Mid_L',
#         5202: 'Occipital_Mid_R',
#         5301: 'Occipital_Inf_L',
#         5302: 'Occipital_Inf_R',
#         5401: 'Fusiform_L',
#         5402: 'Fusiform_R',
#         6101: 'Parietal_Sup_L',
#         6102: 'Parietal_Sup_R',
#         6201: 'Parietal_Inf_L',
#         6222: 'Angular_R',
#         6301: 'Precuneus_L',
#         6302: 'Precuneus_R',
#         8201: 'Temporal_Mid_L',
#         8202: 'Temporal_Mid_R',
#         8301: 'Temporal_Inf_L',
#         8302: 'Temporal_Inf_R',
#         9001: 'Cerebelum_Crus1_L',
#         9002: 'Cerebelum_Crus1_R',
#         9031: 'Cerebelum_4_5_L',
#         9032: 'Cerebelum_4_5_R',
#         9041: 'Cerebelum_6_L',
#         9042: 'Cerebelum_6_R',
#         9130: 'Vermis_6'
#     }
    
#     main_rois = {'Calcarine': 1,
#             'Lingual': 2,
#             'Occipital_Sup': 3,
#             'Occipital_Mid': 4,
#             'Occipital_Inf': 5,
#             'Fusiform': 6,
#             'Temporal_Mid': 7,
#             'Temporal_Inf': 8,
#             'Other': 9}

#     colors = [
#         "black",
#         "#ff7f0e",  # Safety orange.
#         "salmon",
#         "#1f77b4",  # Muted blue.
#         "#17becf",  # Blue-teal.
#         "#9467bd",  # Muted purple.
#         "#d62728",  # Brick red.
#         "#2ca02c",  # Cooked asparagus green.
#         "#bcbd22",  # Curry yellow-green.
#         "#8c564b",  # Chestnut brown.
#     ]

#     brain_path = '/media/miplab-nas-shadow/Data2/Movies_Emo/Preprocessed_data/sub-S01/ses-1/pp_sub-S01_ses-1_YouAgain.feat/filtered_func_data_res_MNI.nii'
#     brain_img = nib.load(brain_path)
#     template_img = datasets.load_mni152_template()
#     template_img = image.resample_to_img(source_img=template_img,
#                                         target_img=brain_img,
#                                         interpolation='nearest')
#     template_data = np.asanyarray(template_img.dataobj)
    
#     atlas_path = './aal_SPM12/aal/atlas/AAL.nii'
#     atlas_img = nib.load(atlas_path)
#     atlas_data = np.asanyarray(atlas_img.dataobj)
    
#     mask3d_path = f'mask_schaefer1000_{saliency_values.shape[0]}.npy'
#     mask_3d = np.load(mask3d_path, mmap_mode = 'r')
#     masked_anatomical = atlas_data * mask_3d
#     mask_2d = mask_3d.reshape(-1,) 
#     indices = np.where(mask_2d == 1)[0]

#     # Step 1: Create a zero-filled array with the same shape as the original 3D data
#     reconstructed_3d = np.zeros(mask_3d.shape, dtype=np.float32)
#     # Step 2: Flatten the zero-filled array (this step is optional and for clarity)
#     reconstructed_flat = reconstructed_3d.flatten()
#     # Step 3: Use the stored indices to place the masked_fMRI values into the flattened array
#     reconstructed_flat[indices] = saliency_values
#     # Step 4: Reshape back to the original 3D shape
#     masked_saliency = reconstructed_flat.reshape(mask_3d.shape)

#     #np.save('saliency_map', masked_saliency)

#     # Step 1: Create a zero-filled array with the same shape as the original 3D data
#     reconstructed_3d_n = np.zeros(mask_3d.shape, dtype=np.float32)
#     # Step 2: Flatten the zero-filled array (this step is optional and for clarity)
#     reconstructed_flat_n = reconstructed_3d_n.flatten()
#     # Step 3: Use the stored indices to place the masked_fMRI values into the flattened array
#     reconstructed_flat_n[indices] = normalize(saliency_values)
#     # Step 4: Reshape back to the original 3D shape
#     masked_saliency_n = reconstructed_flat_n.reshape(mask_3d.shape)

#     labels, counts = np.unique(masked_anatomical, return_counts=True)

#     # Create a dictionary to hold the voxel counts for each category
#     voxel_counts = {key: 0 for key in main_rois}
#     total_voxels = 0

#     # Create a dictionary to hold saliency values for each category
#     saliency_dict = {key: [] for key in main_rois}

#     # Assign counts to categories and collect saliency values
#     for label, count in zip(labels, counts):
#         if label == 0:
#             continue  # Skip background or unlabeled
#         roi_name = brain_regions.get(label, 'Other')

#         # Determine the category for each roi_name
#         matched = False
#         for roi in main_rois.keys():
#             if roi in roi_name:  # Check for partial matches
#                 # Get indices where mask == label
#                 indices = np.where(masked_anatomical == label)
#                 # Get saliency values at those indices
#                 saliency_vals = masked_saliency[indices]
#                 # Append to saliency_dict[roi]
#                 saliency_dict[roi].extend(saliency_vals.tolist())
                
#                 voxel_counts[roi] += count
#                 matched = True
#                 masked_anatomical[masked_anatomical == label] = main_rois[roi]
#                 break

#         if not matched:
#             indices = np.where(masked_anatomical == label)
#             saliency_vals = masked_saliency[indices]
#             saliency_dict['Other'].extend(saliency_vals.tolist())

#             voxel_counts['Other'] += count
#             masked_anatomical[masked_anatomical == label] = main_rois['Other']

#         total_voxels += count
#         print(f"({int(label)}) | {roi_name}: {count}")

#     print(f"Total: {total_voxels} voxels")

#     # Convert counts to percentages
#     roi_percentages = {k: (v / total_voxels * 100) for k, v in voxel_counts.items()}

#     df = pd.DataFrame([
#         {'Region': region, 'Value': value}
#         for region, values in saliency_dict.items()
#         for value in values
#     ])
    
#     categories = df['Region'].unique()
#     positions = np.arange(len(categories))
    
#     # Prepare data for boxplot
#     box_data = [df[df['Region'] == category]['Value'] for category in categories]
    
#     fig, ax = plt.subplots(figsize=(10, 5))
    
#     # Plot the boxplot shifted slightly to the left
#     box_positions = positions - 0.1
#     bp = ax.boxplot(box_data, positions=box_positions, widths=0.1, patch_artist=True, showfliers=False)
    
#     # Customize boxplot appearance
#     for box in bp['boxes']:
#         box.set(facecolor='none', linewidth=1)
#     for whisker in bp['whiskers']:
#         whisker.set(color='black', linewidth=1)
#     for cap in bp['caps']:
#         cap.set(color='black', linewidth=1)
#     for median in bp['medians']:
#         median.set(color='black', linewidth=1)
    
#     # Plot the scatter points shifted slightly to the right
#     strip_positions = positions + 0.1
#     for idx, category in enumerate(categories):
#         y = df[df['Region'] == category]['Value']
#         x = np.random.normal(strip_positions[idx], 0.05, size=len(y))  # Add jitter
#         color = colors[idx+1]  # Default grey color for other categories
        
#         ax.scatter(x, y, alpha=1, s=1, color=color)

#     x_labels = [f'{category}\n({roi_percentages[category]:.1f}%)' for category in categories]
#     # Set x-ticks and labels
#     ax.set_xticks(positions)
#     ax.set_xticklabels(x_labels, rotation=40, fontsize=13)
#     # Customize plot
#     ax.set_title(f'Saliency Distribution by Region', fontsize=15)
#     ax.set_ylabel('Absolute Saliency', fontsize=15)
#     ax.tick_params(axis='y', labelsize=15)
#     ax.grid(color='grey', linestyle=':', linewidth=0.4)
#     plt.tight_layout()
#     plt.show()

#     n_labels = len(main_rois.keys())
#     # Create a custom colormap with a color for each ROI plus one for 'Other'
#     custom_cmap = mcolors.ListedColormap(colors)
#     custom_cmap.set_under(color='black', alpha=0)
    
#     # Custom normalization function
#     class CustomNormalize(mcolors.Normalize):
#         def __init__(self, vmin=None, vmax=None, midpoint=0.4, clip=False):
#             self.midpoint = midpoint
#             super().__init__(vmin, vmax, clip)
        
#         def __call__(self, value, clip=None):
#             x, y = [self.vmin, self.midpoint, self.vmax], [0, 0.5, 1]
#             return np.ma.masked_array(np.interp(value, x, y))
    
#     # Define a custom colormap that emphasizes higher values (above 0.4)
#     n_colors = 256
#     color_array = plt.cm.Reds(np.linspace(0, 1, n_colors))  # Reds colormap
#     color_array[0] = (0, 0, 0, 0)  # Set 0 to transparent black
#     custom_cmap_s = mcolors.ListedColormap(color_array)

#     # Use custom normalization for better contrast
#     norm = CustomNormalize(vmin=0, vmax=1, midpoint=0.4)
    
#     color_bar = True
#     for i in range(10, 90, 10):

#         plt.figure(figsize=(15, 5))
        
#         # Sagittal slices
#         ax1 = plt.subplot(1, 3, 1)
#         ax1.axis('off')
#         plt.imshow(rotate(template_data[i, :, :], 90, reshape=True), cmap='gray')
#         img1 = plt.imshow(rotate(masked_anatomical[i, :, :], 90, reshape=True), cmap=custom_cmap, alpha=0.8, vmin=0.1, vmax=n_labels)
#         plt.title(f'Sagittal slice at x={i}')
        
#         # Coronal slices
#         ax2 = plt.subplot(1, 3, 2)
#         ax2.axis('off')
#         plt.imshow(rotate(template_data[:, i, :], 90, reshape=True), cmap='gray')
#         img2 = plt.imshow(rotate(masked_anatomical[:, i, :], 90, reshape=True), cmap=custom_cmap, alpha=0.8, vmin=0.1, vmax=n_labels)
#         plt.title(f'Coronal slice at y={i}')
        
#         # Axial slices
#         ax3 = plt.subplot(1, 3, 3)
#         ax3.axis('off')
#         plt.imshow(rotate(template_data[:, :, i], 90, reshape=True), cmap='gray')
#         img3 = plt.imshow(rotate(masked_anatomical[:, :, i], 90, reshape=True), cmap=custom_cmap, alpha=0.8, vmin=0.1, vmax=n_labels)
#         plt.title(f'Axial slice at z={i}')
        
#         plt.suptitle('Visual Mask', fontsize=16)
#         plt.show()

#         plt.figure(figsize=(15, 5))
        
#         # Sagittal slices
#         ax1 = plt.subplot(1, 3, 1)
#         ax1.axis('off')
#         plt.imshow(rotate(template_data[i, :, :], 90, reshape=True), cmap='gray')
#         img1 = plt.imshow(rotate(masked_saliency_n[i, :, :], 90, reshape=True), cmap=custom_cmap_s, norm=norm)
#         plt.title(f'Sagittal slice at x={i}')
        
#         # Coronal slices
#         ax2 = plt.subplot(1, 3, 2)
#         ax2.axis('off')
#         plt.imshow(rotate(template_data[:, i, :], 90, reshape=True), cmap='gray')
#         img2 = plt.imshow(rotate(masked_saliency_n[:, i, :], 90, reshape=True), cmap=custom_cmap_s, norm=norm)
#         plt.title(f'Coronal slice at y={i}')
        
#         # Axial slices
#         ax3 = plt.subplot(1, 3, 3)
#         ax3.axis('off')
#         plt.imshow(rotate(template_data[:, :, i], 90, reshape=True), cmap='gray')
#         img3 = plt.imshow(rotate(masked_saliency_n[:, :, i], 90, reshape=True), cmap=custom_cmap_s, norm=norm)
#         plt.title(f'Axial slice at z={i}')

#         if color_bar:
#             cbar = plt.colorbar(img3, ax=[ax1, ax2, ax3], orientation='vertical', fraction=0.046, pad=0.04)
#             cbar.set_label('Normalized Saliency')
#             color_bar = False

#         plt.suptitle('Saliency Distribution', fontsize=16)
#         plt.show()

### STATISTICAL TESTS ###

# Define global variables to hold data, assuming they fit comfortably in memory
# and can be accessed by subprocesses in multiprocessing environment.

def one_sample_permutation_test(lab, pred, alpha = 0.05, num_permutations = 999):
    """
    Conducts a one-sample permutation test to evaluate the significance of observed correlations against a null distribution.

    Parameters:
        lab (numpy array): The ground truth labels, expected to be of shape (TR, voxels).
        pred (numpy array): The model predictions, expected to be of the same shape as 'lab'.
        alpha (float): Significance level for determining statistical significance.
        num_permutations (int): Number of permutations to generate the null distribution.

    Returns:
        numpy array: P-values for each voxel, providing insight into the significance of the observed correlations.
    
    This function computes the correlations between predicted and actual data for each voxel and compares these
    correlations against a null distribution generated by randomly shuffling the labels.
    """
    global tot_lab, tot_pred, valid_indices
    tot_lab, tot_pred = lab, pred

    TR, voxels = tot_pred.shape
    
    # Calculate valid indices only once
    valid_indices = [i for i in range(voxels) if np.std(tot_lab[:, i]) != 0 and np.std(tot_pred[:, i]) != 0]
    
    # Compute observed correlations using valid indices
    correlations_voxel = np.array([
        stats.pearsonr(tot_lab[:, i], tot_pred[:, i])[0] 
        for i in valid_indices
    ])
    mean_correlation_voxel = np.mean(correlations_voxel)
    median_correlation_voxel = np.median(correlations_voxel)
    std_correlation_voxel = np.std(correlations_voxel)

    # Using multiprocessing to handle the computationally intensive task
    with Pool() as pool:
        # Create unique seeds based on the process index, could be random or sequential
        seeds = np.random.randint(0, 1000000, size=num_permutations)
        correlations_voxel_perm = np.asarray(pool.map(compute_permutation, seeds)).T
    mean_correlations_perm = np.concatenate(correlations_voxel_perm[:, :10], axis=0) #for visualisation of the null distribution

    plt.figure(figsize=(8, 4))
    plt.hist(correlations_voxel, bins=30, alpha=0.5, color='green', density=True, label='Prediction Distribution')
    sns.kdeplot(correlations_voxel, fill=True, color='g')
    plt.axvline(x=median_correlation_voxel, color='r', linestyle=':', linewidth=1, label=f'Median: {median_correlation_voxel:.3f}')
    plt.axvline(x=mean_correlation_voxel, color='r', linestyle='--', linewidth=1, label=f'Mean: {mean_correlation_voxel:.3f}')
    plt.axvspan(mean_correlation_voxel - std_correlation_voxel, mean_correlation_voxel + std_correlation_voxel, color='red', alpha=0.05, label=f'std: {std_correlation_voxel:.3f}')
    
    # Plotting the histogram for mean_correlations_perm
    plt.hist(mean_correlations_perm, bins=100, alpha=0.5, color='grey', density=True, label=f'Null Distribution\n(Permutation Test)')
    sns.kdeplot(mean_correlations_perm, fill=True, color='grey')

    plt.xlabel('Correlation Coefficient')
    plt.ylabel('Density')
    plt.title(f'Distribution of Correlations per Voxel (N = {voxels})')
    plt.legend()
    plt.show()
    
    # Calculate the p-values for each observed correlation
    p_values = np.sum(correlations_voxel[:, np.newaxis] < correlations_voxel_perm, axis=1)/num_permutations
    # Assuming you have an array of p-values called `p_values` and other variables `num_permutations`, `voxels`
    plt.figure(figsize=(8, 2))
    # Add a very small value to p_values to avoid log(0) issue, assuming no p-value is exactly 0
    adjusted_p_values = np.maximum(p_values, 1e-6)
    plt.hist(adjusted_p_values, bins=1000, alpha=0.2, color='b', cumulative=True, density=False)
    plt.axvline(alpha, color='k', linestyle=':', linewidth=0.6)
    plt.text(alpha, 170, f' alpha = {alpha}', color='k', ha='left')
    plt.axhline(y=np.sum(p_values < alpha), color='k', linestyle=':', linewidth=0.6)
    plt.text(alpha, 170+np.sum(p_values < alpha), f' {100*np.sum(p_values < alpha)/voxels:.1f}%', color='r', ha='left')
    plt.plot(alpha, np.sum(p_values < alpha), 'x', markersize=6, markeredgewidth=2, color='red')
    plt.axvline(alpha/voxels, color='k', linestyle=':', linewidth=0.6)
    plt.text(alpha/voxels, 170, f' alpha = {alpha}/{voxels}', color='k', ha='left')
    plt.axhline(y=np.sum(p_values < alpha/voxels), color='k', linestyle=':', linewidth=0.6)
    plt.text(alpha/voxels, 170+np.sum(p_values < alpha/voxels), f' {100*np.sum(p_values < alpha/voxels)/voxels:.1f}%', color='r', ha='left')
    plt.plot(alpha/voxels, np.sum(p_values < alpha/voxels), 'x', markersize=6, markeredgewidth=2, color='red')
    plt.xscale('log')
    plt.xlabel('p-value (log scale)')
    plt.ylabel('Cumulative count')
    plt.xlim([0.000001, 1.001])
    plt.title(f'Cumulative Distribution of p-values (Permutation Test with {num_permutations} permutations)')
    plt.show()
    return p_values

# Define the compute_permutation function at the top level
def compute_permutation(seed):
    """
    Compute a permutation of Pearson correlation coefficients between shuffled labels and actual predictions.

    Parameters:
        seed (int): A seed number used to ensure reproducibility of the random shuffling process.

    Returns:
        numpy array: An array of Pearson correlation coefficients calculated between shuffled labels
                     and actual predictions for each valid index.

    This function shuffles the labels array using a given seed to ensure the randomness is reproducible, 
    then calculates the Pearson correlation coefficient between these shuffled labels and the original predictions.
    It operates on the global variables `tot_lab`, `tot_pred`, and `valid_indices`, which must be defined outside
    this function. These variables represent the total labels, total predictions, and indices of valid data points, 
    respectively.
    """
    np.random.seed(seed)  # Set a unique seed for each process
    shuffled_labels = np.copy(tot_lab)
    np.random.shuffle(shuffled_labels)  # Shuffle labels
    perm = np.array([
        stats.pearsonr(shuffled_labels[:, i], tot_pred[:, i])[0] 
        for i in valid_indices
    ])
    return perm

def significant_voxels(p_values, alpha = 0.05, voxels = 4330):
    """
    Identify and visualize (spacially) significant voxels from fMRI data based on provided p-values.

    Parameters:
        p_values (numpy array): Array of p-values for each voxel.
        alpha (float): Significance level for the testing, default is 0.05.
        voxels (int): Total number of voxels considered in the analysis, default is 4330.
    """
    # you can choose another fMRI data as template
    brain_path = '/media/miplab-nas2/Data2/Movies_Emo/Preprocessed_data/sub-S01/ses-1/pp_sub-S01_ses-1_YouAgain.feat/filtered_func_data_res_MNI.nii'
    brain_img = nib.load(brain_path)
    brain_data = np.asanyarray(brain_img.dataobj)[..., 50] #only look at 1 TR

    # path to the mask
    mask_path = '/home/chchan/Michael-Nas2/ml-students2023/resources/vis_mask.nii'
    img_mask = nib.load(mask_path)
    mask_2d = np.asanyarray(img_mask.dataobj).reshape(-1,)
    indices = np.where(mask_2d == 1)[0] # 4330 relevant voxels
    
    #zero-filled array with the same shape as the original 3D data, and flatten
    significant_voxels = np.zeros(brain_data.shape, dtype=np.float32).flatten()
    #set to 1 the voxels that have a p-value < alpha/voxels
    significant_voxels[indices] = np.where(p_values > alpha/voxels, 0, 1)
    #reshape back to the original 3D shape
    significant_voxels = significant_voxels.reshape(brain_data.shape)

    #plot some brain slices
    for i in range(0, 90, 5):
        plt.subplot(1,3,1)
        plt.imshow(np.maximum(significant_voxels[i,:,:],normalize(brain_data[i,:,:])), cmap='gray')
        plt.subplot(1,3,2)
        plt.imshow(np.maximum(significant_voxels[:,i,:],normalize(brain_data[:,i,:])), cmap='gray')
        plt.subplot(1,3,3)
        plt.imshow(np.maximum(significant_voxels[:,:,i],normalize(brain_data[:,:,i])), cmap='gray')
        plt.show()


def two_samples_permutation_test(model1, lab1, pred1, model2, lab2, pred2, num_permutations = 999):
    """
    Conducts a two-sample permutation test to compare two sets of correlations between predicted and actual data.

    Parameters:
        model1 (str): Name or identifier for the first model.
        lab1 (numpy array): Ground truth labels for the first model.
        pred1 (numpy array): Predictions from the first model.
        model2 (str): Name or identifier for the second model.
        lab2 (numpy array): Ground truth labels for the second model.
        pred2 (numpy array): Predictions from the second model.
        num_permutations (int): Number of permutations to perform.

    Returns:
        Displays a visual comparison of correlation distributions and reports the p-value from the permutation test.
    
    This function compares the mean difference in correlations between two different models or conditions using a
    permutation test, providing a statistical comparison of their performance.
    """
    all_encoded, all_labels = [], []
    for key in pred1.keys():
        all_encoded.append(pred1[key])
        all_labels.append(lab1[key])
    tot_pred = np.concatenate(all_encoded, axis=0)
    tot_lab = np.concatenate(all_labels, axis=0)
    voxels = tot_pred.shape[1]
    valid_indices = [i for i in range(voxels) if np.std(tot_lab[:, i]) != 0 and np.std(tot_pred[:, i]) != 0]
    sample1 = np.array([
        stats.pearsonr(tot_lab[:, i], tot_pred[:, i])[0] 
        for i in valid_indices
    ])
    mean1= np.mean(sample1)

    all_encoded, all_labels = [], []
    for key in pred2.keys():
        all_encoded.append(pred2[key])
        all_labels.append(lab2[key])
    tot_pred = np.concatenate(all_encoded, axis=0)
    tot_lab = np.concatenate(all_labels, axis=0)
    voxels = tot_pred.shape[1]
    valid_indices = [i for i in range(voxels) if np.std(tot_lab[:, i]) != 0 and np.std(tot_pred[:, i]) != 0]
    sample2 = np.array([
        stats.pearsonr(tot_lab[:, i], tot_pred[:, i])[0] 
        for i in valid_indices
    ])
    mean2= np.mean(sample2)
    
    test_result = perm_test((sample1, sample2), diff_means, permutation_type='independent', alternative='less', n_resamples=num_permutations)

    plt.figure(figsize=(8, 4))
    plt.hist(sample1, bins=30, alpha=0.5, color='salmon', density=True)
    sns.kdeplot(sample1, fill=True, color='red', label=model1)
    plt.axvline(mean1, color='r', linestyle='--', linewidth=1)
    plt.text(mean1, 8.8, f'Mean: {mean1:.3f} ', color='r', ha='right')
    plt.hist(sample2, bins=30, alpha=0.5, color='lightgreen', density=True)
    sns.kdeplot(sample2, fill=True, color='g', label=model2)
    plt.axvline(mean2, color='g', linestyle='--', linewidth=1)
    plt.text(mean2, 8.8, f' Mean: {mean2:.3f}', color='g', ha='left')
    plt.title(f'Two samples permutation test ({num_permutations} permutations)\nMean difference: {test_result.statistic} (p = {test_result.pvalue:.3e})')
    plt.xlabel('Correlation Coefficient')
    plt.ylabel('Density')
    plt.legend()
    plt.show()

# Define the statistic function
def diff_means(x, y):
    return np.mean(x) - np.mean(y)

from openpyxl import load_workbook

def save_data(excel_file, values_dict, save_model_as, mask_size, num_epochs, lr, encoder_weight):
    values_dict['save_model_as'] = save_model_as
    values_dict['mask_size'] = mask_size
    values_dict['num_epochs'] = num_epochs
    values_dict['lr'] = lr
    values_dict['encoder_weight'] = encoder_weight

    # Load the existing workbook
    workbook = load_workbook(excel_file)

    # Select the active worksheet (or specify the sheet name)
    worksheet = workbook.active

    # Retrieve the headers from the first row of the worksheet
    headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]

    # Ensure that all headers are strings (in case they are None or other types)
    headers = [str(header) if header is not None else '' for header in headers]

    # Check if 'save_model_as' is in headers
    if 'save_model_as' not in headers:
        # Handle the case where 'save_model_as' is not a column in the Excel file
        raise ValueError("'save_model_as' column not found in Excel file.")

    # Find the index of 'save_model_as' column (zero-based index)
    save_model_as_index = headers.index('save_model_as')

    # Collect all existing values in 'save_model_as' column
    existing_save_model_as_values = set()
    for row in worksheet.iter_rows(min_row=2, min_col=save_model_as_index+1, max_col=save_model_as_index+1):
        cell_value = row[0].value
        existing_save_model_as_values.add(cell_value)

    # Check if the 'save_model_as' value already exists
    if save_model_as in existing_save_model_as_values:
        print(f"{save_model_as} already exists in {excel_file}. Skipping save.")
        return  # Do not proceed with appending

    # Prepare the new row data, ordering the values according to the headers
    new_row = [values_dict.get(header, '') for header in headers]

    # Append the new row to the worksheet
    worksheet.append(new_row)

    # Save the workbook to persist the changes
    workbook.save(excel_file)

    print(f"{save_model_as} has been saved to {excel_file}.")
    print(f"num_epochs: {num_epochs}, lr: {lr}, encoder_weight: {encoder_weight}")

def print_dict_tree(d, indent=0):
    """
    Recursively prints the structure of a nested dictionary like a tree.

    Args:
    - d (dict): The nested dictionary to print.
    - indent (int): The indentation level (used for recursion).
    """
    for key, value in d.items():
        if isinstance(value, dict):
            print('  ' * indent + f'{key}')
            print_dict_tree(value, indent + 1)
        elif isinstance(value, np.ndarray):
            print('  ' * indent + f'{key} (shape: {value.shape})')
        else:
            print('  ' * indent + f'{key} (type: {type(value)})')